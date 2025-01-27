import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

import warnings

warnings.simplefilter(action="ignore", category=Warning)


# ---------------- База данных приложения - словарь номеров строк базы ОТК по которым формировались отчеты ------------------
# база данных номеров строк располагается в каталоге проекта или приложения
database = "Справка по рекламациям за период_база данных.txt"

# База ОТК начинается со строки с номером 3. В начале года база ОТК пустая, поэтому ДЕФОЛТНЫЙ словарь
# имеет вид {"0": ["3", "08-01-2025"]}, где "0" - ключ словаря (порядковый номер записи в словаре),
# "3" - номер строки базы ОТК на которой закончилось формирование отчета в последний раз,
# "08-01-2025" - дата записи в словарь (по умолчанию - первый рабочий день).
# Далее, по мере заполнения базы ОТК, в словарь будут добавляться записи.
# Например: {"0": ["3", "08-01-2025"], "1": ["170", "22-01-2025"], "2": ["218", "29-01-2025"]}

try:  # если файл (база данных номеров строк) уже существует
    # открываем базу данных, считываем файл txt и сохраняем словарь в переменную
    with open(database, encoding="utf-8-sig") as file:
        dct_indexes: dict = json.load(file)
        print(dct_indexes)
        # находим длину словаря базы данных номеров строк
        len_dct = len(dct_indexes)
        # определяем номер строки базы ОТК, которая записана в словарь последней и дату последней записи
        index_end, date_end = dct_indexes[str(len_dct - 1)]
        index_end = int(index_end)  # переводим в тип int
        print(index_end, date_end)
except:  # если базы данных нет - создаем
    dct_indexes = {"0": ["3", "08-01-2025"]}
    index_end, date_end = dct_indexes["0"]
    index_end = int(index_end)
    print(index_end, date_end)
    with open(database, "w", encoding="utf-8-sig") as file:
        json.dump(dct_indexes, file, ensure_ascii=False, indent=4)


# --------------------- Константы и расположение файлов базы рекламаций ОТК и справок-отчетов ----------------------
year_now = datetime.today().year  # текущий год

# имя файла базы рекламаций ОТК с учетом текущего года
file = "//Server/otk/1 ГАРАНТИЯ на сервере/" + str(year_now) + "-2019_ЖУРНАЛ УЧЁТА.xlsm"
# file = f"{str(year_now)}-2019_ЖУРНАЛ УЧЁТА.xlsm"


# --------------- Получение итогового датафрейма (справки за период) из файла Excel базы рекламаций ОТК -----------------
# Начальная строка с которой будет осуществляться пропуски в аргументе skiprows=
start_index_skip = 3 if index_end == 3 else 2

df = pd.read_excel(
    file,
    sheet_name=str(year_now),
    header=1,
    usecols=[
        "Месяц регистрации",
        "Дата поступления сообщения в ОТК",
        "Период выявления дефекта (отказа)",
        "Наименование изделия",
        "Обозначение изделия",
        "Заводской номер изделия",
        "Дата изготовления изделия",
        "Пробег, наработка",
        "Заявленный дефект изделия",
        "Количество предъявленных изделий",
    ],
    skiprows=range(start_index_skip, index_end),
)
# header=1 - строку 2 таблицы ОТК делаем заголовками столбцов датафрейма (индексы строк начинаются с 0)
# skiprows=range(start_index_skip, index_end) - пропускаем ранее обработанные строки, начиная с 3 строки таблицы ОТК
# и до строки, на которой закончилось формирование отчета в последний раз.

# Номер строки датафрейма (индекс строки) делаем как в базе данных:
# если база ОТК пустая, то +3 (дефолтное значение), а если заполненные строки есть, то + index_end + 1.
df.index += index_end if index_end == 3 else index_end + 1

# изменяем наименование столбцов датафрейма
df.rename(
    columns={
        "Период выявления дефекта (отказа)": "Период выявления",
        "Количество предъявленных изделий": "Количество",
        "Заявленный дефект изделия": "Заявленный дефект",
    },
    inplace=True,
)

# Удаляем строки в которых нет информации
df_c = df.dropna(subset=["Период выявления"])

# В обозначении изделий убираем перенос строк
df_c["Обозначение изделия"] = df_c["Обозначение изделия"].apply(
    lambda x: x.split("\n")[0] if "\n" in x else x
)

# Изменяем тип данных в столбце "Количество"
df_c["Количество"] = df_c["Количество"].astype("int16")

# Заменяем отсутствующие значения в столбце "Заявленный дефект" на значение "неизвестно"
df_c["Заявленный дефект"].fillna("неизвестно", inplace=True)

# Номер последней строки базы ОТК по которой формировался последний отчет
index_end_new = df_c.index[-1]
print("Последняя строка актуальной базы рекламаций ОТК:", index_end_new)

# Формируем результирующий датафрейм - отчет за период ... df_res = df_c.loc[3:].groupby(...)
df_res = (
    df_c.groupby(
        [
            "Период выявления",
            "Наименование изделия",
            "Обозначение изделия",
            "Заявленный дефект",
        ]
    )["Количество"]
    .sum()
    .to_frame()
)

print(df_res)
print(df_res["Количество"].sum())


# -------------------- Обновляем словарь номеров строк базы ОТК по которым формировались отчеты --------------------
# Добавляем в словарь номер последней строки по которой формировался отчет и дату составления отчета
# Дата составления отчета - сегодняшняя дата
date_end_new = datetime.today().strftime("%d-%m-%Y")
dct_indexes[str(len_dct)] = [str(index_end_new), date_end_new]

# перезаписываем словарь в файл txt базы данных номеров строк
with open(database, "w", encoding="utf-8-sig") as file:
    json.dump(dct_indexes, file, ensure_ascii=False, indent=4)


# --------------------- Запись итогового датафрейма (справки за период) в файл TXT ------------------------
# Файл ТХТ в который будет записываться справка
res_file_txt = "//Server/otk/Support_files_не_удалять!!!/Справка по рекламациям за период.txt"
# res_file_txt = "Справка по рекламациям за период.txt"

with open(res_file_txt, "w", encoding="utf-8") as f:
    print(
        f"\n\n\tСправка по количеству рекламаций за период {date_end} - {date_end_new}",
        file=f,
    )
    f.write(df_res.to_string())

print("Справка в файл TXT записана")


# --------------------- Запись итогового датафрейма (справки за период) в файл Excel ------------------------
# Файл Excel в который будет записываться справка
res_file_excel = f"//Server/otk/ПРОТОКОЛЫ совещаний по качеству/{year_now}/Справка по рекламациям за период.xlsx"
# res_file_excel = "Справка по рекламациям за период.xlsx"

# записываем в файл Excel
df_res.to_excel(res_file_excel)

# ----------------- Редактируем стили и выравнивание в файле Excel справки по рекламациям ----------------------

wb = load_workbook(res_file_excel)  # открываем файл Excel
sheet = wb["Sheet1"]  # делаем активным Лист "Sheet1"

# вставляем дополнительный столбец в позицию 0 (для лучшей визуализации)
sheet.insert_cols(0)

# задаем высоту строки 1 (с названиями столбцов)
sheet.row_dimensions[1].height = 15

# задаем ширину столбцов B, C, D, E, F
sheet.column_dimensions["B"].width = 23
sheet.column_dimensions["C"].width = 20
sheet.column_dimensions["D"].width = 20
sheet.column_dimensions["E"].width = 23
sheet.column_dimensions["F"].width = 10

# столбцы таблицы
cols = "B", "C", "D", "E", "F"

# определяем количество строк в таблице (длина итогового датафрейма)
len_table = len(df_res)

# циклом по столбцам таблицы
for i in cols:
    # активируем перенос текста в ячейках B1, C1, D1, E1, F1 (с названиями столбцов) и выравниваем по центру
    sheet[f"{i + str(1)}"].alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center"
    )
    # циклом по строкам таблицы
    for j in range(1, len_table + 2):
        # задаем стиль границы - тонкая линия и цвет черный
        thins = Side(border_style="thin", color="000000")
        # применяем заданный стиль границы к верхней, нижней, левой и правой границе ячеек по циклу
        sheet[f"{i + str(j)}"].border = Border(
            top=thins, bottom=thins, left=thins, right=thins
        )
        # изменяем шрифт в ячейках с жирного на обычный и устанавливаем Times New Roman размером 10
        sheet[f"{i + str(j)}"].font = Font(name="Times New Roman", size=10, bold=False)

for i in ("B", "C", "D", "E"):
    for j in range(2, len_table + 2):
        # выравниваем текст в ячейках "B", "C", "D", "E" по левому краю по верху с переносом текста
        sheet[f"{i + str(j)}"].alignment = Alignment(
            wrap_text=True, horizontal="left", vertical="top"
        )
for j in range(2, len_table + 2):
    # выравниваем текст в ячейке "F" по центру
    sheet[f"F{str(j)}"].alignment = Alignment(horizontal="center", vertical="center")


# Объединяем ячейки после таблицы для внесения текста
sheet.merge_cells(f"B{len_table + 3}:F{len_table + 3}")
# Записываем текст в объединенную ячейку
sheet[f"B{len_table + 3}"] = (
    f"Справка по количеству рекламаций за период {date_end} - {date_end_new}\n"
    f"Строки базы рекламаций ОТК: с {index_end + 1} по {index_end_new}"
)
# Устанавливаем выравнивание по левому краю с переносом текста
sheet[f"B{len_table + 3}"].alignment = Alignment(
    wrap_text=True, horizontal="left", vertical="center"
)
# Изменяем шрифт в ячейке на Times New Roman размером 12
sheet[f"B{len_table + 3}"].font = Font(name="Times New Roman", size=12)
# Задаем высоту строки
sheet.row_dimensions[len_table + 3].height = 30


# сохраняем изменения
wb.save(res_file_excel)
print("Отредактированный файл Excel со справкой записан")
