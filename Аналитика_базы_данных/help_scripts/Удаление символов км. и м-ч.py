# Программа для удаления символов "км." и "м/ч" в графе "Пробег" файла "Анализ дефектов вн 536 ЯМЗ_2022-2023.xlsx"

import openpyxl


filename = r'//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТЧЕТЫ по дефектности БЗА/Анализ дефектов вн 536 ЯМЗ_2022-2023.xlsx'
# //Server/otk/ОТЧЕТНОСТЬ БЗА/ОТЧЕТЫ по дефектности БЗА/Анализ дефектов вн 536 ЯМЗ_2022-2023.xlsx
wb = openpyxl.load_workbook(filename)   # открываем Книгу
ws = wb['Данные_2021-2023']             # открываем Лист в Книге

row_start = 2   # начинаем со строки 2 (строку заголовков не берем)
row_end = 444   # заканчиваем последней строкой
coll = 8        # номер столбца "Пробег"
lst_mch, lst_km = [], []
for i in range(row_start, row_end + 1):
    value_cell = str(ws.cell(i, coll).value).replace(' ', '').rstrip('.')    # убираем точку в конце и пробелы (210 м/ч. -> 210м/ч)
    if value_cell.endswith('м/ч'):               # если строка заканчивается на м/ч
        value_cell = int(value_cell[:-3]) * 9    # срезом убираем м/ч, переводим в int и умножаем на 9
        lst_mch.append(value_cell)
        ws.cell(i, coll).value = value_cell
    elif value_cell.endswith('км'):
        value_cell = int(value_cell[:-2])        # срезом убираем км и переводим в int
        lst_km.append(value_cell)
        ws.cell(i, coll).value = value_cell
    else:
        value_cell = int(value_cell)             # переводим в int
        lst_km.append(value_cell)
        ws.cell(i, coll).value = value_cell

mch = len(lst_mch)
km = len(lst_km)
print(mch)
print(km)
print(mch + km)
print()
print(lst_mch)
print(lst_km)

wb.save(filename)      # сохраняем файл ОТК
wb.close()             # закрываем открытый файл ОТК
