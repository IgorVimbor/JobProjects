import pandas as pd
import openpyxl
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox

class ExcelFileHandler:
    """
    Класс ExcelFileHandler отвечает за загрузку Excel файлов,
    чтение листов и сохранение изменений с сохранением форматирования.
    """
    def __init__(self):
        self.current_file = None  # Текущий открытый файл
        self.current_sheet = None  # Текущий выбранный лист
        self.sheets = []  # Список листов в файле
        self.original_data = None  # Данные из Excel в виде DataFrame

    def load_workbook_sheets(self, file_path):
        """
        Загружает книгу Excel и возвращает список листов.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)
            sheets = workbook.sheetnames
            workbook.close()
            return sheets
        except Exception as e:
            raise e

    def read_sheet(self, file_path, sheet_name):
        """
        Читает указанный лист Excel в pandas DataFrame.
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df
        except Exception as e:
            raise e

    def save_changes(self, file_path, sheet_name, df):
        """
        Сохраняет изменения из DataFrame в указанный лист Excel,
        при этом сохраняет исходное форматирование ячеек.
        """
        try:
            workbook = openpyxl.load_workbook(file_path, keep_vba=True)
            worksheet = workbook[sheet_name]

            # Сохраняем стили ячеек
            cell_formats = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.has_style:
                        cell_formats[(cell.row, cell.column)] = cell._style

            # Очищаем лист
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.value = None

            # Записываем заголовки
            for col_idx, column_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                if (1, col_idx) in cell_formats:
                    cell._style = cell_formats[(1, col_idx)]

            # Записываем данные
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if (row_idx, col_idx) in cell_formats:
                        cell._style = cell_formats[(row_idx, col_idx)]

            workbook.save(file_path)
        except Exception as e:
            raise e
