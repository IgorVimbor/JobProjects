import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

class ExcelGridWindow(tk.Tk):
    def __init__(self, excel_file):
        super().__init__()
        self.title(f"Таблица Excel ({excel_file})")
        self.geometry("800x600")

        self.excel_file = excel_file  # Путь к файлу Excel
        self.sheet_name = None  # Имя текущего листа
        self.df = None  # DataFrame для хранения данных
        self.filtered_df = None  # DataFrame для отображения отфильтрованных данных
        self.entry = None  # Поле ввода для редактирования ячеек
        self.column_types = {}  # Словарь для хранения типов данных столбцов
        self.formatting = {}  # Словарь для хранения форматирования ячеек
        self.sheet_names = []  # Список листов в файле
        self.sort_column = None  # Столбец для сортировки
        self.sort_ascending = True  # Направление сортировки

        # Создаем основной фрейм
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Фрейм для выбора листа и фильтров
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=5)

        ttk.Label(control_frame, text="Выберите лист:").pack(side=tk.LEFT, padx=5)
        self.sheet_var = tk.StringVar()
        self.sheet_menu = ttk.OptionMenu(control_frame, self.sheet_var, None, *self.sheet_names, command=self.on_sheet_change)
        self.sheet_menu.pack(side=tk.LEFT, padx=5)

        # Кнопка для обновления данных
        refresh_button = ttk.Button(control_frame, text="Обновить", command=self.load_excel_data)
        refresh_button.pack(side=tk.LEFT, padx=5)

        # Поля для фильтрации
        ttk.Label(control_frame, text="Фильтр по столбцу:").pack(side=tk.LEFT, padx=5)
        self.filter_column_var = tk.StringVar()
        self.filter_column_menu = ttk.OptionMenu(control_frame, self.filter_column_var, None, command=self.update_filter_menu)
        self.filter_column_menu.pack(side=tk.LEFT, padx=5)

        self.filter_value_var = tk.StringVar()
        self.filter_entry = ttk.Entry(control_frame, textvariable=self.filter_value_var)
        self.filter_entry.pack(side=tk.LEFT, padx=5)
        self.filter_entry.bind("<Return>", lambda e: self.apply_filter())

        # Кнопка для экспорта
        export_button = ttk.Button(control_frame, text="Экспорт", command=self.export_data)
        export_button.pack(side=tk.RIGHT, padx=5)

        # Создаем фрейм для таблицы с прокруткой
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True)

        # Создаем Treeview для отображения таблицы
        self.tree = ttk.Treeview(table_frame, show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Добавляем вертикальную полосу прокрутки
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)

        # Добавляем горизонтальную полосу прокрутки
        hsb = ttk.Scrollbar(main_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(fill=tk.X)
        self.tree.configure(xscrollcommand=hsb.set)

        # Кнопка для сохранения изменений
        save_button = ttk.Button(main_frame, text="Сохранить изменения в Excel", command=self.on_save)
        save_button.pack(pady=10)

        # Привязываем двойной клик для редактирования ячеек
        self.tree.bind("<Double-1>", self.on_double_click)
        # Привязываем клик по заголовку для сортировки
        self.tree.bind("<Button-1>", self.on_header_click)
        # Привязываем правый клик для контекстного меню
        self.tree.bind("<Button-3>", self.on_right_click)

        # Загружаем список листов и данные из Excel
        self.load_sheet_names()
        if self.sheet_names:
            self.sheet_var.set(self.sheet_names[0])
            self.on_sheet_change(self.sheet_names[0])

    def load_sheet_names(self):
        """Загружаем список листов из файла Excel."""
        try:
            file_extension = os.path.splitext(self.excel_file)[1].lower()
            if file_extension in ['.xlsx', '.xlsm']:
                engine = 'openpyxl'
            elif file_extension == '.xls':
                engine = 'xlrd'
            else:
                raise ValueError(f"Неподдерживаемый формат файла: {file_extension}")

            xl = pd.ExcelFile(self.excel_file, engine=engine)
            self.sheet_names = xl.sheet_names
            menu = self.sheet_menu["menu"]
            menu.delete(0, "end")
            for sheet in self.sheet_names:
                menu.add_command(label=sheet, command=lambda s=sheet: self.sheet_var.set(s))

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при чтении листов из файла Excel: {str(e)}")

    def on_sheet_change(self, sheet_name):
        """Обработчик изменения выбранного листа."""
        self.sheet_name = sheet_name
        self.load_excel_data()

    def load_excel_data(self):
        """Загружаем данные из Excel и сохраняем форматирование."""
        try:
            file_extension = os.path.splitext(self.excel_file)[1].lower()
            if file_extension in ['.xlsx', '.xlsm']:
                engine = 'openpyxl'
            elif file_extension == '.xls':
                engine = 'xlrd'
            else:
                raise ValueError(f"Неподдерживаемый формат файла: {file_extension}")

            # Читаем данные из Excel с помощью pandas
            self.df = pd.read_excel(self.excel_file, sheet_name=self.sheet_name, engine=engine)
            self.filtered_df = self.df.copy()  # Изначально отображаем все данные

            # Сохраняем типы данных столбцов
            for col in self.df.columns:
                self.column_types[col] = self.df[col].dtype

            # Обновляем меню фильтров
            self.update_filter_menu()

            # Очищаем существующую таблицу
            self.tree.delete(*self.tree.get_children())
            self.tree["columns"] = list(self.df.columns)

            # Устанавливаем заголовки столбцов и автоматически подстраиваем ширину
            for col in self.df.columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=100, anchor=tk.CENTER)

            # Заполняем таблицу данными
            self.update_table()

            # Сохраняем форматирование ячеек (только для .xlsx и .xlsm)
            if file_extension in ['.xlsx', '.xlsm']:
                self.load_formatting()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при загрузке файла Excel: {str(e)}")

    def update_filter_menu(self):
        """Обновляем меню выбора столбца для фильтрации."""
        menu = self.filter_column_menu["menu"]
        menu.delete(0, "end")
        for col in self.df.columns:
            menu.add_command(label=col, command=lambda c=col: self.filter_column_var.set(c))
        if self.df.columns:
            self.filter_column_var.set(self.df.columns[0])

    def apply_filter(self):
        """Применяем фильтр к данным."""
        filter_column = self.filter_column_var.get()
        filter_value = self.filter_value_var.get().strip().lower()
        if not filter_column or not filter_value:
            self.filtered_df = self.df.copy()
        else:
            self.filtered_df = self.df[self.df[filter_column].astype(str).str.lower().str.contains(filter_value, na=False)]
        self.update_table()

    def on_header_click(self, event):
        """Обработчик клика по заголовку столбца для сортировки."""
        region = self.tree.identify_region(event.x, event.y)
        if region == "heading":
            column = self.tree.identify_column(event.x)
            col_idx = int(column[1:]) - 1
            col_name = self.tree["columns"][col_idx]
            if self.sort_column == col_name:
                self.sort_ascending = not self.sort_ascending
            else:
                self.sort_column = col_name
                self.sort_ascending = True
            self.sort_table()

    def sort_table(self):
        """Сортируем таблицу по выбранному столбцу."""
        if self.sort_column:
            self.filtered_df = self.filtered_df.sort_values(by=self.sort_column, ascending=self.sort_ascending)
            self.update_table()

    def update_table(self):
        """Обновляем отображение таблицы."""
        self.tree.delete(*self.tree.get_children())
        for index, row in self.filtered_df.iterrows():
            self.tree.insert("", tk.END, values=list(row))
        self.auto_size_columns()

    def auto_size_columns(self):
        """Автоматическая настройка ширины столбцов на основе содержимого."""
        for col in self.tree["columns"]:
            max_width = len(str(col)) * 10  # Начальная ширина на основе заголовка
            for item in self.tree.get_children():
                value = self.tree.item(item, "values")[self.tree["columns"].index(col)]
                max_width = max(max_width, len(str(value)) * 10)
            self.tree.column(col, width=max_width)

    def load_formatting(self):
        """Загружаем форматирование ячеек из файла Excel."""
        self.formatting = {}
        book = load_workbook(self.excel_file, data_only=False)
        sheet = book[self.sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=0):
            for col_idx, cell in enumerate(row):
                self.formatting[(row_idx, col_idx)] = {
                    "fill": cell.fill.patternType if cell.fill else None,
                    "fill_color": cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None,
                    "font_bold": cell.font.bold if cell.font else False,
                    "font_italic": cell.font.italic if cell.font else False,
                    "border": cell.border.left.style if cell.border else None,
                    "alignment": cell.alignment.horizontal if cell.alignment else None,
                }

    def on_double_click(self, event):
        """Обработка двойного клика для редактирования ячейки."""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        column = self.tree.identify_column(event.x)
        col_idx = int(column[1:]) - 1  # Преобразуем #1, #2, ... в индексы 0, 1, ...

        # Получаем координаты ячейки
        bbox = self.tree.bbox(item, column)
        if not bbox:
            return

        # Уничтожаем предыдущее поле ввода, если оно существует
        if self.entry:
            self.entry.destroy()

        # Получаем имя столбца
        col_name = self.tree["columns"][col_idx]

        # Создаем поле ввода
        self.entry = tk.Entry(self.tree)
        self.entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
        self.entry.insert(0, self.tree.item(item, "values")[col_idx])
        self.entry.focus_set()

        # Привязываем события для сохранения изменений в ячейке
        self.entry.bind("<Return>", lambda e: self.save_cell(item, col_idx, col_name))
        self.entry.bind("<FocusOut>", lambda e: self.save_cell(item, col_idx, col_name))

    def save_cell(self, item, col_idx, col_name):
        """Сохранение изменений в ячейке с проверкой типа данных и диапазонов."""
        new_value = self.entry.get()
        col_type = self.column_types[col_name]

        # Проверяем, соответствует ли введенное значение типу данных столбца
        try:
            if pd.api.types.is_integer_dtype(col_type):
                new_value = int(new_value)
            elif pd.api.types.is_float_dtype(col_type):
                new_value = float(new_value)
            # Для других типов (например, строки) оставляем как есть
        except ValueError:
            messagebox.showerror("Ошибка", f"Введенное значение '{new_value}' не соответствует типу данных столбца '{col_name}' ({col_type})")
            self.entry.destroy()
            return

        # Проверяем диапазоны значений (пример: возраст от 0 до 150)
        if col_name.lower() == "возраст" and isinstance(new_value, (int, float)):
            if not (0 <= new_value <= 150):
                messagebox.showerror("Ошибка", f"Возраст должен быть в диапазоне от 0 до 150")
                self.entry.destroy()
                return

        # Обновляем значение в таблице
        values = list(self.tree.item(item, "values"))
        values[col_idx] = str(new_value)  # Treeview требует строковые значения
        self.tree.item(item, values=values)
        self.entry.destroy()

    def on_right_click(self, event):
        """Обработчик правого клика для контекстного меню."""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        column = self.tree.identify_column(event.x)
        col_idx = int(column[1:]) - 1

        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Изменить цвет фона", command=lambda: self.change_background_color(item, col_idx))
        menu.post(event.x_root, event.y_root)

    def change_background_color(self, item, col_idx):
        """Изменение цвета фона ячейки."""
        color = colorchooser.askcolor(title="Выберите цвет фона")[1]
        if color:
            row_idx = self.tree.index(item)
            if (row_idx, col_idx) not in self.formatting:
                self.formatting[(row_idx, col_idx)] = {}
            self.formatting[(row_idx, col_idx)].update({
                "fill": "solid",
                "fill_color": f"FF{color[1:].upper()}",  # Преобразуем в формат openpyxl
            })

    def on_save(self):
        """Сохранение изменений в файл Excel с подтверждением."""
        if messagebox.askyesno("Подтверждение", "Вы уверены, что хотите сохранить изменения в файл Excel?"):
            try:
                # Считываем данные из таблицы обратно в DataFrame
                data = []
                for item in self.tree.get_children():
                    values = self.tree.item(item, "values")
                    data.append(values)
                self.df = pd.DataFrame(data, columns=self.df.columns)

                # Пытаемся преобразовать данные в исходные типы
                for col in self.df.columns:
                    try:
                        if pd.api.types.is_integer_dtype(self.column_types[col]):
                            self.df[col] = self.df[col].astype(int)
                        elif pd.api.types.is_float_dtype(self.column_types[col]):
                            self.df[col] = self.df[col].astype(float)
                        # Для других типов (например, строки) оставляем как есть
                    except ValueError:
                        pass  # Если преобразование не удалось, оставляем как есть

                # Определяем, как сохранять файл в зависимости от его формата
                file_extension = os.path.splitext(self.excel_file)[1].lower()
                if file_extension in ['.xlsx', '.xlsm']:
                    # Загружаем существующий файл Excel с помощью openpyxl
                    book = load_workbook(self.excel_file, keep_vba=(file_extension == '.xlsm'))

                    # Удаляем старый лист, если он существует
                    if self.sheet_name in book.sheetnames:
                        del book[self.sheet_name]

                    # Создаем новый лист и записываем данные
                    with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        writer.book = book
                        self.df.to_excel(writer, sheet_name=self.sheet_name, index=False)

                    # Восстанавливаем форматирование ячеек
                    self.restore_formatting(book)

                    # Сохраняем файл с помощью openpyxl, чтобы сохранить макросы
                    book.save(self.excel_file)

                elif file_extension == '.xls':
                    # Для .xls используем движок xlwt (не поддерживает макросы и сложное форматирование)
                    self.df.to_excel(self.excel_file, sheet_name=self.sheet_name, index=False, engine='xlwt')
                else:
                    raise ValueError(f"Неподдерживаемый формат файла для сохранения: {file_extension}")

                messagebox.showinfo("Успех", "Изменения успешно сохранены!")

            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при сохранении файла Excel: {str(e)}")

    def restore_formatting(self, book):
        """Восстанавливаем форматирование ячеек в новом листе."""
        sheet = book[self.sheet_name]
        for (row_idx, col_idx), fmt in self.formatting.items():
            cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)  # +2 для учета заголовков и индекса pandas
            if fmt.get("fill") and fmt.get("fill_color"):
                cell.fill = PatternFill(start_color=fmt["fill_color"][2:], end_color=fmt["fill_color"][2:], fill_type=fmt["fill"])
            if fmt.get("font_bold") or fmt.get("font_italic"):
                cell.font = Font(bold=fmt.get("font_bold"), italic=fmt.get("font_italic"))
            if fmt.get("border"):
                cell.border = Border(left=Side(style=fmt["border"]), right=Side(style=fmt["border"]),
                                     top=Side(style=fmt["border"]), bottom=Side(style=fmt["border"]))
            if fmt.get("alignment"):
                cell.alignment = Alignment(horizontal=fmt["alignment"])

    def export_data(self):
        """Экспорт данных в новый файл."""
        file_path = filedialog.asksaveasfilename(
            title="Сохранить как",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        if not file_path:
            return
        try:
            if file_path.endswith('.csv'):
                self.df.to_csv(file_path, index=False)
            else:
                self.df.to_excel(file_path, index=False, engine='openpyxl')
            messagebox.showinfo("Успех", f"Данные успешно экспортированы в {file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте данных: {str(e)}")

    def select_sheet(excel_file):
        """Функция для выбора листа из файла Excel."""
        try:
            file_extension = os.path.splitext(excel_file)[1].lower()
            if file_extension in ['.xlsx', '.xlsm']:
                engine = 'openpyxl'
            elif file_extension == '.xls':
                engine = 'xlrd'
            else:
                raise ValueError(f"Неподдерживаемый формат файла: {file_extension}")

            xl = pd.ExcelFile(excel_file, engine=engine)
            sheet_names = xl.sheet_names

            dialog = tk.Toplevel()
            dialog.title("Выбор листа")
            dialog.geometry("300x100")
            dialog.transient()  # Делаем окно модальным
            dialog.grab_set()

            tk.Label(dialog, text="Выберите лист для работы:").pack(pady=5)
            sheet_var = tk.StringVar(value=sheet_names[0])
            sheet_menu = tk.OptionMenu(dialog, sheet_var, *sheet_names)
            sheet_menu.pack(pady=5)

            result = [None]  # Для хранения результата

            def on_ok():
                result[0] = sheet_var.get()
                dialog.destroy()

            def on_cancel():
                result[0] = None
                dialog.destroy()

            tk.Button(dialog, text="OK", command=on_ok).pack(side=tk.LEFT, padx=10, pady=10)
            tk.Button(dialog, text="Отмена", command=on_cancel).pack(side=tk.RIGHT, padx=10, pady=10)

            dialog.wait_window()
            return result[0]

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при чтении листов из файла Excel: {str(e)}")
            return None

def main():
    root = tk.Tk()
    root.withdraw()  # Скрываем главное окно, пока не выберем файл

    excel_file_path = filedialog.askopenfilename(
        title="Выберите файл Excel",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )
    if not excel_file_path:
        root.destroy()
        return  # Пользователь отменил выбор

    app = ExcelGridWindow(excel_file_path)
    app.mainloop()

if __name__ == "__main__":
    main()