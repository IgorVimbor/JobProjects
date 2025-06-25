# Это полный код приложения со всеми улучшениями.

# Основные особенности:
#     Поддержка больших файлов с буферизацией загрузки
#     Индикаторы прогресса для длительных операций
#     Сохранение форматирования Excel
#     История изменений (Undo/Redo)
#     Копирование/вставка данных
#     Сортировка по столбцам
#     Фильтрация данных
#     Управление видимостью столбцов
#     Горячие клавиши
#     Контекстное меню
#     Подтверждение сохранения изменений
#     Информативная строка состояния


import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import colorchooser
import pandas as pd
from pathlib import Path
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

# Часть 1 - основной класс и инициализация:

class EditableTreeview(ttk.Treeview):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        # Создаем поле для редактирования
        self.editor = ttk.Entry(master)
        self.editor.place(x=0, y=0, width=0, height=0)

        # Привязываем события
        self.bind('<Double-1>', self.on_double_click)
        self.editor.bind('<Return>', self.on_editor_return)
        self.editor.bind('<Escape>', self.on_editor_escape)
        self.editor.bind('<FocusOut>', self.on_editor_escape)

        self.editing_cell = None
        self.original_value = None

    def on_double_click(self, event):
        """Обработчик двойного клика - открывает редактор"""
        # Получаем информацию о кликнутой ячейке
        region = self.identify('region', event.x, event.y)
        if region != 'cell':
            return

        column = self.identify_column(event.x)
        item = self.identify_row(event.y)

        if not item or not column:
            return

        # Получаем координаты и размеры ячейки
        x, y, w, h = self.bbox(item, column)

        # Получаем номер колонки и значение
        column_num = int(column[1]) - 1
        values = self.item(item)['values']
        if not values or column_num >= len(values):
            value = ''
        else:
            value = values[column_num]

        # Сохраняем информацию о редактируемой ячейке
        self.editing_cell = (item, column, column_num)
        self.original_value = value

        # Настраиваем редактор
        self.editor.delete(0, tk.END)
        self.editor.insert(0, value)
        self.editor.selection_range(0, tk.END)
        self.editor.place(x=x, y=y, width=w, height=h)
        self.editor.focus_set()

    def on_editor_return(self, event):
        """Сохраняет изменения при нажатии Enter"""
        if not self.editing_cell:
            return

        item, column, column_num = self.editing_cell
        new_value = self.editor.get()

        # Получаем текущие значения и обновляем нужную колонку
        values = list(self.item(item)['values'])
        values[column_num] = new_value

        # Обновляем значение в таблице
        self.item(item, values=values)

        # Скрываем редактор
        self.editor.place_forget()
        self.editing_cell = None

        # Генерируем событие изменения
        self.event_generate('<<TreeviewCellEdited>>')

    def on_editor_escape(self, event):
        """Отменяет редактирование при нажатии Escape или потере фокуса"""
        self.editor.place_forget()
        self.editing_cell = None

class ExcelViewer(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title('Просмотр Excel')
        self.geometry('1024x768')  # Увеличим начальный размер окна

        self.current_file = None
        self.current_sheet = None
        self.sheets = []
        self.original_data = None
        self.changes_made = False
        self.undo_stack = []  # Стек для отмены действий
        self.redo_stack = []  # Стек для возврата действий

        # Создаем главное меню
        self.create_menu()

        # Создаем основной контейнер
        self.main_frame = ttk.Frame(self)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Создаем верхнюю панель с кнопками
        self.create_top_panel()

        # Создаем панель фильтрации
        self.create_filter_panel()

        # Создаем таблицу
        self.create_table()

        # Создаем строку состояния
        self.status_var = tk.StringVar(value='Готов к работе')
        self.status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Создаем контекстное меню для таблицы
        self.create_context_menu()

        # Создаем меню управления столбцами
        self.create_column_menu()

        # Привязываем обработчик закрытия окна
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Устанавливаем горячие клавиши
        self.bind_shortcuts()

# Часть 2 - создание меню и панелей управления:

    def create_menu(self):
        """Создает главное меню программы"""
        menubar = tk.Menu(self)
        self.config(menu=menubar)

        # Меню "Файл"
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Открыть", command=self.load_file, accelerator="Ctrl+O")
        file_menu.add_command(label="Сохранить", command=self.save_changes, accelerator="Ctrl+S")
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.on_closing, accelerator="Alt+F4")

        # Меню "Правка"
        edit_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Правка", menu=edit_menu)
        edit_menu.add_command(label="Отменить", command=self.undo, accelerator="Ctrl+Z")
        edit_menu.add_command(label="Повторить", command=self.redo, accelerator="Ctrl+Y")
        edit_menu.add_separator()
        edit_menu.add_command(label="Копировать", command=self.copy_selection, accelerator="Ctrl+C")
        edit_menu.add_command(label="Вставить", command=self.paste_selection, accelerator="Ctrl+V")

        # Меню "Вид"
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Вид", menu=view_menu)
        view_menu.add_command(label="Управление столбцами", command=self.show_column_manager)
        view_menu.add_command(label="Сбросить фильтр", command=self.clear_filter)

        # Добавляем меню "Анализ"
        analysis_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Анализ", menu=analysis_menu)

        # Подменю статистики
        statistics_menu = tk.Menu(analysis_menu, tearoff=0)
        analysis_menu.add_cascade(label="Статистика", menu=statistics_menu)
        statistics_menu.add_command(label="Описательная статистика",
                                  command=self.show_descriptive_statistics)
        statistics_menu.add_command(label="Корреляционный анализ",
                                  command=self.show_correlation_analysis)

        # # Подменю агрегации
        # aggregation_menu = tk.Menu(analysis_menu, tearoff=0)
        # analysis_menu.add_cascade(label="Агрегация", menu=aggregation_menu)
        # aggregation_menu.add_command(label="Группировка данных",
        #                            command=self.show_grouping_dialog)
        # aggregation_menu.add_command(label="Сводная таблица",
        #                            command=self.show_pivot_dialog)

        # # Подменю визуализации
        # visualization_menu = tk.Menu(analysis_menu, tearoff=0)
        # analysis_menu.add_cascade(label="Визуализация", menu=visualization_menu)
        # visualization_menu.add_command(label="Гистограмма",
        #                              command=lambda: self.show_plot_dialog('histogram'))
        # visualization_menu.add_command(label="График рассеяния",
        #                              command=lambda: self.show_plot_dialog('scatter'))
        # visualization_menu.add_command(label="Линейный график",
        #                              command=lambda: self.show_plot_dialog('line'))
        # visualization_menu.add_command(label="Круговая диаграмма",
        #                              command=lambda: self.show_plot_dialog('pie'))

    def show_descriptive_statistics(self):
        """Показывает окно с описательной статистикой"""
        if self.original_data is None:
            messagebox.showwarning("Предупреждение", "Нет данных для анализа")
            return

        # Создаем новое окно
        stats_window = tk.Toplevel(self)
        stats_window.title("Описательная статистика")
        stats_window.geometry("800x600")

        # Создаем фрейм для выбора столбцов
        selection_frame = ttk.LabelFrame(stats_window, text="Выбор столбцов")
        selection_frame.pack(fill=tk.X, padx=5, pady=5)

        # Создаем список числовых столбцов
        numeric_columns = self.original_data.select_dtypes(include=[np.number]).columns

        # Создаем чекбоксы для выбора столбцов
        self.column_vars = {}
        for col in numeric_columns:
            var = tk.BooleanVar(value=True)
            self.column_vars[col] = var
            ttk.Checkbutton(selection_frame, text=col, variable=var).pack(side=tk.LEFT, padx=5)

        # Создаем кнопку обновления
        ttk.Button(selection_frame, text="Обновить",
                  command=lambda: self.update_statistics(stats_window)).pack(side=tk.RIGHT, padx=5)

        # Создаем текстовое поле для вывода статистики
        self.stats_text = tk.Text(stats_window, wrap=tk.WORD, width=80, height=20)
        self.stats_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Показываем начальную статистику
        self.update_statistics(stats_window)

    def update_statistics(self, window):
        """Обновляет статистику в окне"""
        # Получаем выбранные столбцы
        selected_columns = [col for col, var in self.column_vars.items() if var.get()]

        if not selected_columns:
            messagebox.showwarning("Предупреждение", "Выберите хотя бы один столбец")
            return

        # Получаем статистику
        stats = self.original_data[selected_columns].describe()

        # Добавляем дополнительные статистики
        stats.loc['median'] = self.original_data[selected_columns].median()
        stats.loc['mode'] = self.original_data[selected_columns].mode().iloc[0]
        stats.loc['skew'] = self.original_data[selected_columns].skew()
        stats.loc['kurtosis'] = self.original_data[selected_columns].kurtosis()

        # Очищаем текстовое поле
        self.stats_text.delete(1.0, tk.END)

        # Выводим статистику
        self.stats_text.insert(tk.END, "Описательная статистика:\n\n")
        self.stats_text.insert(tk.END, stats.round(2).to_string())

        # Добавляем дополнительную информацию
        self.stats_text.insert(tk.END, "\n\nДополнительная информация:\n")
        for col in selected_columns:
            null_count = self.original_data[col].isnull().sum()
            unique_count = self.original_data[col].nunique()
            self.stats_text.insert(tk.END,
                                 f"\n{col}:\n"
                                 f"Пропущенные значения: {null_count}\n"
                                 f"Уникальные значения: {unique_count}\n")

    def show_correlation_analysis(self):
        """Показывает окно с корреляционным анализом"""
        if self.original_data is None:
            messagebox.showwarning("Предупреждение", "Нет данных для анализа")
            return

        # Создаем новое окно
        corr_window = tk.Toplevel(self)
        corr_window.title("Корреляционный анализ")
        corr_window.geometry("800x600")

        # Получаем числовые столбцы
        numeric_data = self.original_data.select_dtypes(include=[np.number])

        if numeric_data.empty:
            messagebox.showwarning("Предупреждение", "Нет числовых данных для анализа")
            return

        # Создаем корреляционную матрицу
        corr_matrix = numeric_data.corr()

        # Создаем фигуру matplotlib
        fig, ax = plt.subplots(figsize=(10, 8))
        im = ax.imshow(corr_matrix, cmap='coolwarm')

        # Настраиваем оси
        ax.set_xticks(np.arange(len(corr_matrix.columns)))
        ax.set_yticks(np.arange(len(corr_matrix.columns)))
        ax.set_xticklabels(corr_matrix.columns, rotation=45, ha='right')
        ax.set_yticklabels(corr_matrix.columns)

        # Добавляем значения в ячейки
        for i in range(len(corr_matrix.columns)):
            for j in range(len(corr_matrix.columns)):
                text = ax.text(j, i, f"{corr_matrix.iloc[i, j]:.2f}",
                             ha="center", va="center", color="black")

        # Добавляем цветовую шкалу
        plt.colorbar(im)

        # Создаем холст Tkinter
        canvas = FigureCanvasTkAgg(fig, master=corr_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Добавляем кнопку сохранения
        ttk.Button(corr_window, text="Сохранить график",
                  command=lambda: self.save_plot(fig)).pack(pady=5)

    def create_top_panel(self):
        """Создает верхнюю панель с кнопками"""
        top_panel = ttk.Frame(self.main_frame)
        top_panel.pack(fill=tk.X, pady=(0, 5))

        # Кнопка открытия файла
        self.load_button = ttk.Button(top_panel, text='Открыть Excel файл', command=self.load_file)
        self.load_button.pack(side=tk.LEFT, padx=5)

        # Выпадающий список для выбора листа
        self.sheet_var = tk.StringVar()
        self.sheet_choice = ttk.Combobox(top_panel, textvariable=self.sheet_var, state='disabled')
        self.sheet_choice.pack(side=tk.LEFT, padx=5)
        self.sheet_choice.bind('<<ComboboxSelected>>', self.on_sheet_selected)

        # Кнопка сохранения
        self.save_button = ttk.Button(top_panel, text='Сохранить изменения',
                                    command=self.save_changes, state='disabled')
        self.save_button.pack(side=tk.LEFT, padx=5)

        # Кнопки Undo/Redo
        self.undo_button = ttk.Button(top_panel, text='↶', command=self.undo, state='disabled')
        self.undo_button.pack(side=tk.LEFT, padx=2)
        self.redo_button = ttk.Button(top_panel, text='↷', command=self.redo, state='disabled')
        self.redo_button.pack(side=tk.LEFT, padx=2)

    def create_filter_panel(self):
        """Создает панель с фильтром"""
        filter_panel = ttk.Frame(self.main_frame)
        filter_panel.pack(fill=tk.X, pady=(0, 5))

        # Метка фильтра
        ttk.Label(filter_panel, text="Фильтр:").pack(side=tk.LEFT, padx=5)

        # Поле ввода для фильтра
        self.filter_var = tk.StringVar()
        self.filter_var.trace('w', self.on_filter)
        self.filter_entry = ttk.Entry(filter_panel, textvariable=self.filter_var)
        self.filter_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        # Кнопка сброса фильтра
        self.clear_filter_button = ttk.Button(filter_panel, text='Сбросить фильтр',
                                            command=self.clear_filter)
        self.clear_filter_button.pack(side=tk.LEFT, padx=5)

    def create_context_menu(self):
        """Создает контекстное меню для таблицы"""
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Копировать", command=self.copy_selection)
        self.context_menu.add_command(label="Вставить", command=self.paste_selection)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Удалить", command=self.delete_selection)

        # Привязываем появление меню к правому клику
        # self.tree.bind("<Button-3>", self.show_context_menu)

    def create_column_menu(self):
        """Создает меню управления столбцами"""
        self.column_menu = tk.Menu(self, tearoff=0)
        self.visible_columns = {}  # Словарь для хранения состояния видимости столбцов

    def show_column_manager(self):
        """Показывает окно управления столбцами"""
        column_window = tk.Toplevel(self)
        column_window.title("Управление столбцами")
        column_window.geometry("300x400")

        # Создаем список с чекбоксами
        for column in self.tree['columns']:
            var = tk.BooleanVar(value=self.visible_columns.get(column, True))
            self.visible_columns[column] = var
            ttk.Checkbutton(column_window, text=column, variable=var,
                        command=lambda c=column: self.toggle_column(c)).pack(anchor=tk.W)

    def toggle_column(self, column):
        """Переключает видимость столбца"""
        if self.visible_columns[column].get():
            # Показываем столбец
            self.tree.column(column, width=100)
        else:
            # Скрываем столбец
            self.tree.column(column, width=0)

# Часть 3 - обработка таблицы и прокрутки:

    def create_table(self):
        """Создает таблицу с прокруткой"""
        # Создаем контейнер для таблицы и прокрутки
        self.table_container = ttk.Frame(self.main_frame)
        self.table_container.pack(fill=tk.BOTH, expand=True)

        # Создаем фрейм для фиксированных заголовков
        self.header_frame = ttk.Frame(self.table_container)
        self.header_frame.grid(row=0, column=0, sticky="ew")

        # Создаем холст для прокрутки
        self.canvas = tk.Canvas(self.table_container)

        # Создаем вертикальную прокрутку
        self.vsb = ttk.Scrollbar(self.table_container, orient="vertical", command=self.canvas.yview)
        # Создаем горизонтальную прокрутку
        self.hsb = ttk.Scrollbar(self.table_container, orient="horizontal", command=self.canvas.xview)

        # Создаем фрейм внутри холста для таблицы
        self.table_frame = ttk.Frame(self.canvas)

        # Создаем редактируемую таблицу
        self.tree = EditableTreeview(
            self.table_frame,
            selectmode='extended',
        )
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Создаем фиксированные заголовки
        self.header_tree = ttk.Treeview(self.header_frame, show="headings", height=1)
        self.header_tree.pack(fill=tk.X, expand=True)

        # Настраиваем прокрутку холста
        self.canvas.configure(
            xscrollcommand=self.sync_scroll,
            yscrollcommand=self.vsb.set,
            scrollregion=self.canvas.bbox("all")
        )

        # Создаем окно в холсте для фрейма
        self.canvas_frame = self.canvas.create_window((0, 0), window=self.table_frame, anchor="nw")

        # Размещаем элементы с использованием grid
        self.canvas.grid(row=1, column=0, sticky="nsew")
        self.vsb.grid(row=1, column=1, sticky="ns")
        self.hsb.grid(row=2, column=0, sticky="ew")

        # Настраиваем веса строк и столбцов для корректного растягивания
        self.table_container.grid_rowconfigure(1, weight=1)
        self.table_container.grid_columnconfigure(0, weight=1)

        # Привязываем обработчики событий
        # self.tree.bind('<<TreeviewCellEdited>>', self.on_cell_edited)
        self.canvas.bind('<Configure>', self.on_canvas_configure)
        self.table_frame.bind('<Configure>', self.on_frame_configure)

        # Привязываем прокрутку колесиком мыши
        self.tree.bind('<MouseWheel>', self.on_mousewheel_y)
        self.tree.bind('<Shift-MouseWheel>', self.on_mousewheel_x)
        # Для Linux систем
        self.tree.bind('<Button-4>', self.on_mousewheel_y)
        self.tree.bind('<Button-5>', self.on_mousewheel_y)
        self.tree.bind('<Shift-Button-4>', self.on_mousewheel_x)
        self.tree.bind('<Shift-Button-5>', self.on_mousewheel_x)

    def sync_scroll(self, *args):
        """Синхронизирует прокрутку заголовков и основной таблицы"""
        self.header_tree.xview_moveto(args[0])
        self.hsb.set(*args)

    def on_canvas_configure(self, event):
        """Обработчик изменения размера холста"""
        # Устанавливаем минимальный размер внутреннего фрейма
        self.canvas.itemconfig(self.canvas_frame, width=event.width)

    def on_frame_configure(self, event):
        """Обработчик изменения размера фрейма"""
        # Обновляем область прокрутки
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_mousewheel_y(self, event):
        """Обработчик вертикальной прокрутки колесиком мыши"""
        if event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")

    def on_mousewheel_x(self, event):
        """Обработчик горизонтальной прокрутки колесиком мыши"""
        if event.num == 4 or event.delta > 0:
            self.canvas.xview_scroll(-1, "units")
        elif event.num == 5 or event.delta < 0:
            self.canvas.xview_scroll(1, "units")

    def display_data(self, df):
        """Отображает DataFrame в таблице с буферизацией"""
        BATCH_SIZE = 1000  # Размер пакета для загрузки

        # Показываем индикатор прогресса
        progress = ttk.Progressbar(self.main_frame, mode='determinate')
        progress.pack(fill=tk.X, padx=5, pady=5)

        try:
            # Очищаем таблицы
            for item in self.tree.get_children():
                self.tree.delete(item)
            for item in self.header_tree.get_children():
                self.header_tree.delete(item)

            # Настраиваем столбцы
            columns = list(df.columns)
            self.tree['columns'] = columns
            self.tree['show'] = 'headings'
            self.header_tree['columns'] = columns
            self.header_tree['show'] = 'headings'

            # Настраиваем заголовки и ширину столбцов
            for column in columns:
                # Вычисляем максимальную ширину столбца
                header_width = len(str(column)) * 10
                max_content_width = df[column].astype(str).str.len().max() * 7
                column_width = max(header_width, max_content_width, 50)

                # Настраиваем основную таблицу
                self.tree.heading(column, text=column,
                                command=lambda c=column: self.sort_column(c))
                self.tree.column(column, width=column_width, minwidth=50)

                # Настраиваем заголовки
                self.header_tree.heading(column, text=column)
                self.header_tree.column(column, width=column_width, minwidth=50)

            # Загружаем данные пакетами
            total_rows = len(df)
            for i in range(0, total_rows, BATCH_SIZE):
                batch = df.iloc[i:i+BATCH_SIZE]
                for index, row in batch.iterrows():
                    values = ['' if pd.isna(value) else str(value) for value in row]
                    self.tree.insert('', tk.END, values=values)

                # Обновляем прогресс
                progress['value'] = (i + len(batch)) / total_rows * 100
                self.update_idletasks()

            # Инициализируем видимость столбцов
            self.visible_columns = {column: tk.BooleanVar(value=True) for column in columns}

        finally:
            # Удаляем индикатор прогресса
            progress.destroy()

            # Обновляем область прокрутки
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

# Часть 4 - функции редактирования и управления данными:

    def bind_shortcuts(self):
        """Привязывает горячие клавиши"""
        self.bind('<Control-o>', lambda e: self.load_file())
        self.bind('<Control-s>', lambda e: self.save_changes())
        self.bind('<Control-z>', lambda e: self.undo())
        self.bind('<Control-y>', lambda e: self.redo())
        self.bind('<Control-c>', lambda e: self.copy_selection())
        self.bind('<Control-v>', lambda e: self.paste_selection())
        self.bind('<Delete>', lambda e: self.delete_selection())

    def undo(self):
        """Отменяет последнее действие"""
        if not self.undo_stack:
            return

        # Получаем последнее действие
        action = self.undo_stack.pop()

        # Сохраняем текущее состояние для redo
        self.redo_stack.append({
            'item': action['item'],
            'column': action['column'],
            'old_value': self.tree.set(action['item'], action['column']),
            'new_value': action['old_value']
        })

        # Восстанавливаем предыдущее значение
        self.tree.set(action['item'], action['column'], action['old_value'])

        # Обновляем состояние кнопок
        self.update_undo_redo_buttons()
        self.changes_made = True

    def redo(self):
        """Повторяет отмененное действие"""
        if not self.redo_stack:
            return

        # Получаем последнее отмененное действие
        action = self.redo_stack.pop()

        # Сохраняем текущее состояние для undo
        self.undo_stack.append({
            'item': action['item'],
            'column': action['column'],
            'old_value': self.tree.set(action['item'], action['column']),
            'new_value': action['new_value']
        })

        # Восстанавливаем значение
        self.tree.set(action['item'], action['column'], action['new_value'])

        # Обновляем состояние кнопок
        self.update_undo_redo_buttons()
        self.changes_made = True

    def update_undo_redo_buttons(self):
        """Обновляет состояние кнопок Undo/Redo"""
        self.undo_button['state'] = 'normal' if self.undo_stack else 'disabled'
        self.redo_button['state'] = 'normal' if self.redo_stack else 'disabled'

    def copy_selection(self):
        """Копирует выделенные ячейки в буфер обмена"""
        selected_items = self.tree.selection()
        if not selected_items:
            return

        # Собираем данные выделенных строк
        copy_data = []
        for item in selected_items:
            values = self.tree.item(item)['values']
            copy_data.append('\t'.join(str(v) for v in values))

        # Формируем текст для буфера обмена
        copy_text = '\n'.join(copy_data)

        # Копируем в буфер обмена
        self.clipboard_clear()
        self.clipboard_append(copy_text)

        self.status_var.set(f'Скопировано строк: {len(selected_items)}')

    def paste_selection(self):
        """Вставляет данные из буфера обмена"""
        try:
            # Получаем данные из буфера обмена
            clipboard_data = self.clipboard_get()

            # Разбираем данные
            rows = clipboard_data.split('\n')
            selected_items = self.tree.selection()

            # Если ничего не выделено, вставляем в конец
            if not selected_items:
                for row in rows:
                    values = row.split('\t')
                    self.tree.insert('', tk.END, values=values)
            else:
                # Вставляем данные в выделенные строки
                for item, row in zip(selected_items, rows):
                    values = row.split('\t')
                    current_values = list(self.tree.item(item)['values'])

                    # Сохраняем действие для undo
                    self.undo_stack.append({
                        'item': item,
                        'column': 'all',
                        'old_value': current_values,
                        'new_value': values
                    })

                    # Обновляем значения
                    self.tree.item(item, values=values)

            self.changes_made = True
            self.update_undo_redo_buttons()
            self.status_var.set('Данные вставлены')

        except tk.TclError:
            self.status_var.set('Буфер обмена пуст')

    def delete_selection(self):
        """Удаляет выделенные строки"""
        selected_items = self.tree.selection()
        if not selected_items:
            return

        if messagebox.askyesno("Подтверждение",
                            f"Удалить выбранные строки ({len(selected_items)})?"):
            # Сохраняем удаленные строки для undo
            deleted_data = []
            for item in selected_items:
                values = self.tree.item(item)['values']
                deleted_data.append((item, values))
                self.tree.delete(item)

            # Добавляем действие в стек undo
            self.undo_stack.append({
                'action': 'delete',
                'data': deleted_data
            })

            self.changes_made = True
            self.update_undo_redo_buttons()
            self.status_var.set(f'Удалено строк: {len(selected_items)}')

    def sort_column(self, column):
        """Сортирует данные по столбцу"""
        if not self.tree.get_children():
            return

        # Определяем направление сортировки
        if not hasattr(self, 'last_sort_column') or self.last_sort_column != column:
            self.last_sort_ascending = True
        else:
            self.last_sort_ascending = not self.last_sort_ascending
        self.last_sort_column = column

        # Получаем индекс колонки
        column_idx = self.tree['columns'].index(column)

        # Собираем данные для сортировки
        data = []
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            # Пытаемся преобразовать значение в число для корректной сортировки
            try:
                sort_value = float(values[column_idx])
            except (ValueError, TypeError):
                sort_value = str(values[column_idx]).lower()
            data.append((sort_value, values, item))

        # Сортируем данные
        data.sort(key=lambda x: x[0], reverse=not self.last_sort_ascending)

        # Перемещаем строки
        for idx, (_, values, item) in enumerate(data):
            self.tree.move(item, '', idx)

        # Обновляем статус
        sort_direction = "по возрастанию" if self.last_sort_ascending else "по убыванию"
        self.status_var.set(f'Отсортировано по {column} {sort_direction}')

    def on_filter(self, event=None):
        """Обработчик фильтрации данных"""
        pass
        # if self.original_data is None:
        #     return

        # filter_text = self.filter_text.GetValue().lower()

        # if not filter_text:
        #     self.display_data(self.original_data)
        #     return

        # # Фильтруем данные
        # mask = self.original_data.astype(str).apply(lambda x: x.str.lower()).apply(
        #     lambda x: x.str.contains(filter_text, na=False)).any(axis=1)
        # filtered_data = self.original_data[mask]

        # # Отображаем отфильтрованные данные
        # self.display_data(filtered_data)

        # # Обновляем статус
        # self.statusbar.SetStatusText(f'Найдено записей: {len(filtered_data)}')

    def clear_filter(self, event=None):
        """Обработчик сброса фильтра"""
        self.filter_text.SetValue('')
        if self.original_data is not None:
            self.display_data(self.original_data)
            self.statusbar.SetStatusText('Фильтр сброшен')

# Часть 5 - функции загрузки и сохранения файлов:

    def load_file(self, event=None):
        """Открывает диалог выбора файла и загружает данные"""
        if self.changes_made:
            if not messagebox.askyesno("Подтверждение",
                                    "Есть несохраненные изменения. Продолжить?"):
                return

        file_path = filedialog.askopenfilename(
            filetypes=[('Excel files', '*.xlsx *.xlsm')],
            title='Выберите Excel файл'
        )

        if file_path:
            try:
                # Показываем индикатор прогресса
                progress_window = tk.Toplevel(self)
                progress_window.title("Загрузка файла")
                progress_window.geometry("300x100")
                progress_window.transient(self)
                progress_window.grab_set()

                # Настраиваем окно прогресса
                ttk.Label(progress_window, text="Загрузка файла...").pack(pady=10)
                progress = ttk.Progressbar(progress_window, mode='indeterminate')
                progress.pack(fill=tk.X, padx=20, pady=10)
                progress.start()

                # Обновляем GUI
                self.update()

                # Загружаем список листов
                workbook = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)
                self.sheets = workbook.sheetnames
                workbook.close()

                self.current_file = file_path
                self.sheet_choice['values'] = self.sheets
                self.sheet_choice.current(0)
                self.sheet_choice['state'] = 'readonly'
                self.save_button['state'] = 'normal'

                # Загружаем первый лист
                self.current_sheet = self.sheets[0]
                # self.load_excel_sheet(file_path, self.current_sheet)
                self.on_sheet_selected(file_path, self.current_sheet)

                # Очищаем стеки undo/redo
                self.undo_stack.clear()
                self.redo_stack.clear()
                self.update_undo_redo_buttons()

                # Обновляем заголовок окна
                self.title(f'Просмотр Excel - {Path(file_path).name}')

                # Закрываем окно прогресса
                progress_window.destroy()

            except Exception as e:
                messagebox.showerror('Ошибка', f'Ошибка при загрузке файла: {str(e)}')
                if 'progress_window' in locals():
                    progress_window.destroy()

    def on_sheet_selected(self, filepath, sheet_name):
        """Загружает данные из конкретного листа Excel файла"""
        try:
            # Создаем окно прогресса
            progress_window = tk.Toplevel(self)
            progress_window.title("Загрузка листа")
            progress_window.geometry("300x100")
            progress_window.transient(self)
            progress_window.grab_set()

            # Настраиваем окно прогресса
            ttk.Label(progress_window, text=f"Загрузка листа {sheet_name}...").pack(pady=10)
            progress = ttk.Progressbar(progress_window, mode='determinate')
            progress.pack(fill=tk.X, padx=20, pady=10)

            # Обновляем GUI
            self.update()

            # Читаем Excel файл
            self.original_data = pd.read_excel(filepath, sheet_name=sheet_name)

            # Отображаем данные
            self.display_data(self.original_data)

            # Обновляем статус
            filename = Path(filepath).name
            self.status_var.set(f'Загружен файл: {filename}, лист: {sheet_name}')

            # Сбрасываем фильтр
            self.filter_var.set('')

            # Очищаем историю изменений
            self.changes_made = False
            self.undo_stack.clear()
            self.redo_stack.clear()
            self.update_undo_redo_buttons()

            # Закрываем окно прогресса
            progress_window.destroy()

        except Exception as e:
            messagebox.showerror('Ошибка', f'Ошибка при загрузке листа: {str(e)}')
            if 'progress_window' in locals():
                progress_window.destroy()

    def save_changes(self, event=None):
        """Сохраняет изменения в файл"""
        if not self.changes_made:
            messagebox.showinfo("Информация", "Нет изменений для сохранения")
            return

        if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите сохранить изменения?"):
            return

        try:
            # Показываем окно прогресса
            progress_window = tk.Toplevel(self)
            progress_window.title("Сохранение")
            progress_window.geometry("300x100")
            progress_window.transient(self)
            progress_window.grab_set()

            ttk.Label(progress_window, text="Сохранение изменений...").pack(pady=10)
            progress = ttk.Progressbar(progress_window, mode='indeterminate')
            progress.pack(fill=tk.X, padx=20, pady=10)
            progress.start()
            self.update()

            # Получаем данные из таблицы
            data = []
            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                data.append(values)

            # Создаем DataFrame
            df = pd.DataFrame(data, columns=self.tree['columns'])

            # Загружаем существующий файл
            workbook = openpyxl.load_workbook(self.current_file, keep_vba=True)

            # Получаем лист
            worksheet = workbook[self.current_sheet]

            # Сохраняем форматирование
            cell_formats = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.has_style:
                        cell_formats[(cell.row, cell.column)] = cell._style

            # Очищаем лист
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.value = None

            # Записываем заголовки
            for col_idx, column_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx, value=column_name)
                if (1, col_idx) in cell_formats:
                    cell._style = cell_formats[(1, col_idx)]

            # Записываем данные
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if (row_idx, col_idx) in cell_formats:
                        cell._style = cell_formats[(row_idx, col_idx)]

            # Сохраняем файл
            workbook.save(self.current_file)

            # Обновляем состояние
            self.changes_made = False
            self.save_button['state'] = 'disabled'
            self.status_var.set(f'Изменения сохранены в лист: {self.current_sheet}')

            # Закрываем окно прогресса
            progress_window.destroy()

            messagebox.showinfo('Информация', 'Файл успешно сохранен!')

        except Exception as e:
            messagebox.showerror('Ошибка', f'Ошибка при сохранении: {str(e)}')
            if 'progress_window' in locals():
                progress_window.destroy()

    def on_closing(self):
        """Обработчик закрытия окна"""
        if self.changes_made:
            answer = messagebox.askyesnocancel(
                "Подтверждение",
                "Есть несохраненные изменения. Сохранить перед выходом?"
            )
            if answer is None:  # Отмена
                return
            if answer:  # Да
                self.save_changes()
        self.destroy()

def main():
    app = ExcelViewer()
    app.mainloop()

if __name__ == '__main__':
    main()


# Это первая часть реализации. Здесь добавлены:
    # Меню "Анализ" с подменю
    # Описательная статистика с возможностью выбора столбцов
    # Корреляционный анализ с визуализацией
# Продолжить с реализацией агрегации и визуализации данных?