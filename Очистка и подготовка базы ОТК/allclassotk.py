import openpyxl
from openpyxl.styles import PatternFill

filename = '//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК_текущий год.xlsx'
wb = openpyxl.load_workbook(filename)   # открываем Книгу


class Clear_list:
    def clear_otk(self, row_start, row_end, col_start=3, col_end=21):
        """ функция очистки таблицы ОТК по АСП и Запчасти по каждому Листу """
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                # очищаем значения по АСП в таблице ОТК
                self.ws.cell(row=i, column=j).value = None
            # очищаем значения по Запчасти
            self.ws.cell(row=i, column=24).value = None

    def change_value_god(self, a, b):
        """ вспомогательная функция изменения года в формуле ячейки Excel """
        frm = self.ws.cell(a, b).value   # получаем строку с формулой
        self.ws.cell(a, b).value = None  # удаляем старую формулу
        # делаем список из двух частей строки (разбиваем строку по пробелу)
        lst = frm.split(' ')
        # временная переменная (второй элемент списка - подстрока где год)
        st = lst[1]
        # увеличиваем год и собираем подстроку
        lst[1] = str(int(st.split("'")[0]) + 1) + "'" + st.split("'")[1]
        # меняем значение в ячейке с формулой
        self.ws.cell(a, b).value = ' '.join(lst)

    def change_frm_asp(self):
        """ функция изменения года в формулах по расчету месячной отгрузки по ПОТРЕБИТЕЛЯМ АСП """
        num_row = (21, 38, 53, 67, 76, 90, 99,
                   106)  # список строк в которых меняем год в формуле
        for row in num_row:                          # по каждой строке списка строк num_row
            for col in range(3, 22):                 # по столбцам от 3 до 21 (это АСП)
                # измененяем год в формулах
                self.change_value_god(row, col)

    def change_frm_zp(self, row_start, row_end):
        """ функция изменения года в формулах по расчету месячной отгрузки по ВИДАМ ИЗДЕЛИЙ в колонке АСП и ЗАПЧАСТЬ"""
        for i in range(row_start, row_end + 1):
            # измененяем год в формулах в колонке АСП
            self.change_value_god(i, 23)
            # измененяем год в формулах в колонке ЗАПЧАСТЬ
            self.change_value_god(i, 25)

    def start_clear(self):
        # ОЧИСТКА ТАБЛИЦЫ И ИЗМЕНЕНИЕ ГОДА В НАИМЕНОВАНИИ ЛИСТОВ И ТАБЛИЦ
        name_list = wb.sheetnames     # список наименований всех Листов

        # список кортежей строк для очистки
        #          ТКР       ПК        ВН         МН        ГП
        tp_row = [(4, 19), (22, 36), (39, 51), (54, 65), (68, 74),
                  # ЦМФ      штанга       ХХХ         YYY
                  (77, 88), (107, 111), (93, 97), (100, 104)]

        for name in name_list[1:]:    # циклом по списку имен Листов
            self.ws = wb[name]        # открываем Лист в Книге
            for tp in tp_row:
                row_start, row_end = tp
                # очищаем таблицы по изделиям на всех листах
                self.clear_otk(row_start, row_end)

            # изменяем имя таблицы (увеличиваем год на 1)
            self.ws.cell(1, 3).value = ' '.join(
                [str(int(i) + 1) if i.isdigit() else i for i in self.ws.cell(1, 3).value.split()])
            # изменяем год на Листе (увеличиваем год на 1)
            self.ws.title = ' '.join(
                [str(int(i) + 1) if i.isdigit() else i for i in name.split()])

        # ИЗМЕНЕНИЕ ГОДА В ФОРМУЛАХ ЯЧЕЕК ПО РАСЧЕТУ ОТГРУЗКИ ЗА МЕСЯЦ ПО АСП И ЗАПЧАСТИ
        name_list = wb.sheetnames         # список наименований всех Листов

        # циклом по списку имен Листов
        for name in name_list[1: len(name_list) - 1]:
            self.ws = wb[name]            # открываем Лист в Книге
            self.change_frm_asp()         # изменяем формулы по конвейерам АСП
            for tp in tp_row:
                row_start, row_end = tp
                # изменяем формулы по колонке АСП и ЗАПЧАСТЬ по изделиям
                self.change_frm_zp(row_start, row_end)

        wb.save(filename)  # сохраняем файл ОТК
        wb.close()         # закрываем открытый файл ОТК


class Copy_and_clear_gp:
    """ класс подготовки таблиц отгрузки по ГОДАМ по ПОТРЕБИТЕЛЯМ на следующий год на Листе Гарантийный парк """
    ws = wb['Гарантийный парк']   # открываем Лист Гарантийный парк

    def copy_some_row(self, row_start, row_end, k, col_start=2, col_end=13):
        """ Лист Гарантийный парк - функция для копирования отгрузки текущего года в ПРЕДЫДУЩИЙ текущему год """
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                # копируем значения
                self.ws.cell(
                    row=i + k, column=j).value = self.ws.cell(row=i, column=j).value

    def clear_some_row(self, row_start, row_end=None, col_start=2, col_end=13):
        """ Лист Гарантийный парк - функция очистки НОВОГО текущего года"""
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                # удаляем значения
                self.ws.cell(row=i, column=j).value = None

    def copy_one_row(self, row, k, col_start=2, col_end=13):
        """ Лист Гарантийный парк - функция для копирования отгрузки текущего года в ПРЕДЫДУЩИЙ текущему год """
        for j in range(col_start, col_end + 1):
            self.ws.cell(row + k, column=j).value = self.ws.cell(row,
                                                                 column=j).value  # копируем значения

    def clear_one_row(self, row, col_start=2, col_end=13):
        """ Лист Гарантийный парк - функция очистки НОВОГО текущего года"""
        for j in range(col_start, col_end + 1):
            self.ws.cell(row, column=j).value = None      # удаляем значения

    def rename_row(self, row, col=2):
        """ Лист Гарантийный парк - функция изменения года"""
        self.ws.cell(row, col).value = str(
            int(self.ws.cell(row, col).value) + 1)   # увеличиваем значение года на один

    def start_all_gp(self):
        # заливка желтым цветом ячеек по строкам и столбцам таблицы "Расчет гарантийного парка на текущий год"
        # светло-желтый цвет: RGB (255, 255, 155), в шестнадцатеричном значении FFFF9B
        for row in range(2, 142):
            for col in range(28, 40):
                # удаляем значения в строке 32 таблицы
                self.ws.cell(32, col).value = None
                self.ws.cell(row, col).fill = PatternFill(
                    fill_type='solid', fgColor='FFFF9B')  # заливка желтым цветом

        # перечень строк в которых будет изменяться год
        num_st = (3, 14, 23, 33, 41, 48, 56, 60, 62, 65, 73, 75, 82, 86, 88, 91, 95, 97, 100, 107, 112, 118, 124,
                  128, 133, 137, 139, 142, 146, 148, 151, 155, 157, 160, 164, 166, 169, 173, 175, 178, 182, 184,
                  187, 191, 193, 196, 200, 202, 205, 209, 211, 214, 219, 222, 226, 230, 232)
        for n in num_st:
            self.rename_row(n)          # изменяем год в колонках по АСП
        for m in num_st[:3]:
            self.rename_row(m, col=14)  # изменяем год в колонках по ЗАПЧАСТИ

        # БЗА - копируем отгрузку из прошлого года в позапрошлый (год[-1] -> год[-2])
        self.copy_some_row(15, 22, 9, col_end=25)
        # БЗА - копируем отгрузку из текущего года в прошлый (год -> год[-1])
        self.copy_some_row(6, 13, 9, col_end=25)
        # БЗА - очищаем ячейки текущего года
        self.clear_some_row(6, 13, col_end=25)

        # список кортежей строк и коэффициента смещения для копирования отгрузки и очистки в МНОГОСТРОЧНЫХ конвейерах
        #  --------- ММЗ --------    ------ Ростсельмаш -------
        tp_row_some = [(42, 47, 7), (35, 40, 7), (108, 111, 5), (103, 106, 5),
                       # --------- ЯМЗ - ---------    --------- БАЗ - ---------
                       (125, 127, 4), (121, 123, 4), (220, 221, 3), (217, 218, 3)]

        for tp in tp_row_some:                          # циклом по списку кортежей tp_row_some
            row_start, row_end, k = tp                  # кортеж распаковываем в переменные
            # копируем отгрузку в МНОГОСТРОЧНЫХ конвейерах
            self.copy_some_row(row_start, row_end, k)

        # циклом по НЕЧЕТНЫМ кортежам списка tp_row_some
        for tp in tp_row_some[1::2]:
            # срезом по кортежу распаковываем в переменные
            row_start, row_end = tp[:2]
            # очищаем ячейки текущего года в МНОГОСТРОЧНЫХ конвейерах
            self.clear_some_row(row_start, row_end)

        # список кортежей строки и коэффициента для копирования отгрузки в ОДНОСТРОЧНЫХ конвейерах
            #  ---- МАЗ ----    --Гомсельмаш--    -----УРАЛ-----    -----КАМАЗ----    ------БелАЗ-----
        tp_row_one = [(61, 2), (59, 2), (74, 2), (68, 6), (87, 2), (85, 2), (96, 2), (94, 2), (138, 2), (136, 2),
                      # ------МЗКТ------    ------КрАЗ------    --ХТЗ Белгород--    ------БЗКТ------    --Салео Гомель--
                      (147, 2), (145, 2), (156, 2), (154, 2), (165, 2), (163, 2), (174, 2), (172, 2), (183, 2), (181, 2),
                      # -------ПТЗ------    -------ЧСДМ-----    ------Тула------    ---ХТЗ Харьков---
                      (192, 2), (190, 2), (201, 2), (199, 2), (210, 2), (208, 2), (231, 2), (229, 2)]

        for tp in tp_row_one:           # циклом по НЕЧЕТНЫМ кортежам списка tp_row_one
            row, k = tp                 # кортеж распаковываем в переменные
            # копируем отгрузку в ОДНОСТРОЧНЫХ конвейерах
            self.copy_one_row(row, k)

        for tp in tp_row_one[1::2]:     # циклом по НЕЧЕТНЫМ кортежам списка tp_row_one
            # очищаем ячейки текущего года в ОДНОСТРОЧНЫХ конвейерах
            self.clear_one_row(tp[0])
            # tp[0] - номер строки для удаления значений (первая цифра в нечетных кортежах)

        wb.save(filename)      # сохраняем файл ОТК
        wb.close()             # закрываем открытый файл ОТК


if __name__ == '__main__':
    cl = Clear_list()
    cl.start_clear()
    clgp = Copy_and_clear_gp()
    clgp.start_all_gp()
