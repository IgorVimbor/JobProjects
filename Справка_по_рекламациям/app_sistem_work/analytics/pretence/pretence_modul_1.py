import os
import time
import win32com.client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import warnings

import paths_work  # импортируем файл с путями до базы данных, отчетов и др.
from analytics.pretence.pretence_modul_2 import ActsFromJournal  # импортируем класс ActsFromJournal из tmp_2

# Отключаем предупреждения Pandas в консоли
warnings.simplefilter(action="ignore", category=Warning)


class Date_to_act:
    """
    Класс формирует датафрейм с индексом - номер акта и двумя столбцами - дата акта и дата уведомления,
    записывает датафрейм в файл Excel и редактирует стили таблицы в записанном файле
    """
    def __init__(self, client: str, product: str, acts_claim: str, acts_research: str) -> None:
        """
        Параметры:
            client (str): потребитель (например, ЯМЗ - эксплуатация)
            product (str): наименование (-я) изделия (-ий) (например, водяной насос)
            acts_claim (_type_): номер (-а) акта (-ов) рекламаций потребителя
            acts_research (list): номер (-а) акта (-ов) исследования
        """
        self.client = client
        self.product = product.split(", ") if "," in product else [product]
        self.acts_claim = acts_claim.split() if acts_claim else ""
        self.acts_research = acts_research.split()

        # Путь к файлу с базой данных с учетом текущего года
        self.file_database = paths_work.file_database

        # Путь к файлу Журнала претензий по потребителю
        self.file_journal_pretence = f"{paths_work.journal_pretence}_{client.split()[0]}.xlsx"

        # Путь к файлу с таблицей для ПЭО с номерами и датами актов исследования и датами уведомления
        self.file_out = paths_work.peo_otchet

        # Формируем строку типа "ЯМЗ 2025" для работы класса ActsFromJournal из tmp_2
        self.sheet_name = f"{client.split()[0]} {paths_work.year_now}"

        # Получаем результаты работы класса ActsFromJournal из tmp_2
        self.acts_checker = ActsFromJournal(self.sheet_name, self.acts_research)
        _, self.numbers_acts, self.years_list_acts = self.acts_checker.calculate_results()
        # годы поиска по базе - ключи словаря self.years_list_acts, списки номеров актов по годам - значения словаря

        self.result = None  # итоговый датафрейм с номерами и развернутыми датами актов и датами уведомлений
        self.df_journal = None  # итоговый датафрейм с номерами и короткими датами актов и датами уведомлений


    @staticmethod
    def get_num(str_in):
        """Статический метод выделяет номер акта и переводит в числовой тип"""
        _, _, num = str(str_in).strip().split()
        act = int(num) if "-" not in num else int(num.split("-")[0])
        return act


    def get_frame(self):
        """
        Метод считывает данные из базы рекламаций, формирует датафрейм по годам актов исследования и видам изделий,
        и формирует итоговый датафрейм
        :return: Датафрейм с индексом - номер акта и двумя столбцами - дата акта и дата уведомления.
        """
        # Сводный датафрейм по годам, указанным в списке годов актов исследования
        df_all = pd.DataFrame()
        for value in self.years_list_acts.keys():
            # Формируем датафрейм по конкретному году
            df = pd.read_excel(
                self.file_database,
                sheet_name=value,
                usecols=[
                    "Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия",
                    "Дата поступления сообщения в ОТК",
                    "Период выявления дефекта (отказа)",
                    "Наименование изделия",
                    "Номер акта исследования",
                    "Дата акта исследования",
                ],
                header=1,
            )
            # Добавляем датафрейм по конкретному году в сводный датафрейм
            df_all = pd.concat([df_all, df], ignore_index=True)

        # Сводный датафрейм по изделиям, указанным в списке изделий
        df_client = pd.DataFrame()
        for value in self.product:
            # Из сводного датафрейма по годам формируем датафрейм по наименованию потребителя и изделию
            df_temp = df_all[
                (df_all["Период выявления дефекта (отказа)"] == self.client)
                & (df_all["Наименование изделия"].str.strip() == value)
            ]
            # Добавляем датафрейм по конкретному изделию в сводный датафрейм
            df_client = pd.concat([df_client, df_temp], ignore_index=True)

        # Удаляем пустые строки, в которых нет номеров актов
        df_client = df_client.dropna(subset=["Номер акта исследования"])

        # Переводим номер акта в числовой тип
        df_client["Номер акта исследования"] = df_client["Номер акта исследования"].map(self.get_num)

        # Датафрейм с датами уведомления и номерами актов исследования
        df_client = df_client[
            [
                "Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия",
                "Номер акта исследования",
                "Дата акта исследования",
                "Дата поступления сообщения в ОТК",
            ]
        ]
        # Итоговый датафрейм с номерами актов и датами уведомления, сортированный по номеру акта
        # Если номера актов рекламаций введены, то формируем датафрейм по номерам актов рекламаций и исследования
        if self.acts_claim:
            res_df = df_client[
                (df_client["Номер акта исследования"].isin(self.numbers_acts))
                & (df_client["Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия"].isin(self.acts_claim))
                ].sort_values("Номер акта исследования")
        else:  # Иначе, только по номерам актов исследования
            res_df = df_client[df_client["Номер акта исследования"].isin(self.numbers_acts)].sort_values("Номер акта исследования")

        # Изменяем название столбца с номером актов рекламаций
        res_df.rename(columns={"Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия": "Номер акта рекламации"}, inplace=True)

        # Устанавливаем индексом столбец Номер акта рекламации
        res_df.set_index("Номер акта рекламации", inplace=True)

        # Т.к. некоторые акты исследования составлены на много рекламаций, то убираем дублирование - оставляем только одну строку
        # res_df = res_df.drop_duplicates(subset=["Номер акта исследования", "Дата акта исследования"])

        # Изменяем формат вывода даты на 'dd.mm.yyyy'
        for col in ["Дата поступления сообщения в ОТК", "Дата акта исследования"]:
            res_df[col] = pd.to_datetime(res_df[col]).dt.strftime("%d.%m.%Y")

        # Датафрейм содержит повторяющиеся строки с номерами актов, т.к., например, акт с номером 75 будет и в 2025 и 2024 и 2023 году
        # Удаляем лишние строки из итогового датафрейма - фильтруем по номеру акта и его дате (году)

        # Преобразуем даты в год
        years = pd.to_datetime(res_df['Дата акта исследования']).dt.year.astype(str)

        # Получаем индексы (номера актов) как отдельный Series
        act_numbers = res_df["Номер акта исследования"]

        # Создаем маску: для каждого года и номера акта проверяем соответствие по словарю self.years_list_acts
        mask = [act_num in self.years_list_acts.get(year, [])
                for act_num, year in zip(act_numbers, years)]

        # Фильтруем датафрейм
        res_df = res_df[mask]

        return res_df


    def excel_close_write(self, df: pd.DataFrame):
        """
        Метод для записи итогового датафрейма в файл Excel (file_out).
        Если файл открыт, то перед записью файл закрывается и производится запись.
        :param df (pd.DataFrame): итоговый датафрейм с номерами и датами актов, датами сообщений.
        """
        try:
            # Пытаемся переименовать файл по существующему имени, чтобы проверить доступность
            os.rename(self.file_out, self.file_out)
            df.to_excel(self.file_out)  # Записываем в файл
            # print("Файл записан сразу.")
        except PermissionError:
            # Если файл открыт, закрываем его через COM и записываем заново
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks.Open(self.file_out)
            workbook.Close(SaveChanges=False)  # Закрываем без сохранения
            df.to_excel(self.file_out)  # Записываем в файл
            # print("Файл закрыт и записан.")


    def transform_excel(self, df: pd.DataFrame):
        """
        Метод для редактирования файла Excel (file_out) - вставляет столбец перед столбцом с номерами актов,
        изменяет высоту первой строки с названиями столбцов и их ширину, активирует перенос текста в ячейках
        с названиями столбцов, выравнивает текст в ячейках таблицы по центру и рисует границы по всей таблице.
        :param df (pd.DataFrame): итоговый датафрейм с номерами, датами актов и датами сообщений.
        """
        wb = load_workbook(self.file_out)  # Открываем файл Excel
        sheet = wb["Sheet1"]  # Активируем лист "Sheet1"

        # Вставляем столбец в позицию 1 (счет начинается с 1)
        sheet.insert_cols(1)

        # Задаем высоту строки 1 (с названиями столбцов)
        sheet.row_dimensions[1].height = 30

        cols = ("B", "C", "D", "E")
        for col in cols:
            # Задаем ширину столбцов B, C, D, Е
            sheet.column_dimensions[col].width = 18
            # Активируем перенос текста в ячейках B1, C1, D1, Е1 и выравниваем по центру
            sheet[f"{col}1"].alignment = Alignment(wrap_text=True, horizontal="center")

        # Определяем количество строк в таблице (длина итогового датафрейма)
        len_table = len(df)

        # Изменяем шрифт в ячейках с номерами актов с жирного на обычный
        # Начинаем со 2 строки, т.к. строка 1 - это названия столбцов
        for i in range(2, len_table + 2):
            sheet[f"B{i}"].font = Font(bold=False)

        # Рисуем границы в ячейках столбцов C, D, Е (по количеству строк в таблице)
        for col in ("C", "D", "E"):
            for row in range(2, len_table + 2):
                thins = Side(border_style="thin", color="000000")
                sheet[f"{col}{row}"].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                sheet[f"{col}{row}"].alignment = Alignment(horizontal="center")

        # Сохраняем изменения
        wb.save(self.file_out)


    def get_small_frame(self):
        """
        Метод возвращает датафрейм, в котором даты актов исследования и даты уведомления приведены к виду дд.мм.гг
        """
        temp_df = self.result.copy().reset_index()  # делаем копию итогового датафрейма

        # Убираем из года число 20, т.е из "дд.мм.гггг" делаем "дд.мм.гг"
        temp_df["Дата акта исследования"] = temp_df["Дата акта исследования"].apply(lambda x: x.replace(x[6:8], ""))
        # temp_df["Дата поступления сообщения в ОТК"] = temp_df["Дата поступления сообщения в ОТК"].apply(lambda x: x.replace(x[6:8], ""))

        # Добавляем новый столбец с объединенными номером и датой акта исследования
        temp_df["comb_data"] = temp_df["Номер акта исследования"].astype(str) + " от " + temp_df["Дата акта исследования"]

        # Создаем новый датафрейм с нужными столбцами для вывода на экран приложения
        df_journal = pd.DataFrame({
            "Номер акта рекламации": temp_df["Номер акта рекламации"],
            "Номер и дата акта исследования": temp_df["comb_data"],
            "Дата уведомления": temp_df["Дата поступления сообщения в ОТК"]
        })

        del temp_df  # Удаляем временный датафрейм
        return df_journal


    def write_journal(self) -> None:
        """
        Метод записывает в файл Журнала претензий данные в столбцы Номер и дата акта исследования и Дата уведомления
        """
        # Загружаем файл Журнал претензий по потребителю
        wb = load_workbook(self.file_journal_pretence)
        ws = wb[self.sheet_name]  # или wb['имя_листа'] если нужен конкретный лист

        # Определяем последнюю строку
        last_row = ws.max_row

        # Заполняем нужные ячейки
        # Для столбца "Номер и дата акта исследования"
        value = self.df_journal["Номер и дата акта исследования"]
        ws.cell(row=last_row, column=8).value = '\n'.join(map(str, value.to_list()))
        ws.cell(row=last_row, column=8).alignment = Alignment(wrap_text=True)

        # Для столбца "Дата уведомления"
        value = self.df_journal["Дата уведомления"]
        ws.cell(row=last_row, column=9).value = '\n'.join(map(str, value.to_list()))
        ws.cell(row=last_row, column=9).alignment = Alignment(wrap_text=True)

        # Автоматически подгоняем высоту строки
        ws.row_dimensions[last_row].height = None   # Сбрасываем фиксированную высоту

        # Сохраняем изменения в Журнале претензий
        wb.save(self.file_journal_pretence)
        # print("Номера и даты актов исследования и даты уведомлений внесены в Журнал.")

        # Открываем Журнал претензий после записи данных
        # Нормализация пути (устраняет проблемы с косыми чертами)
        normalized_path = os.path.abspath(self.file_journal_pretence)
        os.startfile(normalized_path)
        # os.startfile(self.file_journal_pretence)


    def close_journal_write_value(self):
        """Метод вызывает функцию записи данных в Журнал претензий. Перед записью проверяет открыт Журнал или закрыт.
        Если открыт, то закрывает его записывает в него данные и открывает файл после записи"""
        try:
            # Пытаемся открыть файл Журнала претензий в режиме записи, чтобы проверить доступность
            with open(self.file_journal_pretence, 'a') as f:
                pass
            # Если Журнал претензий (файл Excel) закрыт, то сразу обрабатываем его
            self.write_journal()

        except PermissionError:
            # Если файл открыт, закрываем его через COM и записываем данные
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks.Open(self.file_journal_pretence)
            workbook.Save()  # Сохраняем перед закрытием (можно заменить на workbook.Close(False) для отмены сохранения)
            workbook.Close()  # Закрываем файл

            # Даем небольшую паузу, чтобы Excel успел освободить файл
            # time.sleep(1)
            # Теперь обрабатываем файл
            self.write_journal()
