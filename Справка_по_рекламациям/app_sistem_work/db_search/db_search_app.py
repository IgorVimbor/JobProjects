# Основной модуль приложения <Поиск по базе ОТК>
import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

import db_search.db_search_modul as t
import paths_work  # импортируем файл с путями до базы данных, отчетов и др.


class AppSearch(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__(master)

        # импортируем путь до файла базы рекламаций ОТК с учетом текущего года
        self.file_database = paths_work.file_database

        # импортируем путь до файла отчета по результатам поиска
        self.file_report = paths_work.db_search_report

        self.title('ПОИСК ПО БАЗЕ РЕКЛАМАЦИЙ ОТК')
        # меняем логотип Tkinter (перышко) на логотип БЗА
        self.iconbitmap('IconBZA.ico')

        # self.geometry('650x630')  # размер окна приложения
        width = 650
        heigh = 630
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        self.geometry("%dx%d+%d+%d" % (width, heigh, (screenwidth - width) / 2, (screenheight - heigh) / 3))

        # Делаем окно растягивающимся
        self.columnconfigure(0, weight=1, minsize=250)
        self.rowconfigure([0, 10], weight=1, minsize=100)

        # Пустая строка перед фреймом
        lbl_null_0 = tk.Label(self, text='')
        lbl_null_0.pack()

        # Создается рамка `frm_form` для ярлыка с текстом и полей для ввода информации.
        self.frm_form = tk.Frame(self, relief=tk.SUNKEN, borderwidth=3)
        self.frm_form.pack()

        # ЗАПОЛНЯЕМ ПЕРВЫЙ ФРЕЙМ
        lbl = tk.Label(master=self.frm_form,
                       text='1. Введите ГОД в котором будет осуществляться поиск:   ', font=("Arial Bold", 10))
        self.ent_god = tk.Entry(master=self.frm_form, width=15)

        # Используем менеджер геометрии grid для размещения ярлыков и поля ввода номеров двигателей
        lbl.grid(row=0, column=0, sticky='w')
        self.ent_god.grid(row=0, column=1, sticky='w')

        # Пустая строка между фреймами
        lbl_null = tk.Label(self, text='')
        lbl_null.pack()

        # Создается новая рамка `frm_form_1` для ярлыков с текстом и полей для ввода информации.
        self.frm_form_1 = tk.Frame(self, relief=tk.SUNKEN, borderwidth=3)
        self.frm_form_1.pack()

        # ЗАПОЛНЯЕМ ВТОРОЙ ФРЕЙМ
        lbl_1 = tk.Label(master=self.frm_form_1,
                         text='2. ПОИСК ОСУЩЕСТВЛЯЕТСЯ ПО:', font=("Arial Bold", 10))
        lbl_null_1 = tk.Label(master=self.frm_form_1, text='')  # пустая строка
        lbl_dvig_1 = tk.Label(
            master=self.frm_form_1, text='- НОМЕРУ ДВИГАТЕЛЯ.', font=("Arial Bold", 10))
        lbl_dvig_2 = tk.Label(master=self.frm_form_1, text='ВНИМАНИЕ! Если в номере двигателя есть БУКВЫ, '
                              'то вводите только ЦИФРЫ. Буквы НЕ надо.', font=("Arial Bold", 10))
        lbl_dvig_3 = tk.Label(master=self.frm_form_1, text='Введите через пробел номера ДВИГАТЕЛЕЙ и нажмите НАЧАТЬ ПОИСК.',
                              font=("Arial Bold", 10))
        self.ent_dvig = tk.Entry(master=self.frm_form_1, width=100)

        # Используем менеджер геометрии grid для размещения ярлыков и поля ввода номеров двигателей
        lbl_1.grid(row=0, column=0, sticky='w')
        lbl_null_1.grid(row=1, column=0)
        lbl_dvig_1.grid(row=2, column=0, sticky='w')
        lbl_dvig_2.grid(row=3, column=0, sticky='w')
        lbl_dvig_3.grid(row=4, column=0, sticky='w')
        self.ent_dvig.grid(row=5, column=0)

        lbl_null_2 = tk.Label(master=self.frm_form_1, text='')  # пустая строка
        lbl_act_1 = tk.Label(
            master=self.frm_form_1, text='- ПО НОМЕРУ АКТА РЕКЛАМАЦИИ.', font=("Arial Bold", 10))
        lbl_act_2 = tk.Label(master=self.frm_form_1, text='Введите через пробел номера АКТОВ РЕКЛАМАЦИЙ и нажмите НАЧАТЬ ПОИСК.',
                             font=("Arial Bold", 10))
        self.ent_act = tk.Entry(master=self.frm_form_1, width=100)
        lbl_null_3 = tk.Label(master=self.frm_form_1, text='')  # пустая строка

        # Используем менеджер геометрии grid для размещения ярлыков и поля ввода номеров актов
        lbl_null_2.grid(row=6, column=0)
        lbl_act_1.grid(row=7, column=0, sticky='w')
        lbl_act_2.grid(row=8, column=0, sticky='w')
        self.ent_act.grid(row=9, column=0)
        lbl_null_3.grid(row=10, column=0)

        # Создает новую рамку frm_buttons для размещения кнопок НАЧАТЬ ПОИСК и ОЧИСТИТЬ СТРОКУ
        self.frm_buttons = tk.Frame(self)
        self.frm_buttons.pack(fill=tk.X, ipadx=5, ipady=5)
        self.bnt_1 = tk.Button(master=self.frm_buttons, text='НАЧАТЬ ПОИСК',
                              font=("Arial Bold", 10), command=self.get_itog)
        self.bnt_2 = tk.Button(master=self.frm_buttons, text='ОЧИСТИТЬ СТРОКИ',
                              font=("Arial Bold", 10), command=self.clear_strok)
        self.bnt_2.pack(side=tk.RIGHT, ipadx=10)
        self.bnt_1.pack(side=tk.RIGHT, padx=10, ipadx=10)

        # Пустая строка между фреймами
        lbl_null_4 = tk.Label(self, text='')
        lbl_null_4.pack()

        # Создается новая рамка frm_form_2 для ярлыков с текстом и поля для вывода результата
        self.frm_form_2 = tk.Frame(self, relief=tk.SUNKEN, borderwidth=3)
        self.frm_form_2.pack()

        # ЗАПОЛНЯЕМ ТРЕТИЙ ФРЕЙМ
        lbl_2 = tk.Label(master=self.frm_form_2, text='3. РЕЗУЛЬТАТ ПОИСКА:',
                         font=("Arial Bold", 10))
        self.text_1 = tk.Text(master=self.frm_form_2, width=85, height=7,
                             background='white', font=("Arial Bold", 10))
        lbl_3 = tk.Label(master=self.frm_form_2, text='Для получения отчета с подробной информацией нажмите СДЕЛАТЬ ОТЧЕТ',
                         font=("Arial Bold", 10))
        lbl_4 = tk.Label(master=self.frm_form_2, text='')   # пустая строка

        # Используем менеджер геометрии grid для размещения ярлыков и поля вывода результата
        lbl_2.grid(row=0, column=0, sticky='w')
        self.text_1.grid(row=1, column=0)
        lbl_3.grid(row=2, column=0, sticky='w')
        lbl_4.grid(row=3, column=0)

        # Создает новую рамку frm_buttons_2 для размещения кнопок СДЕЛАТЬ ОТЧЕТ и СБРОСИТЬ РЕЗУЛЬТАТ
        self.frm_buttons_2 = tk.Frame(self)
        self.frm_buttons_2.pack(fill=tk.X, ipadx=5, ipady=5)
        self.bnt_3 = tk.Button(master=self.frm_buttons_2, text='СДЕЛАТЬ ОТЧЕТ',
                              font=("Arial Bold", 10), command=self.otchet_output)
        self.bnt_4 = tk.Button(master=self.frm_buttons_2, text='СБРОСИТЬ РЕЗУЛЬТАТ',
                              font=("Arial Bold", 10), command=self.clear_res)
        self.bnt_4.pack(side=tk.RIGHT, ipadx=10)
        self.bnt_3.pack(side=tk.RIGHT, padx=10, ipadx=10)

        # Добавляем строку "Development by IGOR VASILENOK" внизу слева
        lbl_dev = tk.Label(self, text="Development by IGOR VASILENOK")
        lbl_dev.pack(side=tk.BOTTOM, anchor='w', padx=5, pady=5)


    def get_value(self):
        """функция возвращает номер года, списки двигателей и актов из строк вводимой информации"""
        self.god = self.ent_god.get()     # получение текста из строки ввода года

        if not self.god:
            messagebox.showinfo('ОШИБКА', 'Не введен год поиска', parent=self)

        self.inf_1 = self.ent_dvig.get()  # получение текста из строки ввода номеров двигателей
        self.inf_2 = self.ent_act.get()   # получение текста из строки ввода номеров актов

        self.num1 = self.inf_1.split()    # формируем список двигателей из строки ввода
        self.num2 = self.inf_2.split()    # формируем список актов из строки ввода

        return self.god, self.num1, self.num2


    def get_itog(self):
        """функция подготовки выборки из базы и вывода на экран"""
        self.god, self.num1, self.num2 = self.get_value()

        # открываем файл базы рекламаций ОТК
        self.wb = load_workbook(self.file_database)
        # открываем лист по номеру года
        self.sheet = self.wb[self.god]

        # список двигателей и актов рекламаций из всей базы
        # колонка 20 - номера двигателей
        self.cells_dvigs = tuple(self.sheet.cell(row=i, column=20) for i in range(3, self.sheet.max_row+1))
        # колонка 13 - номера актов рекламаций ПРИОБРЕТАТЕЛЯ изделия
        self.cells_acts = tuple(self.sheet.cell(row=i, column=13) for i in range(3, self.sheet.max_row+1))

        # вспомогательные списки двигателей и актов рекламаций
        self.dvigs = tuple(str(cell.value) for cell in self.cells_dvigs if cell.value)
        self.acts = tuple(cell.value for cell in self.cells_acts)

        self.text_1.delete('1.0', tk.END)  # очищаем поле вывода перед выводом новых результатов

        for n in self.num1:  # перебираем входящий список двигателей и ищем двигатель в списке dvigs
            if n not in self.dvigs:  # если в списке нет - печатаем результат
                # вывод текста
                self.text_1.insert(1.0, f'Двигателя {n} в базе нет\n')

            for i, cell in enumerate(self.cells_dvigs):
                # если в списке cells_dvigs двигатель есть - печатаем номер строки таблицы Excel
                if n in str(cell.value):
                    row = i + 3
                    self.text_1.insert(1.0, f'Двигатель {str(cell.value)}: строка - {row}\n')

        for n in self.num2:  # перебираем входящий список актов и ищем акт во вспомогательном списке acts
            if n not in self.acts:  # если во вспомогательном списке нет - печатаем результат
                # вывод текста
                self.text_1.insert(1.0, f'Акта {n} в базе нет\n')

            for i, cell in enumerate(self.cells_acts):
                # если в списке cells_acts акт есть - печатаем номер строки таблицы Excel
                if n in str(cell.value):
                    row = i + 3
                    # вывод текста
                    self.text_1.insert(1.0, f'Акт {cell.value}: строка - {row}\n')


    def clear_strok(self):  # функция очистки строк ввода номеров двигателей и актов
        if hasattr(self, 'god') and self.god:
            self.ent_god.delete(0, tk.END)
        if hasattr(self, 'inf_1') and self.inf_1:
            self.ent_dvig.delete(0, tk.END)
        if hasattr(self, 'inf_2') and self.inf_2:
            self.ent_act.delete(0, tk.END)


    def clear_res(self):  # функция очистки поля вывода результата
        self.text_1.delete('1.0', tk.END)


    def otchet_output(self):  # функция подготовки выборки из базы и печати отчета
        num1 = self.inf_1.split()
        num2 = self.inf_2.split()

        # записываем отчет в файл txt
        with open(self.file_report, "w", encoding="utf-8") as res_file:
            print('\n' * 2, file=res_file)

            if num1:  # если есть список двигателей из строки ввода
                # печатаем номера двигателей
                print('Номера двигателей:', ', '.join(str(n)
                      for n in num1), file=res_file)
                print(file=res_file)

            for n in num1:  # ищем двигатель во вспомогательном списке dvigs
                if n not in self.dvigs:   # если во вспомогательном списке нет - печатаем результат
                    print('Двигателя', n, 'в базе нет', file=res_file)
                    print(file=res_file)

                for i, cell in enumerate(self.cells_dvigs):
                    # если в списке cells_dvigs двигатель есть - печатаем информацию из ячеек Excel
                    if n in str(cell.value):
                        row = i + 3

                        print('Двигатель', str(cell.value), '| строка -', row, '|', file=res_file)
                        print('-' * 80, file=res_file)

                        print(
                            self.sheet.cell(row, t.ind('Наименование изделия')).value, '|',
                            self.sheet.cell(row, t.ind('Обозначение изделия')).value, '|',
                            'зав.№:', self.sheet.cell(row, t.ind('Заводской номер изделия')).value,
                            self.sheet.cell(row, t.ind('Дата изготовления изделия')).value, '|',
                            file=res_file
                        )

                        print(
                            'Где выявлен дефект:',
                            self.sheet.cell(row, t.ind('Период выявления дефекта')).value, '|', end=' ',
                            file=res_file
                        )

                        if self.sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value:
                            print(
                                'Р/А №:',
                                self.sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value,
                                '|', end=' ', file=res_file
                            )
                        else:
                            print('Р/А №: акта нет', '|', end=' ', file=res_file)

                        print(
                            'Дефект:',
                            self.sheet.cell(row, t.ind('Заявленный дефект изделия')).value,
                            file=res_file
                        )

                        if self.sheet.cell(row, t.ind('Номер акта исследования')).value:
                            print(
                                'Акт БЗА:',
                                self.sheet.cell(row, t.ind('Номер акта исследования')).value, 'от',
                                self.sheet.cell(row, t.ind('Дата акта исследования')).value.strftime('%d.%m.%Y'),
                                file=res_file
                            )
                        else:
                            print('Акт БЗА: акта нет', file=res_file)

                        print(
                            'Решение БЗА:',
                            self.sheet.cell(row, t.ind('Причины возникновения дефектов')).value, end=' ',
                            file=res_file
                        )

                        if self.sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value:
                            print(
                                '| отчетность БЗА -',
                                self.sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value,
                                file=res_file
                            )
                        else:
                            print(
                                '|',
                                self.sheet.cell(row, t.ind('Пояснения к причинам возникновения дефектов')).value,
                                file=res_file
                            )

                        print('-' * 80, file=res_file)
                        print(file=res_file)

            if num2:   # если есть список актов из строки ввода
                print('Номер акта:', ', '.join(n for n in num2),
                      file=res_file)  # печатаем номера актов рекламаций
                print(file=res_file)

            for n in num2:   # ищем акт во вспомогательном списке acts
                if n not in self.acts:  # если во вспомогательном списке нет - печатаем результат
                    print('Акта', n, 'в базе нет', file=res_file)
                    print(file=res_file)

                for i, cell in enumerate(self.cells_acts):
                    # если в списке cells_acts есть акт - печатаем информацию из ячеек Excel
                    if n in str(cell.value):
                        row = i + 3

                        print('Акт', cell.value, '| строка -', row, '|', file=res_file)
                        print('-' * 80, file=res_file)

                        print(
                            self.sheet.cell(row, t.ind('Наименование изделия')).value, '|',
                            self.sheet.cell(row, t.ind('Обозначение изделия')).value, '|',
                            'зав.№:', self.sheet.cell(row, t.ind('Заводской номер изделия')).value,
                            self.sheet.cell(row, t.ind('Дата изготовления изделия')).value, '|',
                            file=res_file
                        )

                        print(
                            'Где выявлен дефект:',
                            self.sheet.cell(row, t.ind('Период выявления дефекта')).value, '|', end=' ',
                            file=res_file
                        )

                        if self.sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value:
                            print(
                                'Р/А №:',
                                self.sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value,
                                '|', end=' ', file=res_file
                            )
                        else:
                            print('Р/А №: акта нет', '|', end=' ', file=res_file)

                        print(
                            'Дефект:',
                            self.sheet.cell(row, t.ind('Заявленный дефект изделия')).value,
                            file=res_file
                        )

                        if self.sheet.cell(row, t.ind('Номер акта исследования')).value:
                            print(
                                'Акт БЗА:',
                                self.sheet.cell(row, t.ind('Номер акта исследования')).value,
                                'от', self.sheet.cell(row, t.ind('Дата акта исследования')).value.strftime('%d.%m.%Y'),
                                file=res_file
                            )
                        else:
                            print('Акт БЗА: акта нет', file=res_file)

                        print(
                            'Решение БЗА:', self.sheet.cell(row, t.ind('Причины возникновения дефектов')).value, end=' ',
                            file=res_file
                        )

                        if self.sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value:
                            print(
                                '| отчетность БЗА -',
                                self.sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value,
                                file=res_file
                            )
                        else:
                            print(
                                '|',
                                self.sheet.cell(row, t.ind('Пояснения к причинам возникновения дефектов')).value,
                                file=res_file
                            )

                        print('-' * 80, file=res_file)
                        print(file=res_file)

        try:
            # Показываем сообщение о создании и сохранении файла и вопросом об его открытии
            answer = messagebox.askyesno(
                "ИНФОРМАЦИЯ",
                f"Отчет успешно создан и сохранен в файл:\n\n{self.file_report}\n\nОткрыть файл?",
                parent=self
            )

            if answer:  # Если пользователь нажал "Да"
                self._open_file(self.file_report)

            # Закрываем окно после успешного завершения
            self.destroy()

        except:
            messagebox.showinfo(
                "СООБЩЕНИЕ",
                "Возникла ОШИБКА в работе приложения!!!\n\n"
                "Возможно, отсутствуют данные для составления отчета.\n"
                "Обратитесь к разработчику приложения!",
                parent=self
            )

            self.destroy()

    def _open_file(self, file_path):
        """Открывает файл с помощью стандартного приложения Windows"""
        try:
            # Нормализация пути (устраняет проблемы с косыми чертами)
            normalized_path = os.path.abspath(file_path)
            os.startfile(normalized_path)
        except Exception as e:
            messagebox.showwarning(
                "Ошибка открытия",
                f"Не удалось открыть файл:\n{str(e)}",
                parent=self
            )
