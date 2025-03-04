import openpyxl
import xlwings
from openpyxl.styles import PatternFill


# открываем таблицу ОСиМ и нужный Лист
file_osim = "//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК/Отгрузка для ОТК_текущий год.xlsx"
# флаг True - считываем только значение ячейки
wb1 = openpyxl.load_workbook(file_osim, data_only=True)
name_list = wb1.sheetnames  # список имен всех Листов таблицы ОСиМ
name_osim = name_list[0]  # имя первого Листа
ws1 = wb1[name_osim]  # открываем первый Лист таблицы ОСиМ по имени

# открываем таблицу ОТК и нужный Лист
file_otk = "//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК_текущий год.xlsx"
wb2 = openpyxl.load_workbook(file_otk)
# используем title, чтобы имя Листа таблицы ОСиМ начиналось с большой буквы и по этому имени открываем Лист в файле ОТК
name_otk = name_osim.title()
ws2 = wb2[name_otk]  # открываем Лист в файле ОТК


class ExcelSheetCopier:
    """
    класс для копирования значений отгрузки из файла ОСиМ в файл ОТК на Лист конкретного месяца
    row_start: строка начала диапазона по изделию (ТКР, ПК, ВН и т.д.) в файле ОСиМ
    row_end: строка окончания диапазона по изделию в файле ОСиМ
    k: поправочный коэффициент - разница номеров строк начала диапазона в таблице ОТК и ОСиМ
       ТКР = 1, ПК = 2, ВН = 3, МН = 4, ГП = 5, ЦМФ = 6, коромысло и штанга = 11
    col_start: столбец начала диапазона по изделию в АСП = 3, т.к совпадает в таблицах ОСиМ и ОТК
    col_end: столбец окончания диапазона по изделию в АСП = 21, т.к совпадает
    """

    def copy_in_otk(self, row_start, row_end, k, col_start=3, col_end=21):
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                # -------------------- копируем значения по АСП ------------------------
                # временная переменная - значение ячеек в таблице ОСиМ
                t_1 = ws1.cell(row=i, column=j).value
                # если значение в ячейке есть, то переводим в int, иначе пустая строка
                cell = int(t_1) if t_1 else ""
                # копируем значения по АСП из таблицы ОСиМ в ОТК
                ws2.cell(row=i + k, column=j).value = cell

            # ---------------------- копируем значения по Запчасти ---------------------
            # временная переменная (22 - номер столбца в ОСиМ)
            t_2 = ws1.cell(row=i, column=22).value
            # проверка значения и перевод в int
            cell = int(t_2) if t_2 else ""
            # копируем значения по Запчасти в ОТК (24 - столбец в ОТК)
            ws2.cell(row=i + k, column=24).value = cell

        wb2.save(file_otk)  # сохраняем файл ОТК
        wb1.close()  # закрываем файл ОСиМ
        wb2.close()  # закрываем файл ОТК


class DataCopierGarant:
    """
    класс для копирования значений отгрузки из Листа конкретного месяца на Лист "Гарантийный парк" файла ОТК
    гарантийный парк по потребителям на Листе "Гарантийный парк" расчитывается формулами в Excel
    """

    def copy_garant(self):
        # ----------------- сохраняем значения суммарной отгрузки по изделиям в списки и временные переменные -----------------
        # модулем xlwings в таблице ОТК фиксируем значения в ячейках, где используются формулы
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(file_otk)
        excel_book.save()
        excel_book.close()
        excel_app.quit()

        wb1 = openpyxl.open(file_otk, read_only=True, data_only=True)
        ws1 = wb1[name_otk]

        # номера строк с суммарной месячной отгрузкой:
        # ТКР = 23, ПК = 43, ВН = 63, МН = 80, ГП = 89, ЦМФ = 103, коромысло = 122, штанга = 123
        num_strok = (23, 43, 63, 80, 89, 103, 122, 123)

        # список значений суммарной отгрузки по изделиям на АСП
        asp = list(ws1.cell(i, 22).value for i in num_strok)
        # список значений отгрузки по изделиям в ЗАПЧАСТЬ
        zap = list(ws1.cell(i, 24).value for i in num_strok)

        # список значений отгрузки по изделиям на ММЗ
        mmz = list(ws1.cell(i, 3).value for i in num_strok)

        maz = ws1.cell(row=43, column=4).value  # отгрузка на МАЗ
        gomsel = ws1.cell(row=43, column=5).value  # ГОМСЕЛЬМАШ
        mzkt = ws1.cell(row=43, column=6).value  # МЗКТ
        belaz = ws1.cell(row=43, column=7).value  # БелАЗ
        saleo = ws1.cell(row=103, column=8).value  # САЛЕО-Гомель
        xxx_rb = ws1.cell(row=105, column=9).value  # ХХХ-РБ

        ural = ws1.cell(row=43, column=10).value  # УРАЛ
        rostsel_1 = ws1.cell(row=43, column=11).value  # РОСТСЕЛЬМАШ (ПК)
        rostsel_2 = ws1.cell(row=23, column=11).value  # РОСТСЕЛЬМАШ (ТКР)
        rostsel_3 = ws1.cell(row=63, column=11).value  # РОСТСЕЛЬМАШ (ВН)
        rostsel_4 = ws1.cell(row=80, column=11).value  # РОСТСЕЛЬМАШ (МН)
        kamaz_1 = ws1.cell(row=43, column=12).value  # КАМАЗ (ПК)
        kamaz_2 = ws1.cell(row=63, column=12).value  # КАМАЗ (ВН)
        kamaz_3 = ws1.cell(row=80, column=12).value  # КАМАЗ (МН)
        ymz_1 = ws1.cell(row=43, column=13).value  # ЯМЗ (ПК)
        ymz_2 = ws1.cell(row=63, column=13).value  # ЯМЗ (ВН)
        ymz_3 = ws1.cell(row=80, column=13).value  # ЯМЗ (МН)
        ptz = ws1.cell(row=43, column=14).value  # ПТЗ С-Петербург
        paz_1 = ws1.cell(row=43, column=15).value  # ПАЗ (ПК)
        paz_2 = ws1.cell(row=63, column=15).value  # ПАЗ (ВН)
        csdm = ws1.cell(row=43, column=16).value  # ЧСДМ
        tula = ws1.cell(row=43, column=17).value  # Тула
        baz = ws1.cell(row=43, column=18).value  # БАЗ
        xx_rf_1 = ws1.cell(row=105, column=19).value  # ХХ-РФ-1
        xx_rf_2 = ws1.cell(row=105, column=20).value  # ХХ-РФ-2
        xx_rf_3 = ws1.cell(row=105, column=21).value  # ХХ-РФ-3

        wb1.close()  # закрываем файл ОТК

        # ----------------- копируем цифры отгрузки по потребителям и изделиям на лист "Гарантийный парк" -----------------
        wb2 = openpyxl.open(file_otk)  # повторно открываем файл ОТК
        ws2 = wb2["Гарантийный парк"]  # делаем активным лист "Гарантийный парк"

        # словарь номеров столбцов АСП на листе "Гарантийный парк" по наименованию месяца
        num_col = {
            "январь": 2,
            "февраль": 3,
            "март": 4,
            "апрель": 5,
            "май": 6,
            "июнь": 7,
            "июль": 8,
            "август": 9,
            "сентябрь": 10,
            "октябрь": 11,
            "ноябрь": 12,
            "декабрь": 13,
        }

        # наименование месяца в который производится копирование данных
        month = name_otk.split()[0].lower()

        # номер колонки в которую производится копирование данных
        col_asp = num_col[month]  # АСП
        col_zap = num_col[month] + 12  # ЗАПЧАСТЬ
        # номера столбцов ЗАПЧАСТИ отличаются от номеров столбцов АСП на +12

        # создаем итераторы для перебора списков значений отгрузки по АСП, ЗАПЧАСТИ и ММЗ
        it_asp = iter(asp)
        it_zap = iter(zap)
        it_mmz_1 = iter(mmz)

        for i in range(6, 14):
            # копируем значения по АСП
            ws2.cell(row=i, column=col_asp).value = next(it_asp)
            # копируем значения по ЗАПЧАСТИ
            ws2.cell(row=i, column=col_zap).value = next(it_zap)

        for i in range(41, 49):
            # копируем значения по ММЗ
            ws2.cell(row=i, column=col_asp).value = next(it_mmz_1)

        # копируем значения по МАЗ
        ws2.cell(row=71, column=col_asp).value = maz
        # копируем значения по ГОМСЕЛЬМАШ
        ws2.cell(row=80, column=col_asp).value = gomsel
        # копируем значения по МЗКТ
        ws2.cell(row=89, column=col_asp).value = mzkt
        # копируем значения по БелАЗ
        ws2.cell(row=98, column=col_asp).value = belaz
        # копируем значения по САЛЕО-Гомель
        ws2.cell(row=107, column=col_asp).value = saleo
        # копируем значения по ХХХ-РБ
        ws2.cell(row=116, column=col_asp).value = xxx_rb

        # копируем значения по УРАЛ
        ws2.cell(row=125, column=col_asp).value = ural
        # копируем значения по РОСТСЕЛЬМАШ (ПК)
        ws2.cell(row=134, column=col_asp).value = rostsel_1
        # копируем значения по РОСТСЕЛЬМАШ (ТКР)
        ws2.cell(row=135, column=col_asp).value = rostsel_2
        # копируем значения по РОСТСЕЛЬМАШ (ВН)
        ws2.cell(row=136, column=col_asp).value = rostsel_3
        # копируем значения по РОСТСЕЛЬМАШ (МН)
        ws2.cell(row=137, column=col_asp).value = rostsel_4
        # копируем значения по КАМАЗ (ПК)
        ws2.cell(row=152, column=col_asp).value = kamaz_1
        # копируем значения по КАМАЗ (ВН)
        ws2.cell(row=153, column=col_asp).value = kamaz_2
        # копируем значения по КАМАЗ (МН)
        ws2.cell(row=154, column=col_asp).value = kamaz_3
        # копируем значения по ЯМЗ (ПК)
        ws2.cell(row=167, column=col_asp).value = ymz_1
        # копируем значения по ЯМЗ (ВН)
        ws2.cell(row=168, column=col_asp).value = ymz_2
        # копируем значения по ЯМЗ (МН)
        ws2.cell(row=169, column=col_asp).value = ymz_3
        # копируем значения по ПТЗ С-Петербург
        ws2.cell(row=182, column=col_asp).value = ptz
        # копируем значения по ПАЗ (ПК)
        ws2.cell(row=191, column=col_asp).value = paz_1
        # копируем значения по ПАЗ (ВН)
        ws2.cell(row=192, column=col_asp).value = paz_2
        # копируем значения по ЧСДМ
        ws2.cell(row=203, column=col_asp).value = csdm
        # копируем значения по Туле
        ws2.cell(row=212, column=col_asp).value = tula
        # копируем значения по БАЗ
        ws2.cell(row=221, column=col_asp).value = baz
        # копируем значения по ХХ-РФ-1
        ws2.cell(row=230, column=col_asp).value = xx_rf_1
        # копируем значения по ХХ-РФ-2
        ws2.cell(row=239, column=col_asp).value = xx_rf_2
        # копируем значения по ХХ-РФ-3
        ws2.cell(row=248, column=col_asp).value = xx_rf_3

        # после копирования данных и расчета гарантийного парка заливаем столбцы белым цветом
        # номера столбцов отличаются от номеров столбцов АСП на +26
        col_color = num_col[month] + 26

        for row in range(2, 155):
            # заливка белым цветом: : RGB (255, 255, 255), в шестнадцатиричном значении FFFFFF
            ws2.cell(row, col_color).fill = PatternFill(
                fill_type="solid", fgColor="FFFFFF"
            )

        wb2.save(file_otk)  # сохраняем изменения
        wb2.close()  # закрываем файл ОТК

        # ----------------- копируем цифры отгрузки по потребителям и изделиям на лист "Данные2" файла отчета -----------------
        # открываем файл отчета и лист "Данные2"
        file_otchet = "//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТЧЕТЫ БЗА по дефектности за месяц-квартал-год/Отчет по дефектности БЗА_текущий месяц.xlsx"
        wb3 = openpyxl.load_workbook(file_otchet)
        ws3 = wb3["Данные2"]  # открываем Лист в файле ОТК

        # словарь номеров столбцов на листе "Данные2" по наименованию месяца
        num_col_onchet = {
            "январь": 2,
            "февраль": 5,
            "март": 8,
            "апрель": 11,
            "май": 14,
            "июнь": 17,
            "июль": 20,
            "август": 23,
            "сентябрь": 26,
            "октябрь": 29,
            "ноябрь": 32,
            "декабрь": 35,
        }

        # номер колонки в которую производится копирование данных
        col_otchet = num_col_onchet[month]

        # создаем итераторы для перебора списка отгрузки по ММЗ
        it_mmz_2 = iter(mmz)

        # копируем значения по ММЗ
        for i in range(6, 14):
            ws3.cell(row=i, column=col_otchet).value = next(it_mmz_2)

        # копируем значения по МАЗ
        ws3.cell(row=16, column=col_otchet).value = maz
        # копируем значения по ГОМСЕЛЬМАШ
        ws3.cell(row=20, column=col_otchet).value = gomsel
        # копируем значения по МЗКТ
        ws3.cell(row=24, column=col_otchet).value = mzkt
        # копируем значения по БелАЗ
        ws3.cell(row=28, column=col_otchet).value = belaz
        # копируем значения по САЛЕО-Гомель
        ws3.cell(row=32, column=col_otchet).value = saleo
        # копируем значения по ХХХ-РБ
        ws3.cell(row=36, column=col_otchet).value = xxx_rb

        # копируем значения по УРАЛ
        ws3.cell(row=40, column=col_otchet).value = ural
        # копируем значения по РОСТСЕЛЬМАШ (ПК)
        ws3.cell(row=44, column=col_otchet).value = rostsel_1
        # # копируем значения по РОСТСЕЛЬМАШ (ТКР)
        # ws3.cell(row=135, column=col_otchet).value = rostsel_2
        # # копируем значения по РОСТСЕЛЬМАШ (ВН)
        # ws3.cell(row=136, column=col_otchet).value = rostsel_3
        # # копируем значения по РОСТСЕЛЬМАШ (МН)
        # ws3.cell(row=137, column=col_otchet).value = rostsel_4
        # копируем значения по КАМАЗ (ПК)
        ws3.cell(row=48, column=col_otchet).value = kamaz_1
        # # копируем значения по КАМАЗ (ВН)
        # ws3.cell(row=153, column=col_otchet).value = kamaz_2
        # # копируем значения по КАМАЗ (МН)
        # ws3.cell(row=154, column=col_otchet).value = kamaz_3
        # копируем значения по ЯМЗ (ПК)
        ws3.cell(row=52, column=col_otchet).value = ymz_1
        # копируем значения по ЯМЗ (ВН)
        ws3.cell(row=53, column=col_otchet).value = ymz_2
        # копируем значения по ЯМЗ (МН)
        ws3.cell(row=54, column=col_otchet).value = ymz_3
        # копируем значения по ПТЗ С-Петербург
        ws3.cell(row=57, column=col_otchet).value = ptz
        # копируем значения по ПАЗ (ПК)
        ws3.cell(row=61, column=col_otchet).value = paz_1
        # # копируем значения по ПАЗ (ВН)
        # ws3.cell(row=192, column=col_otchet).value = paz_2
        # копируем значения по ЧСДМ
        ws3.cell(row=65, column=col_otchet).value = csdm
        # копируем значения по Туле
        ws3.cell(row=69, column=col_otchet).value = tula
        # копируем значения по БАЗ
        ws3.cell(row=73, column=col_otchet).value = baz
        # копируем значения по ХХ-РФ-1
        ws3.cell(row=77, column=col_otchet).value = xx_rf_1
        # копируем значения по ХХ-РФ-2
        ws3.cell(row=81, column=col_otchet).value = xx_rf_2
        # копируем значения по ХХ-РФ-3
        ws3.cell(row=85, column=col_otchet).value = xx_rf_3

        wb3.save(file_otchet)  # сохраняем изменения
        wb3.close()  # закрываем файл отчета


if __name__ == "__main__":
    pr = ExcelSheetCopier()
    pr.copy_in_otk(3, 21, 1)  # ТКР
    pr.copy_in_otk(23, 40, 2)  # ПК
    pr.copy_in_otk(42, 59, 3)  # ВН
    pr.copy_in_otk(61, 75, 4)  # МН
    pr.copy_in_otk(77, 83, 5)  # ГП
    pr.copy_in_otk(85, 96, 6)  # ЦМФ
    pr.copy_in_otk(110, 114, 11)  # штанга и коромысло

    print("Данные по отгрузке скопированы на лист месяца файла ОТК")

    grp = DataCopierGarant()
    grp.copy_garant()

    print(
        'Данные скопированы на лист "Гарантийный парк" файла ОТК и "Данные2" файла отчета'
    )
