import openpyxl
import xlwings
from openpyxl.styles import PatternFill


# открываем таблицу ОСиМ и нужный Лист
filename1 = '//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК/Отгрузка для ОТК_текущий год.xlsx'
# флаг True - считываем только значение ячейки
wb1 = openpyxl.load_workbook(filename1, data_only=True)
name_list = wb1.sheetnames  # список имен всех Листов таблицы ОСиМ
name_osim = name_list[0]    # имя первого Листа
ws1 = wb1[name_osim]        # получаем первый Лист таблицы ОСиМ по имени

# открываем таблицу ОТК и нужный Лист
filename2 = '//Server/otk/ОТЧЕТНОСТЬ БЗА/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК/ОТГРУЗКА+ГАРАНТИЙНЫЙ ПАРК_текущий год.xlsx'
wb2 = openpyxl.load_workbook(filename2)
# чтобы имя Листа таблицы ОСиМ начиналось с большой буквы <title()> и по этому имени открываем Лист в таблице ОТК
name_otk = name_osim.title()
ws2 = wb2[name_otk]


class Pre_copy:
    """
    класс для копирования значений отгрузки из файла ОСиМ в файл ОТК на Лист "Месяц"
    row_start: строка начала диапазона по изделию (ТКР, ПК, ВН и т.д.)
    row_end: строка окончания диапазона
    k: поправочный коэффициент - разница номеров строк начала диапазона в таблице ОТК и ОСиМ
       ТКР = 1, ПК = 4, ВН = 9, МН = 13, ГП = 18, ЦМФ = 21, штанга = 21
    col_start: столбец начала диапазона по изделию в АСП = 3, т.к совпадает в таблицах ОСиМ и ОТК
    col_end: столбец окончания диапазона по изделию в АСП = 21, т.к совпадает
    """

    def copy_in_otk(self, row_start, row_end, k, col_start=3, col_end=21):
        for i in range(row_start, row_end + 1):
            for j in range(col_start, col_end + 1):
                # копируем значения по АСП
                # временная переменная - значение ячеек в таблице ОСиМ
                t_1 = ws1.cell(row=i, column=j).value
                # если значение в ячейке есть -> в int, иначе пустая строка
                cell = int(t_1) if t_1 else ''
                # копируем значения по АСП из таблицы ОСиМ в ОТК
                ws2.cell(row=i + k, column=j).value = cell

            # копируем значения по Запчасти
            # временная переменная (22 - номер столбца в ОСиМ)
            t_2 = ws1.cell(row=i, column=22).value
            # проверка значения и перевод в int
            cell = int(t_2) if t_2 else ''
            # копируем значения по Запчасти в ОТК (24 - столбец в ОТК)
            ws2.cell(row=i + k, column=24).value = cell

        wb2.save(filename2)  # сохраняем файл ОТК
        wb1.close()
        wb2.close()


class Copy_value:
    """
    класс для копирования значений отгрузки из файла ОТК из Листа "Месяц" на Лист "Гарантийный парк"
    гарантийный парк по потребителям на Листе "Гарантийный парк" расчитывается формулами в Excel
    """

    def copy_garant(self):
        # модулем xlwings в таблице ОТК фиксируем значения в ячейках, где используются формулы
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(filename2)
        excel_book.save()
        excel_book.close()
        excel_app.quit()

        wb1 = openpyxl.open(filename2, read_only=True, data_only=True)
        ws1 = wb1[name_otk]

        # номера строк с суммарной месячной отгрузкой: ТКР = 20, ПК = 37, ВН = 52, МН = 66, ГП = 75, ЦМФ = 89
        num_strok = (20, 37, 52, 66, 75, 89)

        # список значений суммарной отгрузки по изделиям на АСП
        asp = list(ws1.cell(i, 22).value for i in num_strok)
        # список значений отгрузки по изделиям в ЗАПЧАСТЬ
        zap = list(ws1.cell(i, 24).value for i in num_strok)
        # список значений отгрузки по изделиям на ММЗ
        mmz = list(ws1.cell(i, 3).value for i in num_strok)

        maz = ws1.cell(row=37, column=4).value     # копируем значения по МАЗ
        # копируем значения по ГОМСЕЛЬМАШ
        gomsel = ws1.cell(row=37, column=7).value
        ural = ws1.cell(row=37, column=12).value   # копируем значения по УРАЛ
        kamaz = ws1.cell(row=37, column=10).value  # копируем значения по КАМАЗ
        # копируем значения по РОСТСЕЛЬМАШ (ПК)
        rostsel_1 = ws1.cell(row=37, column=11).value
        # копируем значения по РОСТСЕЛЬМАШ (ТКР)
        rostsel_2 = ws1.cell(row=20, column=11).value
        # копируем значения по РОСТСЕЛЬМАШ (ВН)
        rostsel_3 = ws1.cell(row=52, column=11).value
        # копируем значения по РОСТСЕЛЬМАШ (МН)
        rostsel_4 = ws1.cell(row=66, column=11).value
        # копируем значения по ЯМЗ (ПК)
        ymz_1 = ws1.cell(row=37, column=17).value
        # копируем значения по ЯМЗ (ВН)
        ymz_2 = ws1.cell(row=52, column=17).value
        # копируем значения по ЯМЗ (МН)
        ymz_3 = ws1.cell(row=66, column=17).value
        belaz = ws1.cell(row=37, column=5).value   # копируем значения по БелАЗ
        mzkt = ws1.cell(row=37, column=6).value    # копируем значения по МЗКТ
        kraz = ws1.cell(row=37, column=14).value   # копируем значения по КрАЗ
        # копируем значения по ХТЗ Белгород
        htz_b = ws1.cell(row=37, column=9).value
        bzkt = ws1.cell(row=37, column=13).value   # копируем значения по БЗКТ
        # копируем значения по САЛЕО-Гомель
        saleo = ws1.cell(row=89, column=8).value
        # копируем значения по ПТЗ С-Петербург
        ptz = ws1.cell(row=37, column=15).value
        csdm = ws1.cell(row=37, column=16).value   # копируем значения по ЧСДМ
        tula = ws1.cell(row=37, column=18).value   # копируем значения по Туле
        baz = ws1.cell(row=37, column=21).value    # копируем значения по БАЗ
        # копируем значения по ХТЗ Харьков
        htz_h = ws1.cell(row=37, column=20).value

        wb1.close()

        wb2 = openpyxl.open(filename2)
        # делаем активным лист "Гарантийный парк"
        ws2 = wb2['Гарантийный парк']

        # словари номеров колонок по АСП и ЗАПЧАСТИ по наименованию месяца
        num_col_asp = {'январь': 2, 'февраль': 3, 'март': 4, 'апрель': 5, 'май': 6, 'июнь': 7,
                       'июль': 8, 'август': 9, 'сентябрь': 10, 'октябрь': 11, 'ноябрь': 12, 'декабрь': 13}
        num_col_zap = {'январь': 14, 'февраль': 15, 'март': 16, 'апрель': 17, 'май': 18, 'июнь': 19,
                       'июль': 20, 'август': 21, 'сентябрь': 22, 'октябрь': 23, 'ноябрь': 24, 'декабрь': 25}

        # наименование месяца в который производится копирование данных
        month = name_otk.split()[0].lower()

        # номер колонки в которую производится копирование данных
        col_asp = num_col_asp[month]
        col_zap = num_col_zap[month]

        # создаем итераторы для перебора списков значений
        it_asp = iter(asp)
        it_zap = iter(zap)
        it_mmz = iter(mmz)

        for i in range(6, 12):
            ws2.cell(row=i, column=col_asp).value = next(
                it_asp)  # копируем значения по АСП
            ws2.cell(row=i, column=col_zap).value = next(
                it_zap)  # копируем значения по ЗАПЧАСТИ

        for i in range(35, 41):
            ws2.cell(row=i, column=col_asp).value = next(
                it_mmz)  # копируем значения по ММЗ

        # копируем значения по МАЗ
        ws2.cell(row=59, column=col_asp).value = maz
        # копируем значения по ГОМСЕЛЬМАШ
        ws2.cell(row=68, column=col_asp).value = gomsel
        # копируем значения по УРАЛ
        ws2.cell(row=85, column=col_asp).value = ural
        # копируем значения по КАМАЗ
        ws2.cell(row=94, column=col_asp).value = kamaz
        # копируем значения по РОСТСЕЛЬМАШ (ПК)
        ws2.cell(row=103, column=col_asp).value = rostsel_1
        # копируем значения по РОСТСЕЛЬМАШ (ТКР)
        ws2.cell(row=104, column=col_asp).value = rostsel_2
        # копируем значения по РОСТСЕЛЬМАШ (ВН)
        ws2.cell(row=105, column=col_asp).value = rostsel_3
        # копируем значения по РОСТСЕЛЬМАШ (МН)
        ws2.cell(row=106, column=col_asp).value = rostsel_4
        # копируем значения по ЯМЗ (ПК)
        ws2.cell(row=121, column=col_asp).value = ymz_1
        # копируем значения по ЯМЗ (ВН)
        ws2.cell(row=122, column=col_asp).value = ymz_2
        # копируем значения по ЯМЗ (МН)
        ws2.cell(row=123, column=col_asp).value = ymz_3
        # копируем значения по БелАЗ
        ws2.cell(row=136, column=col_asp).value = belaz
        # копируем значения по МЗКТ
        ws2.cell(row=145, column=col_asp).value = mzkt
        # копируем значения по КрАЗ
        ws2.cell(row=154, column=col_asp).value = kraz
        # копируем значения по ХТЗ Белгород
        ws2.cell(row=163, column=col_asp).value = htz_b
        # копируем значения по БЗКТ
        ws2.cell(row=172, column=col_asp).value = bzkt
        # копируем значения по САЛЕО-Гомель
        ws2.cell(row=181, column=col_asp).value = saleo
        # копируем значения по ПТЗ С-Петербург
        ws2.cell(row=190, column=col_asp).value = ptz
        # копируем значения по ЧСДМ
        ws2.cell(row=199, column=col_asp).value = csdm
        # копируем значения по Туле
        ws2.cell(row=208, column=col_asp).value = tula
        # копируем значения по БАЗ
        ws2.cell(row=217, column=col_asp).value = baz
        # копируем значения по ХТЗ Харьков
        ws2.cell(row=229, column=col_asp).value = htz_h

        # словарь номеров колонок для заливки ячеек белым цветом после расчета гарантийного парка
        mth_color = {'январь': 28, 'февраль': 29, 'март': 30, 'апрель': 31, 'май': 32, 'июнь': 33,
                     'июль': 34, 'август': 35, 'сентябрь': 36, 'октябрь': 37, 'ноябрь': 38, 'декабрь': 39}
        for row in range(2, 142):
            # заливка белым цветом: : RGB (255, 255, 255), в шестнадцатеричном значении FFFFFF
            ws2.cell(row, mth_color[month]).fill = PatternFill(
                fill_type='solid', fgColor='FFFFFF')

        wb2.save(filename2)
        wb2.close()


if __name__ == '__main__':
    pr = Pre_copy()
    pr.copy_in_otk(3, 18, 1)     # ТКР
    pr.copy_in_otk(20, 34, 2)    # ПК
    pr.copy_in_otk(36, 48, 3)    # ВН
    pr.copy_in_otk(50, 61, 4)    # МН
    pr.copy_in_otk(63, 69, 5)    # ГП
    pr.copy_in_otk(71, 82, 6)    # ЦМФ
    pr.copy_in_otk(103, 107, 4)  # штанга и коромысло

    grp = Copy_value()
    grp.copy_garant()
