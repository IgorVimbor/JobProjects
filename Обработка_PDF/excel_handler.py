import pandas as pd
from openpyxl import load_workbook

class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def write_data(self, data):
        try:
            # Пытаемся загрузить существующий файл
            workbook = load_workbook(self.excel_path)
            sheet = workbook.active
        except FileNotFoundError:
            # Если файл не существует, создаем новый
            df = pd.DataFrame(columns=list(data.keys()))
            df.to_excel(self.excel_path, index=False)
            workbook = load_workbook(self.excel_path)
            sheet = workbook.active

        # Находим первую пустую строку
        row = sheet.max_row + 1

        # Записываем данные
        for col, value in enumerate(data.values(), 1):
            sheet.cell(row=row, column=col, value=value)

        workbook.save(self.excel_path)