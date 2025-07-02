import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime

import app_total.db_search.db_search_modul as t


year_now = datetime.today().year  # текущий год

# файл базы рекламаций ОТК с учетом текущего года
# file = "//Server/otk/1 ГАРАНТИЯ на сервере/" + str(year_now) + "-2019_ЖУРНАЛ УЧЁТА.xlsm"
file_database = f"D:/РАБОТА/{str(year_now)}-2019_ЖУРНАЛ УЧЁТА.xlsm"

# файл отчета по результатам поиска
# file_report = "//Server/otk/Support_files_не_удалять!!!/Отчет по результатам поиска по базе ОТК.txt"
file_report = "АНАЛИТИЧЕСКАЯ_СИСТЕМА_УК/Отчет по результатам поиска по базе ОТК.txt"
# -------------------------------------------------------------------------------------------

def get_value():
    """функция возвращает номер года, списки двигателей и актов из строк вводимой информации"""
    global god
    global inf_1
    global inf_2

    god = ent_god.get()     # получение текста из строки ввода года
    if not god:
        messagebox.showinfo('ОШИБКА', 'Не введен год поиска')
    inf_1 = ent_dvig.get()  # получение текста из строки ввода номеров двигателей
    inf_2 = ent_act.get()   # получение текста из строки ввода номеров актов

    num1 = inf_1.split()    # формируем список двигателей из строки ввода
    num2 = inf_2.split()    # формируем список актов из строки ввода

    return god, num1, num2


def get_itog():
    """функция подготовки выборки из базы и вывода на экран"""
    global wb
    global sheet
    global cells_dvigs
    global cells_acts
    global dvigs
    global acts

    god, num1, num2 = get_value()

    # открываем файл базы рекламаций ОТК
    wb = load_workbook(file_database)
    # открываем лист по номеру года
    sheet = wb[god]

    # список двигателей и актов рекламаций из всей базы
    # колонка 20 - номера двигателей
    cells_dvigs = tuple(sheet.cell(row=i, column=20) for i in range(3, sheet.max_row+1))
    # колонка 13 - номера актов рекламаций ПРИОБРЕТАТЕЛЯ изделия
    cells_acts = tuple(sheet.cell(row=i, column=13) for i in range(3, sheet.max_row+1))

    # вспомогательные списки двигателей и актов рекламаций
    dvigs = tuple(str(cell.value) for cell in cells_dvigs if cell.value)
    acts = tuple(cell.value for cell in cells_acts)

    for n in num1:           # перебираем входящий список двигателей и ищем двигатель в списке dvigs
        if n not in dvigs:   # если в списке нет - печатаем результат
            # вывод текста
            text_1.insert(1.0, f'Двигателя {n} в базе нет\n')

        for i, cell in enumerate(cells_dvigs):
            # если в списке cells_dvigs двигатель есть - печатаем номер строки таблицы Excel
            if n in str(cell.value):
                row = i + 3
                text_1.insert(1.0, f'Двигатель {str(cell.value)}: строка - {row}\n')

    for n in num2:           # перебираем входящий список актов и ищем акт во вспомогательном списке acts
        if n not in acts:    # если во вспомогательном списке нет - печатаем результат
            # вывод текста
            text_1.insert(1.0, f'Акта {n} в базе нет\n')

        for i, cell in enumerate(cells_acts):
            # если в списке cells_acts акт есть - печатаем номер строки таблицы Excel
            if n in str(cell.value):
                row = i + 3
                # вывод текста
                text_1.insert(1.0, f'Акт {cell.value}: строка - {row}\n')


def clear_strok():  # функция очистки строк ввода номеров двигателей и актов
    if god:
        ent_god.delete(0, tk.END)
    if inf_1:
        ent_dvig.delete(0, tk.END)
    if inf_2:
        ent_act.delete(0, tk.END)


def clear_res():  # функция очистки поля вывода результата
    text_1.delete('1.0', tk.END)


def otchet_output():  # функция подготовки выборки из базы и печати отчета
    num1 = inf_1.split()
    num2 = inf_2.split()
    # записываем отчет в файл txt
    with open(file_report, "w", encoding="utf-8") as res_file:
        print('\n' * 2, file=res_file)

        if num1:  # если есть список двигателей из строки ввода
            # печатаем номера двигателей
            print('Номера двигателей:', ', '.join(str(n)
                  for n in num1), file=res_file)
            print(file=res_file)

        for n in num1:  # ищем двигатель во вспомогательном списке dvigs
            if n not in dvigs:   # если во вспомогательном списке нет - печатаем результат
                print('Двигателя', n, 'в базе нет', file=res_file)
                print(file=res_file)

            for i, cell in enumerate(cells_dvigs):
                # если в списке cells_dvigs двигатель есть - печатаем информацию из ячеек Excel
                if n in str(cell.value):
                    row = i + 3

                    print('Двигатель', str(cell.value), '| строка -', row, '|', file=res_file)
                    print('-' * 80, file=res_file)

                    print(
                        sheet.cell(row, t.ind('Наименование изделия')).value, '|',
                        sheet.cell(row, t.ind('Обозначение изделия')).value, '|',
                        'зав.№:', sheet.cell(row, t.ind('Заводской номер изделия')).value,
                        sheet.cell(row, t.ind('Дата изготовления изделия')).value, '|',
                        file=res_file
                    )

                    print(
                        'Где выявлен дефект:',
                        sheet.cell(row, t.ind('Период выявления дефекта')).value, '|', end=' ',
                        file=res_file
                    )

                    if sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value:
                        print(
                            'Р/А №:',
                            sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value,
                            '|', end=' ', file=res_file
                        )
                    else:
                        print('Р/А №: акта нет', '|', end=' ', file=res_file)

                    print(
                        'Дефект:',
                        sheet.cell(row, t.ind('Заявленный дефект изделия')).value,
                        file=res_file
                    )

                    if sheet.cell(row, t.ind('Номер акта исследования')).value:
                        print(
                            'Акт БЗА:',
                            sheet.cell(row, t.ind('Номер акта исследования')).value, 'от',
                            sheet.cell(row, t.ind('Дата акта исследования')).value.strftime('%d.%m.%Y'),
                            file=res_file
                        )
                    else:
                        print('Акт БЗА: акта нет', file=res_file)

                    print(
                        'Решение БЗА:',
                        sheet.cell(row, t.ind('Причины возникновения дефектов')).value, end=' ',
                        file=res_file
                    )

                    if sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value:
                        print(
                            '| отчетность БЗА -',
                            sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value,
                            file=res_file
                        )
                    else:
                        print(
                            '|',
                            sheet.cell(row, t.ind('Пояснения к причинам возникновения дефектов')).value,
                            file=res_file
                        )

                    print('-' * 80, file=res_file)
                    print(file=res_file)

        if num2:   # если есть список актов из строки ввода
            print('Номер акта:', ', '.join(n for n in num2),
                  file=res_file)  # печатаем номера актов рекламаций
            print(file=res_file)

        for n in num2:   # ищем акт во вспомогательном списке acts
            if n not in acts:  # если во вспомогательном списке нет - печатаем результат
                print('Акта', n, 'в базе нет', file=res_file)
                print(file=res_file)

            for i, cell in enumerate(cells_acts):
                # если в списке cells_acts есть акт - печатаем информацию из ячеек Excel
                if n in str(cell.value):
                    row = i + 3

                    print('Акт', cell.value, '| строка -', row, '|', file=res_file)
                    print('-' * 80, file=res_file)

                    print(
                        sheet.cell(row, t.ind('Наименование изделия')).value, '|',
                        sheet.cell(row, t.ind('Обозначение изделия')).value, '|',
                        'зав.№:', sheet.cell(row, t.ind('Заводской номер изделия')).value,
                        sheet.cell(row, t.ind('Дата изготовления изделия')).value, '|',
                        file=res_file
                    )

                    print(
                        'Где выявлен дефект:',
                        sheet.cell(row, t.ind('Период выявления дефекта')).value, '|', end=' ',
                        file=res_file
                    )

                    if sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value:
                        print(
                            'Р/А №:',
                            sheet.cell(row, t.ind('Номер рекламационного акта ПРИОБРЕТАТЕЛЯ изделия')).value,
                            '|', end=' ',
                            file=res_file
                        )
                    else:
                        print('Р/А №: акта нет', '|', end=' ', file=res_file)

                    print(
                        'Дефект:',
                        sheet.cell(row, t.ind('Заявленный дефект изделия')).value,
                        file=res_file
                    )

                    if sheet.cell(row, t.ind('Номер акта исследования')).value:
                        print(
                            'Акт БЗА:',
                            sheet.cell(row, t.ind('Номер акта исследования')).value,
                            'от', sheet.cell(row, t.ind('Дата акта исследования')).value.strftime('%d.%m.%Y'),
                            file=res_file
                        )
                    else:
                        print('Акт БЗА: акта нет', file=res_file)

                    print(
                        'Решение БЗА:', sheet.cell(row, t.ind('Причины возникновения дефектов')).value, end=' ',
                        file=res_file
                    )

                    if sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value:
                        print(
                            '| отчетность БЗА -',
                            sheet.cell(row, t.ind('Месяц отражения в статистике БЗА')).value,
                            file=res_file
                        )
                    else:
                        print(
                            '|',
                            sheet.cell(row, t.ind('Пояснения к причинам возникновения дефектов')).value,
                            file=res_file
                        )

                    print('-' * 80, file=res_file)
                    print(file=res_file)

    # вывод информационного окна о сохранении отчета
    messagebox.showinfo(
        'ИНФОРМАЦИЯ', 'ОТЧЕТ СОХРАНЕН.')


window = tk.Tk()

# Создается окно с заголовком
# меняем логотип Tkinter (перышко) на свой логотип
window.iconbitmap('app_total/IconBZA.ico')
# название заголовка в окне приложения
window.title('ПОИСК ПО БАЗЕ РЕКЛАМАЦИЙ ОТК')
window.geometry('650x600')    # размер окна приложения

# Делаем окно растягивающимся с центрированием по центру
window.columnconfigure(0, weight=1, minsize=250)
window.rowconfigure([0, 10], weight=1, minsize=100)

# Пустая строка перед фреймом
lbl_null_0 = tk.Label(text='')
lbl_null_0.pack()

# Создается рамка `frm_form` для ярлыка с текстом и полей для ввода информации.
frm_form = tk.Frame(relief=tk.SUNKEN, borderwidth=3)
frm_form.pack()

# ЗАПОЛНЯЕМ ПЕРВЫЙ ФРЕЙМ
lbl = tk.Label(master=frm_form,
               text='1. Введите ГОД в котором будет осуществляться поиск:   ', font=("Arial Bold", 10))
ent_god = tk.Entry(master=frm_form, width=15)

# Используем менеджер геометрии grid для размещения ярлыков и поля ввода номеров двигателей
lbl.grid(row=0, column=0, sticky='w')
ent_god.grid(row=0, column=1, sticky='w')

# Пустая строка между фреймами
lbl_null = tk.Label(text='')
lbl_null.pack()

# Создается новая рамка `frm_form_1` для ярлыков с текстом и полей для ввода информации.
frm_form_1 = tk.Frame(relief=tk.SUNKEN, borderwidth=3)
frm_form_1.pack()

# ЗАПОЛНЯЕМ ВТОРОЙ ФРЕЙМ
lbl_1 = tk.Label(master=frm_form_1,
                 text='2. ПОИСК ОСУЩЕСТВЛЯЕТСЯ ПО:', font=("Arial Bold", 10))
lbl_null_1 = tk.Label(master=frm_form_1, text='')  # пустая строка
lbl_dvig_1 = tk.Label(
    master=frm_form_1, text='- НОМЕРУ ДВИГАТЕЛЯ.', font=("Arial Bold", 10))
lbl_dvig_2 = tk.Label(master=frm_form_1, text='ВНИМАНИЕ! Если в номере двигателя есть БУКВЫ, '
                      'то вводите только ЦИФРЫ. Буквы НЕ надо.', font=("Arial Bold", 10))
lbl_dvig_3 = tk.Label(master=frm_form_1, text='Введите через пробел номера ДВИГАТЕЛЕЙ и нажмите НАЧАТЬ ПОИСК.',
                      font=("Arial Bold", 10))
ent_dvig = tk.Entry(master=frm_form_1, width=100)

# Используем менеджер геометрии grid для размещения ярлыков и поля ввода номеров двигателей
lbl_1.grid(row=0, column=0, sticky='w')
lbl_null_1.grid(row=1, column=0)
lbl_dvig_1.grid(row=2, column=0, sticky='w')
lbl_dvig_2.grid(row=3, column=0, sticky='w')
lbl_dvig_3.grid(row=4, column=0, sticky='w')
ent_dvig.grid(row=5, column=0)

lbl_null_2 = tk.Label(master=frm_form_1, text='')  # пустая строка
lbl_act_1 = tk.Label(
    master=frm_form_1, text='- ПО НОМЕРУ АКТА РЕКЛАМАЦИИ.', font=("Arial Bold", 10))
lbl_act_2 = tk.Label(master=frm_form_1, text='Введите через пробел номера АКТОВ РЕКЛАМАЦИЙ и нажмите НАЧАТЬ ПОИСК.',
                     font=("Arial Bold", 10))
ent_act = tk.Entry(master=frm_form_1, width=100)
lbl_null_3 = tk.Label(master=frm_form_1, text='')  # пустая строка

# Используем менеджер геометрии grid для размещения ярлыков и поля ввода номеров актов
lbl_null_2.grid(row=6, column=0)
lbl_act_1.grid(row=7, column=0, sticky='w')
lbl_act_2.grid(row=8, column=0, sticky='w')
ent_act.grid(row=9, column=0)
lbl_null_3.grid(row=10, column=0)

# Создает новую рамку frm_buttons для размещения кнопок НАЧАТЬ ПОИСК и ОЧИСТИТЬ СТРОКУ
frm_buttons = tk.Frame()
frm_buttons.pack(fill=tk.X, ipadx=5, ipady=5)
bnt_1 = tk.Button(master=frm_buttons, text='НАЧАТЬ ПОИСК',
                  font=("Arial Bold", 10), command=get_itog)
bnt_2 = tk.Button(master=frm_buttons, text='ОЧИСТИТЬ СТРОКИ',
                  font=("Arial Bold", 10), command=clear_strok)
bnt_2.pack(side=tk.RIGHT, ipadx=10)
bnt_1.pack(side=tk.RIGHT, padx=10, ipadx=10)

# Пустая строка между фреймами
lbl_null_4 = tk.Label(text='')
lbl_null_4.pack()

# Создается новая рамка frm_form_2 для ярлыков с текстом и поля для вывода результата
frm_form_2 = tk.Frame(relief=tk.SUNKEN, borderwidth=3)
frm_form_2.pack()

# ЗАПОЛНЯЕМ ТРЕТИЙ ФРЕЙМ
lbl_2 = tk.Label(master=frm_form_2, text='3. РЕЗУЛЬТАТ ПОИСКА:',
                 font=("Arial Bold", 10))
text_1 = tk.Text(master=frm_form_2, width=85, height=7,
                 background='white', font=("Arial Bold", 10))
lbl_3 = tk.Label(master=frm_form_2, text='Для получения отчета с подробной информацией нажмите СДЕЛАТЬ ОТЧЕТ',
                 font=("Arial Bold", 10))
lbl_4 = tk.Label(master=frm_form_2, text='')   # пустая строка

# Используем менеджер геометрии grid для размещения ярлыков и поля вывода результата
lbl_2.grid(row=0, column=0, sticky='w')
text_1.grid(row=1, column=0)
lbl_3.grid(row=2, column=0, sticky='w')
lbl_4.grid(row=3, column=0)

# Создает новую рамку frm_buttons_2 для размещения кнопок СДЕЛАТЬ ОТЧЕТ и СБРОСИТЬ РЕЗУЛЬТАТ
frm_buttons_2 = tk.Frame()
frm_buttons_2.pack(fill=tk.X, ipadx=5, ipady=5)
bnt_3 = tk.Button(master=frm_buttons_2, text='СДЕЛАТЬ ОТЧЕТ',
                  font=("Arial Bold", 10), command=otchet_output)
bnt_4 = tk.Button(master=frm_buttons_2, text='СБРОСИТЬ РЕЗУЛЬТАТ',
                  font=("Arial Bold", 10), command=clear_res)
bnt_4.pack(side=tk.RIGHT, ipadx=10)
bnt_3.pack(side=tk.RIGHT, padx=10, ipadx=10)

window.mainloop()
