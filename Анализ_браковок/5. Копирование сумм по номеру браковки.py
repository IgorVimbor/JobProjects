import pandas as pd
import warnings
from openpyxl import load_workbook

# Команда для удаления предупреждений Pandas в консоли
warnings.simplefilter(action="ignore", category=Warning)

file_bush = "C:/Users/BT-0529/Desktop/Литейный брак август2024.xlsx"
# колонки ['Брак', 'Наладка', 'гарант'] - Демидчик И.А.
# колонки ['Литей', 'Втулка', 'стоимость забракованного литья', 'отк'] - Ларионова А.В.

sheet_to_load = ["Брак", "Наладка", "гарант"]
dfs = []
# считываем данные из файла Excel и создаем общий датафрейм с листов 'Брак', 'Наладка', 'гарант'
for sheet_name in sheet_to_load:
    df_chank = pd.read_excel(
        file_bush, sheet_name=sheet_name, usecols=["№брак.", "Сумма"], header=2
    )
    dfs.append(df_chank)

df = pd.concat(dfs, ignore_index=True)
print("Файл бухгалтерии загружен.\n")

# удаляем строки с отсутствующими значениями в столбце "№брак."
df.dropna(subset=["№брак."], inplace=True)
# переводим значения в столбце "№брак." в целочисленные
df = df.astype({"№брак.": "int16", "Сумма": "float16"})
# устанавливаем индексом столбец "№брак."
df = df.set_index("№брак.")

# создаем словарь (ключ - столбец "№брак.", значение - столбец "Сумма" с округлением до 2-х знаков)
dct = {k: round(v, 2) for k, v in df.to_dict()["Сумма"].items()}
print("Создан словарь с номерами актов и суммами.\n")
# --------------------------------------------------------------------------------------------

# Загружаем файл Excel ОТК
file = "//Server/otk/2 ИННА/Списание БРАКА по ЦЕХАМ/ЖУРНАЛ УЧЕТА актов о браке_2020-2024.xlsx"
wb = load_workbook(file)
sheet = wb["2024"]  # активируем лист "2024"
print("Файл ОТК загружен.\n")

# Задаем номера столбцов
column_nom = 1  # номер столбца "Номер_акта" (A = 1, B = 2 и т.д.)
column_sum = 10  # номер столбца "Сумма_по_акту"
print("Началось копирование данных ...\n")
# Обрабатываем строки листа, начиная с третьей строки (первая и вторая строка - заголовок)
for row in range(3, sheet.max_row + 1):
    # переменная cell_nom - значение ячейки в столбце "№_акта_НП", т.е. номер акта
    cell_nom = sheet.cell(row=row, column=column_nom).value

    # если номер акта совпадает с ключом в словаре, изменяем значение в ячейке столбца "Сумма_по_акту"
    if cell_nom in dct:
        # нумерация строк таблицы Excel отличается на +2 от номера акта, поэтому строку находим как row=cell_nom + 2
        sheet.cell(row=cell_nom + 2, column=column_sum).value = dct[cell_nom]

# Сохраняем изменения
wb.save(file)
wb.close()

print("Файл ОТК с новыми данными записан.")
# -----------------------------------------------------------------------------------------------
