from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from urllib.parse import quote

from reclamations.models import Reclamation
from investigations.models import Investigation


class ReclamationExcelExporter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Рекламации и исследования"  # можно задать имя листа

    def _write_reclamation_data(self):
        """Метод для записи данных из базы данных в Excel"""
        # Заголовки для рекламаций
        headers = [
            "ID рекламации",
            "Входящий номер",
            "Дата получения",
            "Статус",
            "Отправитель",
            "Исходящий номер",
            "Изделие",
            "Заводской номер",
            "Дефект",
            # Поля исследования
            "Номер акта исследования",
            "Дата акта",
            "Заключение",
        ]

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col, value=header)

        # Стиль ячейки (серый цвет)
        light_gray_fill = PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        # Получаем рекламации
        reclamations = Reclamation.objects.all().select_related(
            "product_name", "product"
        )

        # Записываем данные
        for row, rec in enumerate(reclamations, 2):
            # Данные рекламации
            self.ws.cell(row=row, column=1, value=rec.id)
            self.ws.cell(row=row, column=2, value=rec.incoming_number)
            self.ws.cell(row=row, column=3, value=rec.message_received_date)
            self.ws.cell(row=row, column=4, value=rec.get_status_display())
            self.ws.cell(row=row, column=5, value=rec.sender)
            self.ws.cell(row=row, column=6, value=rec.sender_outgoing_number)
            self.ws.cell(row=row, column=7, value=str(rec.product))
            self.ws.cell(row=row, column=8, value=rec.product_number)
            self.ws.cell(row=row, column=9, value=rec.claimed_defect)

            # Получаем связанное исследование (если есть)
            try:
                investigation = rec.investigation
                self.ws.cell(row=row, column=10, value=investigation.act_number)
                self.ws.cell(row=row, column=11, value=investigation.act_date)
            # Если исследования нет, оставляем ячейки пустыми
            except Reclamation.investigation.RelatedObjectDoesNotExist:
                self.ws.cell(row=row, column=10, value="")
                self.ws.cell(row=row, column=11, value="")

    def _adjust_column_width(self):
        """Метод для автоматической настройки ширины колонок в файле Excel"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 2
            self.ws.column_dimensions[column_letter].width = adjusted_width

    # def _create_response(self):
    #     """Метод для запуска загрузки файла в браузере"""
    #     current_date = datetime.now().strftime("%Y-%m-%d")
    #     # Название файла на английском языке (могут быть проблемы с кодировкой русского названия файла)
    #     filename = f"reclamations_and_investigations_{current_date}.xlsx"

    #     response = HttpResponse(
    #         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #     )
    #     response["Content-Disposition"] = f"attachment; filename={filename}"
    #     self.wb.save(response)
    #     return response

    def _create_response(self):
        """Метод для запуска загрузки файла в браузере"""
        current_date = datetime.now().strftime("%Y-%m-%d")

        # Базовое имя файла (можно на русском языке)
        base_filename = f"ЖУРНАЛ УЧЕТА_{current_date}.xlsx"

        # Кодируем имя файла для HTTP-заголовка
        encoded_filename = quote(base_filename)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Добавляем два варианта Content-Disposition для лучшей совместимости
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{encoded_filename}; filename={encoded_filename}"
        )

        self.wb.save(response)
        return response

    def export_to_excel(self):
        """Общий метод для экспорта данных из базы данных в Excel"""
        self._write_reclamation_data()
        self._adjust_column_width()
        return self._create_response()
