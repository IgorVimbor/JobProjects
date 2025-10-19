from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reclamations.models import Reclamation
from reports.config.paths import get_excel_exporter_path


# Группировка полей (чек-боксов) на странице
MODEL_GROUP_MAPPING = {
    "reclamation": "Рекламации",
    "investigation": "Исследования",
    "claims": "Претензии",
}


class UniversalExcelExporter:
    """Универсальный экспортер в Excel с возможностью выбора полей"""

    def __init__(self, selected_fields=None, year=None):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Выгрузка данных"
        self.selected_fields = selected_fields or []
        # Доступные поля для экспорта
        self.field_config = self._get_field_configuration()
        self.year = year  # Год данных из базы
        # Путь для сохранения файла с учетом выбранного года
        self.save_path = get_excel_exporter_path(year)

    def _get_field_configuration(self):
        """Упрощенная конфигурация доступных полей для экспорта"""
        # Изначально конфигурация полей была через словари. Например:
        # "reclamation.message_received_date": {
        #     "header": "Дата поступления сообщения",
        #     "model": "reclamation",
        #     "field": "message_received_date",
        #     "type": "direct",
        # },
        # Сейчас сделано через кортежи, что значительно проще.
        return {
            # Прямые поля рекламации
            "reclamation.id": ("Номер строки", "direct"),
            "reclamation.message_received_date": (
                "Дата поступления сообщения",
                "direct",
            ),
            "reclamation.defect_period": ("Период выявления дефекта", "related_str"),
            "reclamation.product_name": ("Наименование изделия", "related_str"),
            "reclamation.product": ("Обозначение изделия", "related_str"),
            "reclamation.products_count": ("Количество изделий", "direct"),
            "reclamation.product_number": ("Номер изделия", "direct"),
            "reclamation.manufacture_date": ("Дата изготовления", "direct"),
            "reclamation.claimed_defect": ("Заявленный дефект", "direct"),
            "reclamation.consumer_act_number": ("Номер акта рекламации", "direct"),
            "reclamation.consumer_act_date": ("Дата акта рекламации", "direct"),
            "reclamation.engine_brand": ("Марка двигателя", "direct"),
            "reclamation.engine_number": ("Номер двигателя", "direct"),
            "reclamation.transport_name": ("Транспортное средство", "direct"),
            "reclamation.mileage_operating_time": ("Пробег/наработка", "direct"),
            "reclamation.receipt_invoice_number": ("Накладная прихода", "direct"),
            "reclamation.receipt_invoice_date": ("Дата накладной прихода", "direct"),
            "reclamation.pkd_number": ("Номер 8D (ПКД)", "direct"),
            # Поля исследований (OneToOne)
            "investigation.act_number": ("Номер акта исследования", "investigation"),
            "investigation.act_date": ("Дата акта исследования", "investigation"),
            "investigation.solution": ("Заключение", "investigation_choice"),
            "investigation.guilty_department": (
                "Виновное подразделение",
                "investigation",
            ),
            "investigation.defect_causes": ("Причины дефекта", "investigation"),
            "investigation.defect_causes_explanation": (
                "Пояснения к причинам",
                "investigation",
            ),
            "investigation.defective_supplier": (
                "Поставщик комплектующего",
                "investigation",
            ),
            "investigation.shipment_invoice_number": (
                "Накладная отгрузки",
                "investigation",
            ),
            "investigation.shipment_invoice_date": (
                "Дата накладной отгрузки",
                "investigation",
            ),
            # Поля претензий (ForeignKey множественные)
            "claims.claim_number": ("Номер претензии", "claims"),
            "claims.claim_date": ("Дата претензии", "claims"),
            "claims.type_money": ("Денежная единица", "claims"),
            "claims.claim_amount_all": ("Сумма претензии", "claims"),
            "claims.claim_amount_act": ("Сумма по акту рекламации", "claims"),
            "claims.costs_all": ("Признано по претензии", "claims"),
            "claims.costs_act": ("Признано по акту", "claims"),
            "claims.response_number": ("Ответ на претензию", "claims"),
            "claims.response_date": ("Дата ответа на претензию", "claims"),
        }

    @classmethod
    def get_available_fields(cls):
        """Список доступных полей"""
        exporter = cls()
        fields = []

        for field_key, (header, field_type) in exporter.field_config.items():
            model = field_key.split(".")[0]
            fields.append(
                {
                    "key": field_key,
                    "header": header,
                    "model": model,
                    "group": MODEL_GROUP_MAPPING.get(model, "Прочее"),
                }
            )

        return fields

    def _get_field_value(self, reclamation, field_key):
        """Метод для получения значения поля по конфигурации"""
        if field_key not in self.field_config:
            return ""

        header, field_type = self.field_config[field_key]
        model_name, field_name = field_key.split(".")

        try:
            if field_type == "direct":
                value = getattr(reclamation, field_name)
                return self._format_value(value)

            elif field_type == "related_str":
                related_obj = getattr(reclamation, field_name)
                return str(related_obj) if related_obj else ""

            elif field_type == "investigation":
                try:
                    investigation = reclamation.investigation
                    value = getattr(investigation, field_name)
                    return self._format_value(value)
                except Reclamation.investigation.RelatedObjectDoesNotExist:
                    return ""

            elif field_type == "investigation_choice":
                try:
                    investigation = reclamation.investigation
                    return getattr(investigation, f"get_{field_name}_display")()
                except Reclamation.investigation.RelatedObjectDoesNotExist:
                    return ""

            elif field_type == "claims":
                # Специальная обработка для дат в претензиях
                claims = reclamation.claims.all()
                if not claims.exists():
                    return ""

                values = []
                for claim in claims:
                    value = getattr(claim, field_name, None)
                    if value is not None:
                        # Для дат возвращаем только первую дату, а не все через запятую
                        if field_name.endswith("_date") and hasattr(value, "strftime"):
                            return value  # Возвращаем как datetime для Excel
                        else:
                            values.append(str(value))

                # Для не-дат объединяем через запятую
                return ", ".join(values) if values else ""

        except Exception as e:
            print(f"Ошибка получения значения для {field_key}: {e}")
            return ""

    def _format_value(self, value):
        """Форматирование считанных значений полей"""
        if value is None:
            return ""
        elif hasattr(value, "strftime"):  # Дата
            return value  # Возвращаем datetime для Excel
        else:
            return str(value)

    def _apply_date_formatting(self):
        """Форматирование полей с датами"""
        date_columns = []

        for col_num, field_key in enumerate(self.selected_fields, 1):
            if field_key in self.field_config:
                # Проверяем по имени поля
                field_name = field_key.split(".")[1]
                if field_name.endswith("_date") or "date" in field_name.lower():
                    date_columns.append(col_num)

        # Применяем форматирование дат
        for col_num in date_columns:
            column_letter = get_column_letter(col_num)
            for row in range(2, self.ws.max_row + 1):
                cell = self.ws[f"{column_letter}{row}"]
                if cell.value and hasattr(cell.value, "strftime"):
                    cell.number_format = "DD.MM.YYYY"

    def _write_data(self):
        """Запись данных в Excel по выбранным полям"""
        if not self.selected_fields:
            return

        # Записываем заголовки
        for col, field_key in enumerate(self.selected_fields, 1):
            if field_key in self.field_config:
                header = self.field_config[field_key][0]
                cell = self.ws.cell(row=1, column=col, value=header)
                # Применяем форматирование к заголовкам
                self._apply_cell_formatting(cell, is_header=True)

        # Получаем данные
        queryset = Reclamation.objects.select_related(
            "product_name", "product"
        ).prefetch_related("investigation", "claims")

        # Применяем фильтр по году если выбран
        if self.year and str(self.year) != "all":
            queryset = queryset.filter(year=self.year)

        # Записываем данные
        for row, reclamation in enumerate(queryset, 2):
            for col, field_key in enumerate(self.selected_fields, 1):
                value = self._get_field_value(reclamation, field_key)
                cell = self.ws.cell(row=row, column=col, value=value)
                # Применяем форматирование
                self._apply_cell_formatting(cell, is_header=False)

    def _apply_cell_formatting(self, cell, is_header=False):
        """Форматирование ячеек таблицы Excel"""

        # Границы для всех ячеек
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        # Центрирование по центру (горизонтально и вертикально) + перенос строк
        alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        if is_header:
            # Заголовки: размер 8, не жирный
            cell.font = Font(size=8, bold=False)
            # Серый фон для заголовков
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
        else:
            # Обычные ячейки: размер 9. Фон по умолчанию - белый.
            cell.font = Font(size=9)

        # Применяем границы и выравнивание ко всем ячейкам
        cell.border = border
        cell.alignment = alignment

    def _adjust_column_width(self):
        """Настройка ширины колонок"""
        # Словарь: название поля → ширина столбца
        custom_widths = {
            "Номер строки": 9,
            "Дата поступления сообщения": 14,
            "Период выявления дефекта": 17,
            "Наименование изделия": 16,
            "Обозначение изделия": 17,
            "Количество изделий": 10,
            "Номер изделия": 10,
            "Дата изготовления": 10,
            "Заявленный дефект": 20,
            "Номер акта рекламации": 14,
            "Марка двигателя": 16,
            "Номер двигателя": 14,
            "Транспортное средство": 16,
            "Пробег/наработка": 14,
            "Номер 8D (ПКД)": 14,
            "Номер акта исследования": 14,
            "Дата акта исследования": 14,
            "Заключение": 11,
            "Виновное подразделение": 14,
            "Причины дефекта": 22,
            "Пояснения к причинам": 26,
            "Поставщик комплектующего": 22,
            "Накладная отгрузки": 11,
            "Номер претензии": 14,
            "Ответ на претензию": 14,
        }

        # Получаем заголовки для сопоставления
        headers = []
        for field_key in self.selected_fields:
            if field_key in self.field_config:
                headers.append(self.field_config[field_key][0])

        # Устанавливаем ширину для каждого столбца
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)

            # Берем ширину из словаря или ставим дефолтную
            width = custom_widths.get(header, 12)  # 12 - дефолтная ширина

            self.ws.column_dimensions[column_letter].width = width

    def _apply_filter_and_freeze_row(self):
        """Устанавливаем автофильтр и закрепляем строку заголовков"""
        if self.ws.max_row > 1:  # Проверяем, что есть данные
            # Фильтр только для первой строки (заголовки)
            header_range = f"A1:{get_column_letter(len(self.selected_fields))}1"
            self.ws.auto_filter.ref = header_range

            # Закрепляем первую строку (заголовки)
            self.ws.freeze_panes = "A2"

    def export_to_excel(self):
        """Основной метод экспорта в файл Excel"""

        if not self.selected_fields:
            raise ValueError("Не выбраны поля для экспорта")

        self._write_data()  # Записываем данные с форматированием
        self._apply_date_formatting()  # Форматируем даты в Excel
        self._adjust_column_width()  # Настраиваем ширину столбцов
        self._apply_filter_and_freeze_row()  # Добавляем фильтры и закрепляем заголовок

        # Сохраняем файл на диск
        self.wb.save(self.save_path)
        return True
