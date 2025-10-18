from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reclamations.models import Reclamation
from reports.config.paths import get_excel_exporter_path


# Группировка полей (чек-боксов) для выбора
MODEL_GROUP_MAPPING = {
    "reclamation": "Рекламации",
    "investigation": "Исследования",
    "claims": "Претензии",
}


class UniversalExcelExporter:
    """Универсальный экспортер Excel с выбором полей"""

    def __init__(self, selected_fields=None, year=None):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Выгрузка данных"
        self.selected_fields = selected_fields or []
        # Доступные поля для экспорта
        self.field_config = self._get_field_configuration()
        self.year = year  # Год данных из базы
        # Путь для сохранения файла с учетом выбранного года
        self.save_path = get_excel_exporter_path(year)

    def _get_field_configuration(self):
        """Конфигурация доступных полей для экспорта"""
        return {
            # Поля Reclamation
            "reclamation.id": {
                "header": "Номер строки",
                "model": "reclamation",
                "field": "id",
                "type": "direct",
            },
            "reclamation.message_received_date": {
                "header": "Дата поступления сообщения",
                "model": "reclamation",
                "field": "message_received_date",
                "type": "direct",
            },
            # "reclamation.status": {
            #     "header": "Статус рекламации",
            #     "model": "reclamation",
            #     "field": "status",
            #     "type": "choice_display",  # Для полей с choices
            # },
            "reclamation.defect_period": {
                "header": "Период выявления дефекта",
                "model": "reclamation",
                "field": "defect_period",
                "type": "related_str",  # Для связанных объектов
            },
            "reclamation.product_name": {
                "header": "Наименование изделия",
                "model": "reclamation",
                "field": "product_name",
                "type": "related_str",
            },
            "reclamation.product": {
                "header": "Обозначение изделия",
                "model": "reclamation",
                "field": "product",
                "type": "related_str",
            },
            "reclamation.products_count": {
                "header": "Количество изделий",
                "model": "reclamation",
                "field": "products_count",
                "type": "direct",
            },
            "reclamation.product_number": {
                "header": "Номер изделия",
                "model": "reclamation",
                "field": "product_number",
                "type": "direct",
            },
            "reclamation.manufacture_date": {
                "header": "Дата изготовления",
                "model": "reclamation",
                "field": "manufacture_date",
                "type": "direct",
            },
            "reclamation.claimed_defect": {
                "header": "Заявленный дефект",
                "model": "reclamation",
                "field": "claimed_defect",
                "type": "direct",
            },
            "reclamation.consumer_act_number": {
                "header": "Номер акта рекламации",
                "model": "reclamation",
                "field": "consumer_act_number",
                "type": "direct",
            },
            "reclamation.consumer_act_date": {
                "header": "Дата акта рекламации",
                "model": "reclamation",
                "field": "consumer_act_date",
                "type": "direct",
            },
            "reclamation.engine_brand": {
                "header": "Марка двигателя",
                "model": "reclamation",
                "field": "engine_brand",
                "type": "direct",
            },
            "reclamation.engine_number": {
                "header": "Номер двигателя",
                "model": "reclamation",
                "field": "engine_number",
                "type": "direct",
            },
            "reclamation.transport_name": {
                "header": "Транспортное средство",
                "model": "reclamation",
                "field": "transport_name",
                "type": "direct",
            },
            "reclamation.mileage_operating_time": {
                "header": "Пробег/наработка",
                "model": "reclamation",
                "field": "mileage_operating_time",
                "type": "direct",
            },
            "reclamation.receipt_invoice_number": {
                "header": "Накладная прихода",
                "model": "reclamation",
                "field": "receipt_invoice_number",
                "type": "direct",
            },
            "reclamation.receipt_invoice_date": {
                "header": "Дата накладной прихода",
                "model": "reclamation",
                "field": "receipt_invoice_date",
                "type": "direct",
            },
            "reclamation.pkd_number": {
                "header": "Номер 8D (ПКД)",
                "model": "reclamation",
                "field": "pkd_number",
                "type": "direct",
            },
            # Поля Investigation
            "investigation.act_number": {
                "header": "Номер акта исследования",
                "model": "investigation",
                "field": "act_number",
                "type": "related_field",  # поля связанной модели OneToOne
            },
            "investigation.act_date": {
                "header": "Дата акта исследования",
                "model": "investigation",
                "field": "act_date",
                "type": "related_field",
            },
            "investigation.solution": {
                "header": "Заключение",
                "model": "investigation",
                "field": "solution",
                "type": "related_choice_display",
            },
            "investigation.guilty_department": {
                "header": "Виновное подразделение",
                "model": "investigation",
                "field": "guilty_department",
                "type": "related_field",
            },
            "investigation.defect_causes": {
                "header": "Причины дефекта",
                "model": "investigation",
                "field": "defect_causes",
                "type": "related_field",
            },
            "investigation.defect_causes_explanation": {
                "header": "Пояснения к причинам",
                "model": "investigation",
                "field": "defect_causes_explanation",
                "type": "related_field",
            },
            "investigation.defective_supplier": {
                "header": "Поставщик комплектующего",
                "model": "investigation",
                "field": "defective_supplier",
                "type": "related_field",
            },
            "investigation.shipment_invoice_number": {
                "header": "Накладная отгрузки",
                "model": "investigation",
                "field": "shipment_invoice_number",
                "type": "related_field",
            },
            "investigation.shipment_invoice_date": {
                "header": "Дата накладной отгрузки",
                "model": "investigation",
                "field": "shipment_invoice_date",
                "type": "related_field",
            },
            # Поля Claim
            "claims.claim_number": {
                "header": "Номер претензии",
                "model": "claims",  # по related_name (claims)
                "field": "claim_number",
                "type": "related_field_multiple",  # поля связанной модели ForeignKey
            },
            "claims.claim_date": {
                "header": "Дата претензии",
                "model": "claims",
                "field": "claim_date",
                "type": "related_field_multiple",
            },
            "claims.type_moneyl": {
                "header": "Денежная единица",
                "model": "claims",
                "field": "type_money",
                "type": "related_field_multiple",
            },
            "claims.claim_amount_all": {
                "header": "Сумма претензии",
                "model": "claims",
                "field": "claim_amount_all",
                "type": "related_field_multiple",
            },
            "claims.claim_amount_act": {
                "header": "Сумма по акту рекламации",
                "model": "claims",
                "field": "claim_amount_act",
                "type": "related_field_multiple",
            },
            "claims.costs_all": {
                "header": "Признано по претензии",
                "model": "claims",
                "field": "costs_all",
                "type": "related_field_multiple",
            },
            "claims.costs_act": {
                "header": "Признано по акту",
                "model": "claims",
                "field": "costs_act",
                "type": "related_field_multiple",
            },
            "claims.response_number": {
                "header": "Ответ на претензию",
                "model": "claims",
                "field": "response_number",
                "type": "related_field_multiple",
            },
            "claims.response_date": {
                "header": "Дата ответа на претензию",
                "model": "claims",
                "field": "response_date",
                "type": "related_field_multiple",
            },
        }

    @classmethod
    def get_available_fields(cls):
        """Получить список всех доступных полей для выбора"""
        exporter = cls()
        fields = []

        for field_key, config in exporter.field_config.items():
            fields.append(
                {
                    "key": field_key,
                    "header": config["header"],
                    "model": config["model"],
                    "group": MODEL_GROUP_MAPPING.get(config["model"], "Прочее"),
                }
            )

        return fields

    def _get_field_value(self, reclamation, field_key):
        """Получить значение поля по конфигурации"""
        config = self.field_config[field_key]

        try:
            if config["type"] == "direct":
                value = getattr(reclamation, config["field"])
                return self._format_value(value)

            elif config["type"] == "choice_display":
                return getattr(reclamation, f"get_{config['field']}_display")()

            elif config["type"] == "related_str":
                related_obj = getattr(reclamation, config["field"])
                return str(related_obj) if related_obj else ""

            elif config["type"] == "related_field":
                # Для OneToOne связей (investigation)
                try:
                    investigation = reclamation.investigation
                    value = getattr(investigation, config["field"])
                    return self._format_value(value)  # с единым форматированием
                except Reclamation.investigation.RelatedObjectDoesNotExist:
                    return ""

            elif config["type"] == "related_choice_display":
                try:
                    investigation = reclamation.investigation
                    return getattr(investigation, f"get_{config['field']}_display")()
                except Reclamation.investigation.RelatedObjectDoesNotExist:
                    return ""

            elif config["type"] == "related_field_multiple":
                # Для ForeignKey связей (claims)
                try:
                    # Выгружаем все претензии через запятую
                    claims = getattr(reclamation, config["model"]).all()
                    if claims.exists():
                        claim_values = []
                        for claim in claims:
                            field_value = getattr(claim, config["field"])
                            if field_value is not None:
                                formatted_value = self._format_value(field_value)
                                claim_values.append(formatted_value)

                        return ", ".join(claim_values) if claim_values else ""
                    return ""
                except Exception:
                    return ""

        except Exception:
            return ""

    def _format_value(self, value):
        """Единое форматирование значений для всех типов полей"""
        if value is None:
            return ""
        # elif hasattr(value, "strftime"):  # Дата
        #     return value.strftime("%d.%m.%Y")
        elif hasattr(value, "strftime"):  # Дата
            return value  # возвращаем datetime как есть, а форматируем ниже
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            # Числа (но не булевы значения)
            return str(value)
        else:
            return str(value)

    def _apply_date_formatting(self):
        """Применить форматирование дат к ячейкам Excel"""
        # Определяем какие столбцы содержат даты
        date_columns = []
        for col_num, field_key in enumerate(self.selected_fields, 1):
            if field_key in self.field_config:
                config = self.field_config[field_key]
                # Проверяем по названию поля в модели
                if (
                    config["field"].endswith("_date")
                    or "date" in config["field"].lower()
                ):
                    date_columns.append(col_num)
        print(date_columns)
        # Применяем форматирование к столбцам с датами
        for col_num in date_columns:
            column_letter = get_column_letter(col_num)
            for row in range(2, self.ws.max_row + 1):  # Пропускаем заголовки
                cell = self.ws[f"{column_letter}{row}"]
                if cell.value is not None:
                    cell.number_format = "DD.MM.YYYY"

    def _apply_cell_formatting(self, cell, is_header=False):
        """Применить форматирование к ячейке"""

        # Границы для всех ячеек
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Центрирование по центру (горизонтально и вертикально) + перенос строк
        alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        if is_header:
            # Заголовки: размер 8, не жирный
            cell.font = Font(size=8, bold=False)
            # Серый фон для заголовков
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
        else:
            # Обычные ячейки: размер 9
            cell.font = Font(size=9)
            # Фон по умолчанию (белый)

        # Применяем границы и выравнивание ко всем ячейкам
        cell.border = border
        cell.alignment = alignment

    def _write_data(self):
        """Записать данные в Excel согласно выбранным полям"""
        if not self.selected_fields:
            return

        # Записываем заголовки
        headers = []
        for field_key in self.selected_fields:
            if field_key in self.field_config:
                headers.append(self.field_config[field_key]["header"])

        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            # Применяем форматирование заголовка
            self._apply_cell_formatting(cell, is_header=True)

        # # Получаем данные с оптимизацией запросов
        # reclamations = (
        #     Reclamation.objects.all()
        #     .select_related("product_name", "product")
        #     .prefetch_related("investigation")
        # )

        # Получаем данные с фильтрацией по году
        queryset = Reclamation.objects.select_related(
            "product_name", "product"
        ).prefetch_related("investigation", "claims")

        # Применяем фильтр по году если выбран
        if self.year and str(self.year) != "all":
            queryset = queryset.filter(year=self.year)

        reclamations = queryset

        # Записываем данные
        for row, reclamation in enumerate(reclamations, 2):
            for col, field_key in enumerate(self.selected_fields, 1):
                if field_key in self.field_config:
                    value = self._get_field_value(reclamation, field_key)
                    cell = self.ws.cell(row=row, column=col, value=value)
                    # Применяем форматирование данных
                    self._apply_cell_formatting(cell, is_header=False)

    def _adjust_column_width(self):
        """Настройка ширины колонок"""
        # Словарь: название поля → ширина столбца
        custom_widths = {
            "Номер строки": 9,
            "Дата поступления сообщения": 14,
            "Период выявления дефекта": 17,
            "Наименование изделия": 16,
            "Обозначение изделия": 17,
            "Количество изделий": 10,
            "Номер изделия": 10,
            "Дата изготовления": 10,
            "Заявленный дефект": 20,
            "Номер акта рекламации": 14,
            "Марка двигателя": 16,
            "Номер двигателя": 14,
            "Транспортное средство": 16,
            "Пробег/наработка": 14,
            "Номер 8D (ПКД)": 14,
            "Номер акта исследования": 14,
            "Дата акта исследования": 14,
            "Заключение": 11,
            "Виновное подразделение": 14,
            "Причины дефекта": 22,
            "Пояснения к причинам": 26,
            "Поставщик комплектующего": 22,
            "Накладная отгрузки": 11,
            "Номер претензии": 14,
            "Ответ на претензию": 14,
        }

        # Получаем заголовки для сопоставления
        headers = []
        for field_key in self.selected_fields:
            if field_key in self.field_config:
                headers.append(self.field_config[field_key]["header"])

        # Устанавливаем ширину для каждого столбца
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)

            # Берем ширину из словаря или ставим дефолтную
            width = custom_widths.get(header, 12)  # 12 - дефолтная ширина

            self.ws.column_dimensions[column_letter].width = width

    def _apply_filter_and_freeze_row(self):
        """Добавить автофильтр к строке заголовков и закрепить заголовки"""
        if self.ws.max_row > 1:  # Проверяем, что есть данные
            # Фильтр только для первой строки (заголовки)
            header_range = f"A1:{get_column_letter(len(self.selected_fields))}1"
            self.ws.auto_filter.ref = header_range

            # Закрепляем первую строку (заголовки)
            self.ws.freeze_panes = "A2"

    def export_to_excel(self):
        """Основной метод экспорта в файл Excel"""

        if not self.selected_fields:
            raise ValueError("Не выбраны поля для экспорта")

        self._write_data()  # Записываем данные с форматированием
        self._apply_date_formatting()  # Форматируем даты в Excel
        self._adjust_column_width()  # Настраиваем ширину столбцов
        self._apply_filter_and_freeze_row()  # Добавляем фильтры и закрепляем заголовок

        # Сохраняем файл на диск
        self.wb.save(self.save_path)
        return True
