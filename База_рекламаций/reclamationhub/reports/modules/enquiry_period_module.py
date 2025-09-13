# reports/modules/enquiry_period_module.py
# Модуль приложения "Справка за период" с основной логикой формирования справки

import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import errno
import os

from reclamations.models import Reclamation
from reports.models import EnquiryPeriod
from reports.config.paths import (
    BASE_REPORTS_DIR,
    get_enquiry_period_txt_path,
    get_enquiry_period_excel_path,
)


class MetadataLoader:
    """Аналог TextDatabaseLoader - работа с Django моделью вместо JSON"""

    def __init__(self):
        # Получаем последнюю запись метаданных или создаем начальную
        self.last_metadata = EnquiryPeriod.objects.order_by("-sequence_number").first()

        if not self.last_metadata:
            # Создаем начальную запись (аналог {"0": ["0", "08-01-2025"]}) - первый рабочий день
            self.last_metadata = EnquiryPeriod.objects.create(
                sequence_number=0,
                last_processed_id=0,
                report_date=date.today(),
            )

        # Переменные
        self.sequence_number = self.last_metadata.sequence_number  # номер справки
        self.last_processed_id = self.last_metadata.last_processed_id  # последний ID
        self.last_report_date = self.last_metadata.report_date  # последняя дата


class DataProcessor(MetadataLoader):
    """Аналог MakeResultDataframe - получение данных из Django ORM вместо Excel"""

    def __init__(self):
        super().__init__()
        self.new_last_id = 0
        self.df_res = pd.DataFrame()
        self.today = date.today()

    def get_result(self):
        # Вместо pd.read_excel - запрос к Django ORM
        queryset = (
            Reclamation.objects.filter(id__gt=self.last_processed_id)
            .select_related("defect_period", "product_name", "product")
            .values(
                "id",  # поле из Reclamation для отслеживания последнего ID
                "defect_period__name",  # Период выявления дефекта - поле name из PeriodDefect
                "product_name__name",  # Наименование изделия - поле name из ProductType
                "product__nomenclature",  # Обозначение изделия - поле nomenclature из Product
                "claimed_defect",  # поле из Reclamation
                "products_count",  # поле из Reclamation
            )
        )

        if not queryset.exists():
            return None  # Нет новых данных

        # Преобразуем в DataFrame
        df = pd.DataFrame(list(queryset))

        # Устанавливаем индекс как ID записи (аналог номера строки Excel)
        df.set_index("id", inplace=True)

        # Переименовываем столбцы как в оригинале
        df.rename(
            columns={
                "defect_period__name": "Период выявления",
                "product_name__name": "Наименование изделия",
                "product__nomenclature": "Обозначение изделия",
                "claimed_defect": "Заявленный дефект",
                "products_count": "Количество",
            },
            inplace=True,
        )

        # Ваша существующая логика обработки
        df_c = df.dropna(subset=["Период выявления"])

        # Очистка обозначения изделий от переносов
        df_c["Обозначение изделия"] = df_c["Обозначение изделия"].apply(
            lambda x: x.split("\n")[0] if pd.notna(x) and "\n" in str(x) else x
        )

        # Приведение типов
        df_c["Количество"] = df_c["Количество"].astype("int16")

        # Заполнение пропусков
        df_c["Заявленный дефект"].fillna("неизвестно", inplace=True)

        # Запоминаем последний обработанный ID
        self.new_last_id = df_c.index.max()

        # Группировка как в оригинале
        self.df_res = (
            df_c.groupby(
                [
                    "Период выявления",
                    "Наименование изделия",
                    "Обозначение изделия",
                    "Заявленный дефект",
                ]
            )["Количество"]
            .sum()
            .to_frame()
        )

        return self.df_res

    def update_metadata(self):
        """Аналог write_to_database - создаем новую запись метаданных"""
        EnquiryPeriod.objects.create(
            sequence_number=self.sequence_number + 1,
            last_processed_id=self.new_last_id,
            report_date=self.today,
        )


class ExcelWriter(DataProcessor):
    """Аналог WriteResult - создание Excel файла"""

    def __init__(self):
        super().__init__()
        self.setup_file_paths()

    def setup_file_paths(self):
        """Настройка путей через общий конфиг reports/config/paths.py"""
        sequence_number = self.sequence_number + 1
        self.txt_file_path = get_enquiry_period_txt_path(sequence_number)
        self.excel_file_path = get_enquiry_period_excel_path(sequence_number)

    def write_to_txt(self):
        """Архивирование отчета в TXT"""
        with open(self.txt_file_path, "w", encoding="utf-8") as f:
            print(
                f"\n\n\tСправка по количеству рекламаций за период с {self.last_report_date.strftime('%d-%m-%Y')} по {self.today.strftime('%d-%m-%Y')}"
                f"\n\tID записей базы рекламаций: {self.last_processed_id + 1} - {self.new_last_id}",
                file=f,
            )
            f.write(self.df_res.to_string())

    def write_to_excel(self):
        """Создание отформатированного Excel файла"""
        self.df_res.to_excel(self.excel_file_path)

        wb = load_workbook(self.excel_file_path)
        sheet = wb["Sheet1"]

        # Вызываем метод форматирования
        self._apply_formatting(sheet)

        wb.save(self.excel_file_path)

    def _apply_formatting(self, sheet):
        """Метод для редактирования стилей и выравнивания в файле Excel справки"""
        # вставляем дополнительный столбец в позицию 0 (для лучшей визуализации)
        sheet.insert_cols(0)

        # задаем высоту строки 1 (с названиями столбцов)
        sheet.row_dimensions[1].height = 15

        # задаем ширину столбцов B, C, D, E, F
        sheet.column_dimensions["B"].width = 23
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 23
        sheet.column_dimensions["F"].width = 10

        # столбцы таблицы
        cols = "B", "C", "D", "E", "F"

        # определяем количество строк в таблице (длина итогового датафрейма)
        len_table = len(self.df_res)

        # циклом по столбцам таблицы
        for i in cols:
            # активируем перенос текста в ячейках B1, C1, D1, E1, F1 (с названиями столбцов) и выравниваем по центру
            sheet[f"{i + str(1)}"].alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
            # циклом по строкам таблицы
            for j in range(1, len_table + 2):
                # задаем стиль границы - тонкая линия и цвет черный
                thins = Side(border_style="thin", color="000000")
                # применяем заданный стиль границы к верхней, нижней, левой и правой границе ячеек по циклу
                sheet[f"{i + str(j)}"].border = Border(
                    top=thins, bottom=thins, left=thins, right=thins
                )
                # изменяем шрифт в ячейках с жирного на обычный и устанавливаем Times New Roman размером 10
                sheet[f"{i + str(j)}"].font = Font(
                    name="Times New Roman", size=10, bold=False
                )

        for i in ("B", "C", "D", "E"):
            for j in range(2, len_table + 2):
                # выравниваем текст в ячейках "B", "C", "D", "E" по левому краю по верху с переносом текста
                sheet[f"{i + str(j)}"].alignment = Alignment(
                    wrap_text=True, horizontal="left", vertical="top"
                )
        for j in range(2, len_table + 2):
            # выравниваем текст в ячейке "F" по центру
            sheet[f"F{str(j)}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Объединяем ячейки после таблицы для внесения текста
        sheet.merge_cells(f"B{len_table + 3}:F{len_table + 3}")
        # Записываем текст в объединенную ячейку
        sheet[f"B{len_table + 3}"] = (
            f"Справка по количеству рекламаций за период с {self.last_report_date.strftime('%d-%m-%Y')} по {self.today.strftime('%d-%m-%Y')}\n"
            f"ID записей базы рекламаций: {self.last_processed_id + 1} - {self.new_last_id}"
        )
        # Устанавливаем выравнивание по левому краю с переносом текста
        sheet[f"B{len_table + 3}"].alignment = Alignment(
            wrap_text=True, horizontal="left", vertical="center"
        )
        # Изменяем шрифт в ячейке на Times New Roman размером 12
        sheet[f"B{len_table + 3}"].font = Font(name="Times New Roman", size=12)
        # Задаем высоту строки
        sheet.row_dimensions[len_table + 3].height = 30

    def generate_full_report(self):
        """Полная генерация справки с обработкой ошибок"""
        try:
            result = self.get_result()

            if result is None:
                return {
                    "success": False,
                    "message": "Нет новых данных для формирования справки",
                    "message_type": "info",
                }

            # Генерируем файлы
            self.write_to_txt()  # Сохраняем справку TXT
            self.write_to_excel()  # Сохраняем справку Excel
            self.update_metadata()  # Обновляем в БД актуальные значения (ID строки и сегодняшюю дату)

            # Формируем результат
            return {
                "success": True,
                "message": f"Справка сформирована",
                "full_message": f"ID записей: {self.last_processed_id + 1} - {self.new_last_id}. Файл находится в папке {BASE_REPORTS_DIR}",
                "excel_path": self.excel_file_path,
                "filename": os.path.basename(self.excel_file_path),
                "message_type": "success",
            }

        except OSError as e:
            # Обработка ошибок файловой системы (включая PermissionError)
            if e.errno == errno.EACCES or "Permission denied" in str(e):
                return {
                    "success": False,
                    "message": "🔒 Возможно у вас открыт файл Excel со справкой. Закройте файл Excel и попробуйте снова.",
                    "message_type": "warning",
                }
            else:
                return {
                    "success": False,
                    "message": f"Ошибка файловой системы: {str(e)}",
                    "message_type": "error",
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"Неожиданная ошибка при формировании отчета: {str(e)}",
                "message_type": "error",
            }


if __name__ == "__main__":

    obj = ExcelWriter()

    # считываем значения из базы данных приложения
    print(f"Дата последней записи: {obj.last_report_date}")
    print(f"Номер последней строки базы рекламаций ОТК: {obj.last_processed_id}")

    # получаем итоговый датафрейм
    result = obj.get_result()
    print(f"Последняя строка актуальной базы рекламаций ОТК: {obj.new_last_id}")

    # записываем в словарь актуальные значения номера строки и сегодняшюю дату
    obj.update_metadata()

    # записываем в файл TXT
    obj.write_to_txt()
    print("Справка в файл TXT записана")

    # записываем в файл Excel
    obj.write_to_excel()
    print("Отредактированный файл Excel со справкой записан")
