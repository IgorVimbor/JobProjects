# reports/modules/culprits_defect_module.py
"""Модуль приложения "Дефекты по виновникам" с основной логикой.

Включает класс:
- `CulpritsDefectProcessor` - Обработка данных для анализа дефектов по виновникам
"""

import os
import json
from django.utils.safestring import mark_safe
import pandas as pd
from datetime import date
from dateutil.relativedelta import relativedelta
import errno
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from investigations.models import Investigation
from reports.config.paths import (
    BASE_REPORTS_DIR,
    culprits_defect_json_db,
    get_culprits_defect_excel_path,
)


class CulpritsDefectProcessor:
    """Обработка данных для анализа дефектов по виновникам"""

    # Названия месяцев
    MONTH_NAMES = {
        1: "январь",
        2: "февраль",
        3: "март",
        4: "апрель",
        5: "май",
        6: "июнь",
        7: "июль",
        8: "август",
        9: "сентябрь",
        10: "октябрь",
        11: "ноябрь",
        12: "декабрь",
    }

    def __init__(self, user_number=None):
        """
        Инициализация процессора.

        user_number: Номер акта, введённый пользователем.
        Если None - используется автоматическое значение из JSON.
        """
        self.today = date.today()
        self.bza_df = pd.DataFrame()
        self.not_bza_df = pd.DataFrame()

        self.dct_act_numbers = {}
        self.max_act_number = None
        self.start_act_number = None

        # Определяем год анализа по прошлому месяцу
        self.prev_month = self.today - relativedelta(months=1)
        self.analysis_year = self.prev_month.year
        self.month_name = self.MONTH_NAMES[self.prev_month.month]

        # Загружаем данные из JSON
        self._load_act_numbers_from_json()

        # Определяем начальный номер акта для фильтрации
        if user_number is not None:
            # Пользователь указал свой номер - используем его
            self.start_act_number = user_number + 1

    def _load_act_numbers_from_json(self):
        """Загрузка номеров актов из JSON файла"""
        try:
            with open(culprits_defect_json_db, encoding="utf-8-sig") as file:
                raw_data = json.load(file)

            # Преобразуем ключи в int (JSON хранит ключи как строки)
            self.dct_act_numbers = {int(k): v for k, v in raw_data.items()}

            # По последнему (максимальному) ключу находим стартовый и максимальный акты
            max_key = max(self.dct_act_numbers)
            self.start_act_number, self.max_act_number = self.dct_act_numbers[max_key]

        except (FileNotFoundError, json.JSONDecodeError):
            # Если файла нет или он повреждён - создаём новый
            self.dct_act_numbers = {1: [1, 1]}
            self.start_act_number = 1
            self.max_act_number = 1
            self._save_act_numbers_to_json(self.dct_act_numbers)

    def _save_act_numbers_to_json(self, dct):
        """Сохранение номеров актов в JSON файл"""
        try:
            # Создаём директорию, если не существует
            os.makedirs(os.path.dirname(culprits_defect_json_db), exist_ok=True)

            with open(culprits_defect_json_db, "w", encoding="utf-8-sig") as file:
                json.dump(dct, file, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Ошибка сохранения JSON: {e}")

    def get_default_act_number(self):
        """
        Получение месяца справки и номеров актов по умолчанию для отображения в форме.
        Возвращает название месяца последней справки, номер стартового и максимального акта из этой справки
        """
        # Находим максимальный (последний) ключ в словаре
        max_key = max(self.dct_act_numbers)

        # Убеждаемся, что ключ — число (защита на случай если ключ строка)
        if isinstance(max_key, str):
            max_key = int(max_key)

        # Определяем название месяца последней справки
        json_month_name = self.MONTH_NAMES[max_key]

        return json_month_name, self.start_act_number, self.max_act_number

    def process_data(self):
        """Основная логика обработки данных с pandas"""

        # =========== Вспомогательные функции ===========

        def get_numeric_part(act_str):
            """Извлекает числовую часть акта для сравнения: '1067-1' -> 1067"""
            try:
                return int(str(act_str).split("-")[0])
            except (ValueError, IndexError, AttributeError):
                return 0

        def join_unique(series, separator=", "):
            """Объединяет уникальные значения в строку"""
            return separator.join(str(i) for i in series.dropna().unique())

        try:
            # =========== Получение данных из модели Investigation ============

            investigations_queryset = Investigation.objects.select_related(
                "reclamation__defect_period", "reclamation__product_name"
            ).values(
                "act_number",
                "act_date",
                "reclamation__defect_period__name",
                "reclamation__product_name__name",
                "reclamation__product_number",
                "reclamation__manufacture_date",
                "reclamation__products_count",
                "solution",
                "fault_type",
                "guilty_department",
                "defect_causes",
                "defect_causes_explanation",
            )

            if not investigations_queryset.exists():
                return False, "Нет данных в таблице исследований"

            # =========== Создание датафрейма ============

            df = pd.DataFrame(list(investigations_queryset))

            df.rename(
                columns={
                    "act_number": "Номер акта исследования",
                    "act_date": "Дата акта исследования",
                    "reclamation__defect_period__name": "Период выявления дефекта",
                    "reclamation__product_name__name": "Обозначение изделия",
                    "reclamation__product_number": "Заводской номер изделия",
                    "reclamation__manufacture_date": "Дата изготовления изделия",
                    "reclamation__products_count": "Количество предъявленных изделий",
                    "solution": "Решение",
                    "fault_type": "Виновник дефекта",
                    "guilty_department": "Виновное подразделение",
                    "defect_causes": "Причины дефектов",
                    "defect_causes_explanation": "Пояснения к причинам дефектов",
                },
                inplace=True,
            )

            # =========== Обработка отсутствующих значений ============

            # Удаляем строки с отсутствующим номером акта исследования
            df.dropna(subset=["Номер акта исследования"], inplace=True)

            if df.empty:
                return False, "Нет записей с номерами актов исследования"

            # Удаляем акты со значением "без исследования"
            df = df[df["Номер акта исследования"] != "без исследования"]

            if df.empty:
                return False, "Нет записей после исключения актов 'без исследования'"

            # =========== Фильтрация датафрейма ============

            # 1. Фильтруем по отчетному месяцу
            df["Дата акта исследования"] = pd.to_datetime(df["Дата акта исследования"])
            prev_month_ts = pd.Timestamp(self.prev_month)

            # Преобразуем даты в формат год-месяц и оставляем только отчетный месяц
            df_filtered = df[  # Создаем новый датафрейм
                df["Дата акта исследования"].dt.to_period("M")
                == prev_month_ts.to_period("M")
            ].copy()  # Добавляем .copy() для избежания SettingWithCopyWarning

            if df_filtered.empty:
                return (
                    False,
                    f"Нет данных за {self.month_name} {self.analysis_year} года",
                )

            # 2. Фильтруем по году анализа
            # Извлекаем год и номер акта исследования ("2025 № 1067" → год=2025, номер=1067)
            df_filtered["Год акта"] = (
                df_filtered["Номер акта исследования"]
                .str.split(" № ")
                .str[0]
                .astype(int)
            )

            df_filtered = df_filtered[df_filtered["Год акта"] == self.analysis_year]

            if df_filtered.empty:
                return False, f"Нет данных за {self.analysis_year} год"

            # 3. Фильтруем по номеру акта
            # Оставляем как строку для корректной работы с "1067-1"
            df_filtered["Номер акта (короткий)"] = (
                df_filtered["Номер акта исследования"].str.split(" № ").str[1]
            )

            # Создаем числовой столбец для сравнения
            df_filtered["act_number_int"] = df_filtered["Номер акта (короткий)"].apply(
                get_numeric_part
            )

            # Оставляем акты с номером >= start_act_number
            df_filtered = df_filtered[
                df_filtered["act_number_int"] >= self.start_act_number
            ]

            if df_filtered.empty:  # Если датафрейм пустой
                if self.prev_month.month == len(self.dct_act_numbers):
                    return (  # Если справка за прошлый месяц уже формировалась - выводим сообщение
                        False,
                        mark_safe(
                            f"Справка за {self.month_name} {self.analysis_year} года <strong>уже формировалась</strong>! "
                            f"Смотри в папке {BASE_REPORTS_DIR}."
                        ),
                    )
                else:  # Если справка не делалась, но датафрейм пустой - выводим сообщение
                    return (
                        False,
                        f"Нет данных за {self.analysis_year} год начиная с акта № {self.start_act_number}",
                    )

            # Находим минимальный и максимальный номера актов (числовые)
            self.start_act_number = int(df_filtered["act_number_int"].min())
            self.max_act_number = int(df_filtered["act_number_int"].max())

            # Сохраняем с числовым ключом в JSON
            self.dct_act_numbers[self.prev_month.month] = [
                self.start_act_number,
                self.max_act_number,
            ]
            self._save_act_numbers_to_json(self.dct_act_numbers)

            # 4. Фильтруем по признано/отклонено
            df_filtered = df_filtered[df_filtered["Решение"] == "ACCEPT"]

            if df_filtered.empty:
                return False, "Нет признанных рекламаций в данных"

            # Убираем служебные столбцы
            df_filtered = df_filtered.drop(columns=["Год акта", "act_number_int"])

            # =========== Группировка и агрегация ============

            df_grouped = df_filtered.groupby(
                [
                    "Виновное подразделение",
                    "Период выявления дефекта",
                    "Обозначение изделия",
                ]
            ).agg(
                {
                    "Заводской номер изделия": join_unique,
                    "Дата изготовления изделия": join_unique,
                    "Количество предъявленных изделий": "sum",
                    "Номер акта (короткий)": join_unique,
                    "Причины дефектов": join_unique,
                    "Пояснения к причинам дефектов": join_unique,
                }
            )

            # Разделение на БЗА ("Не определено") и не-БЗА (виновник определен)
            self.bza_df = df_grouped.loc[
                df_grouped.index.get_level_values("Виновное подразделение")
                == "Не определено"
            ]
            self.not_bza_df = df_grouped.loc[
                df_grouped.index.get_level_values("Виновное подразделение")
                != "Не определено"
            ]

            return True, f"Обработано записей: {len(df_filtered)}"

        except Exception as e:
            return False, f"Ошибка обработки данных: {str(e)}"

    def _prepare_table_data(self, df, include_culprit=False):
        """Подготовка данных для отображения в таблице"""
        if df.empty:
            return []

        result = []
        for index, row in df.iterrows():
            data = {
                "Потребитель": index[1],  # Период выявления дефекта
                "Изделие": index[2],  # Обозначение изделия
                "Заводской_номер": row["Заводской номер изделия"],
                "Дата_изготовления": row["Дата изготовления изделия"],
                "Количество": int(row["Количество предъявленных изделий"]),
                "Номера_актов": row["Номер акта (короткий)"],
                "Причины": row["Причины дефектов"],
                "Пояснения": row["Пояснения к причинам дефектов"],
            }

            # Для таблицы "не-БЗА" - добавляем виновника
            if include_culprit:
                data["Виновник"] = index[0]  # Виновное подразделение

            result.append(data)

        return result

    def generate_analysis(self):
        """Основной метод генерации анализа"""
        try:
            # Обрабатываем данные
            success, message = self.process_data()

            if not success:
                return {"success": False, "message": message, "message_type": "warning"}

            # Подготавливаем данные для таблиц
            bza_data = self._prepare_table_data(self.bza_df, include_culprit=False)
            not_bza_data = self._prepare_table_data(
                self.not_bza_df, include_culprit=True
            )

            return {
                "success": True,
                "message": f"Справка по виновникам дефектов за {self.month_name} {self.analysis_year} года",
                "bza_data": bza_data,
                "not_bza_data": not_bza_data,
                "bza_count": len(bza_data),
                "not_bza_count": len(not_bza_data),
                "start_act_number": self.start_act_number,
                "max_act_number": self.max_act_number,
                "analysis_year": self.analysis_year,
                "message_type": "success",
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"Ошибка при выполнении анализа: {str(e)}",
                "message_type": "warning",
            }

    def save_to_excel_from_data(
        self, bza_data, not_bza_data, start_act_number, max_act_number=None
    ):
        """Сохранение готовых данных в Excel"""
        try:
            if not bza_data and not not_bza_data:
                return {
                    "success": False,
                    "message": "Нет данных для сохранения в файл",
                    "message_type": "warning",
                }

            # Формируем путь к файлу
            excel_path = get_culprits_defect_excel_path()

            # Создаем Excel файл из готовых данных
            self._create_excel_from_data(
                excel_path, bza_data, not_bza_data, start_act_number, max_act_number
            )

            return {
                "success": True,
                "message": f"✅ Справка сохранена в файл Excel",
                "full_message": f"Справка по виновникам дефектов сохранена в папку {BASE_REPORTS_DIR}",
                "excel_path": excel_path,
                "filename": os.path.basename(excel_path),
                "message_type": "success",
            }

        except OSError as e:
            if e.errno == errno.EACCES or "Permission denied" in str(e):
                return {
                    "success": False,
                    "message": "🔒 Возможно у вас открыт файл Excel со справкой. Закройте файл Excel и попробуйте снова.",
                    "message_type": "warning",
                }
            else:
                return {
                    "success": False,
                    "message": f"Ошибка файловой системы: {str(e)}",
                    "message_type": "error",
                }

        except Exception as e:
            return {
                "success": False,
                "message": f"Неожиданная ошибка при сохранении файла: {str(e)}",
                "message_type": "error",
            }

    def _create_excel_from_data(
        self, excel_path, bza_data, not_bza_data, start_act_number, max_act_number=None
    ):
        """Создание Excel файла из готовых данных на одном листе"""

        # Создаем новую книгу
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.title = "Справка по виновникам дефектов"

        current_row = 1

        # 1. Первая строка - информация о номерах актов исследования (старт-стоп)
        ws[f"A{current_row}"] = (
            f"Справка составлена по актам исследования с {self.start_act_number} по {max_act_number}"
        )

        current_row += 2  # +2 для пустой строки

        # Запоминаем строки для форматирования
        first_table_start = None
        first_table_header_row = None
        first_table_data_rows = 0

        second_table_start = None
        second_table_header_row = None
        second_table_data_rows = 0

        # 2. Первая таблица "Дефекты по виновникам" - начинаем с колонки A
        if not_bza_data:
            # Заголовок таблицы
            ws[f"A{current_row}"] = "Дефекты по виновникам"
            first_table_start = current_row
            current_row += 1

            # Заголовки столбцов с "Виновник" первым (A-I)
            headers_1 = [
                "Виновник",
                "Потребитель",
                "Изделие",
                "Заводской_номер",
                "Дата_изготовления",
                "Количество",
                "Номера_актов",
                "Причины",
                "Пояснения",
            ]

            first_table_header_row = current_row
            # Записываем заголовки начиная с колонки A
            for col, header in enumerate(headers_1, 1):  # 1=A, 2=B, 3=C...
                ws.cell(row=current_row, column=col, value=header)

            current_row += 1

            # Данные таблицы
            for row_data in not_bza_data:
                ws.cell(row=current_row, column=1, value=row_data["Виновник"])
                ws.cell(row=current_row, column=2, value=row_data["Потребитель"])
                ws.cell(row=current_row, column=3, value=row_data["Изделие"])
                ws.cell(row=current_row, column=4, value=row_data["Заводской_номер"])
                ws.cell(row=current_row, column=5, value=row_data["Дата_изготовления"])
                ws.cell(row=current_row, column=6, value=row_data["Количество"])
                ws.cell(row=current_row, column=7, value=row_data["Номера_актов"])
                ws.cell(row=current_row, column=8, value=row_data["Причины"])
                ws.cell(row=current_row, column=9, value=row_data["Пояснения"])
                current_row += 1
                first_table_data_rows += 1

            current_row += 1  # Пустая строка после первой таблицы

        # 3. Вторая таблица "БЗА" - начинаем с колонки B
        if bza_data:
            # Заголовок таблицы
            ws[f"A{current_row}"] = "Дефекты без центра ответственности"
            second_table_start = current_row
            current_row += 1

            # Заголовки столбцов БЕЗ "Виновник" (B-I)
            headers_2 = headers_1[1:]

            second_table_header_row = current_row
            # Записываем заголовки начиная с колонки B
            for col, header in enumerate(headers_2, 2):  # 2=B, 3=C, 4=D...
                ws.cell(row=current_row, column=col, value=header)

            current_row += 1

            # Данные таблицы
            for row_data in bza_data:
                # Пропускаем колонку A (Виновник), начинаем с B
                ws.cell(row=current_row, column=2, value=row_data["Потребитель"])
                ws.cell(row=current_row, column=3, value=row_data["Изделие"])
                ws.cell(row=current_row, column=4, value=row_data["Заводской_номер"])
                ws.cell(row=current_row, column=5, value=row_data["Дата_изготовления"])
                ws.cell(row=current_row, column=6, value=row_data["Количество"])
                ws.cell(row=current_row, column=7, value=row_data["Номера_актов"])
                ws.cell(row=current_row, column=8, value=row_data["Причины"])
                ws.cell(row=current_row, column=9, value=row_data["Пояснения"])
                current_row += 1
                second_table_data_rows += 1

        # 4. Применяем форматирование сразу
        self._apply_formatting_to_worksheet(
            ws,
            first_table_start,
            first_table_header_row,
            first_table_data_rows,
            second_table_start,
            second_table_header_row,
            second_table_data_rows,
        )

        # Сохраняем файл
        wb.save(excel_path)

    def _apply_formatting_to_worksheet(
        self,
        ws,
        first_table_start,
        first_table_header_row,
        first_table_data_rows,
        second_table_start,
        second_table_header_row,
        second_table_data_rows,
    ):
        """Форматирование рабочего листа"""

        # Стили
        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        thin_border = Side(border_style="thin", color="000000")
        font_normal = Font(name="Times New Roman", size=8, bold=False)
        font_header = Font(name="Times New Roman", size=8, bold=False)
        font_title = Font(name="Times New Roman", size=9, bold=True)
        # font_info = Font(name="Times New Roman", size=10, bold=True)

        # Настройка ширины столбцов A-I
        column_widths = [10, 15, 14, 20, 20, 9, 21, 30, 30]
        for i, width in enumerate(column_widths, 1):
            col_letter = chr(65 + i - 1)  # A, B, C, D, E, F, G, H, I
            ws.column_dimensions[col_letter].width = width

        # # ✅ 1. Форматируем информационную строку (строка 1)
        # info_cell = ws["A1"]
        # info_cell.font = font_info

        # 2. Форматируем первую таблицу
        if first_table_start:
            # Заголовок таблицы
            title_cell = ws[f"A{first_table_start}"]
            title_cell.font = font_title

            # Заголовки столбцов (A-I)
            if first_table_header_row:
                for col in range(1, 10):  # A-I
                    cell = ws.cell(row=first_table_header_row, column=col)
                    cell.fill = header_fill
                    cell.font = font_header
                    cell.border = Border(
                        top=thin_border,
                        bottom=thin_border,
                        left=thin_border,
                        right=thin_border,
                    )
                    cell.alignment = Alignment(
                        wrap_text=True, horizontal="center", vertical="center"
                    )

                ws.row_dimensions[first_table_header_row].height = 20

                # Данные первой таблицы
                for row in range(
                    first_table_header_row + 1,
                    first_table_header_row + 1 + first_table_data_rows,
                ):
                    for col in range(1, 10):  # A-I
                        cell = ws.cell(row=row, column=col)
                        cell.font = font_normal
                        cell.border = Border(
                            top=thin_border,
                            bottom=thin_border,
                            left=thin_border,
                            right=thin_border,
                        )

                        # Выравнивание: количество (F) по центру, остальное по левому краю
                        if col == 6:  # Колонка "Количество" (F)
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        else:
                            cell.alignment = Alignment(
                                wrap_text=True, horizontal="left", vertical="top"
                            )

        # 3. Форматируем вторую таблицу
        if second_table_start:
            # Заголовок таблицы
            title_cell = ws[f"A{second_table_start}"]
            title_cell.font = font_title

            # Заголовки столбцов (B-I)
            if second_table_header_row:
                for col in range(2, 10):  # B-I
                    cell = ws.cell(row=second_table_header_row, column=col)
                    cell.fill = header_fill
                    cell.font = font_header
                    cell.border = Border(
                        top=thin_border,
                        bottom=thin_border,
                        left=thin_border,
                        right=thin_border,
                    )
                    cell.alignment = Alignment(
                        wrap_text=True, horizontal="center", vertical="center"
                    )

                ws.row_dimensions[second_table_header_row].height = 20

                # Данные второй таблицы
                for row in range(
                    second_table_header_row + 1,
                    second_table_header_row + 1 + second_table_data_rows,
                ):
                    for col in range(2, 10):  # B-I
                        cell = ws.cell(row=row, column=col)
                        cell.font = font_normal
                        cell.border = Border(
                            top=thin_border,
                            bottom=thin_border,
                            left=thin_border,
                            right=thin_border,
                        )

                        # Выравнивание: количество (F) по центру, остальное по левому краю
                        if col == 6:  # Колонка "Количество" (F)
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        else:
                            cell.alignment = Alignment(
                                wrap_text=True, horizontal="left", vertical="top"
                            )
